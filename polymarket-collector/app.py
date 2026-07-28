import asyncio
import html
import json
import os
import shutil
import sqlite3
import threading
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlencode
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import httpx
from fastapi import BackgroundTasks, Body, FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from openpyxl import Workbook

try:
    import truststore

    truststore.inject_into_ssl()
except ImportError:
    truststore = None

APP_DIR = Path(__file__).parent
DB_PATH = APP_DIR / "poly_data.sqlite3"
EXPORT_DIR = APP_DIR / "output"
EXPORT_FILE_PREFIX = "polymarket_data_"
EXPORT_FILE_SUFFIX = ".xlsx"
EXPORT_FETCH_CHUNK_SIZE = 5000
EXPORT_SNAPSHOT_PREFIX = ".export_snapshot_"
SQLITE_TIMEOUT_SECONDS = 10
SQLITE_BUSY_TIMEOUT_MS = SQLITE_TIMEOUT_SECONDS * 1000


def env_int(name: str, default: int) -> int:
    raw_value = os.getenv(name, str(default)).strip()
    try:
        return int(raw_value)
    except ValueError:
        return default


GAMMA_URL = "https://gamma-api.polymarket.com/events/slug/{slug}"
CLOB_BOOK_URL = "https://clob.polymarket.com/book?token_id={token_id}"
COINBASE_CANDLES_URL = os.getenv(
    "COINBASE_CANDLES_URL",
    "https://api.exchange.coinbase.com/products/{product_id}/candles",
).strip()
COINBASE_PRODUCT_ID = os.getenv("COINBASE_PRODUCT_ID", "BTC-USD").strip() or "BTC-USD"
COINBASE_CANDLE_GRANULARITY_SECONDS = env_int("COINBASE_CANDLE_GRANULARITY_SECONDS", 300)
COINBASE_VOLUME_POLL_INTERVAL_SECONDS = env_int("COINBASE_VOLUME_POLL_INTERVAL_SECONDS", 30)
COINBASE_REQUEST_TIMEOUT_SECONDS = env_int("COINBASE_REQUEST_TIMEOUT_SECONDS", 10)
COINBASE_MAX_DELTA_GAP_SECONDS = env_int("COINBASE_MAX_DELTA_GAP_SECONDS", 90)
COINBASE_MISSING_CANDLE_RETRY_COUNT = env_int("COINBASE_MISSING_CANDLE_RETRY_COUNT", 2)
COINBASE_MISSING_CANDLE_RETRY_DELAY_SECONDS = env_int("COINBASE_MISSING_CANDLE_RETRY_DELAY_SECONDS", 2)

try:
    LOCAL_TIMEZONE = ZoneInfo("Asia/Jerusalem")
except ZoneInfoNotFoundError:
    LOCAL_TIMEZONE = timezone(timedelta(hours=3), name="Asia/Jerusalem")

EVENT_CHECK_INTERVAL_SECONDS = 5
BOOK_CHECK_INTERVAL_SECONDS = 2
DEMO_INVESTMENT_USD = Decimal("1.00")
DEMO_FEE_CALCULATION_VERSION = "polymarket-platform-fee-v2-2026-07-21"
DEMO_FEE_SOURCE_MARKET = "MARKET_SNAPSHOT"
DEMO_FEE_SOURCE_FALLBACK = "SIMULATED_CRYPTO_DEFAULT"
DEMO_FEE_SOURCE_BACKFILL = "SIMULATED_BACKFILL"
CLOB_FEE_SOURCE = "CLOB_MARKET_INFO"
DEMO_ENTRY_LIQUIDITY_ROLE = "TAKER"
DEMO_EXIT_LIQUIDITY_ROLE_BY_REASON = {
    "stop_loss": "TAKER",
    "take_profit": "TAKER",
    "event_resolution": "TAKER",
}
POLYMARKET_CRYPTO_TAKER_FEE_RATE = Decimal("0.07")
MONEY_QUANT = Decimal("0.00000001")
FEE_QUANT = Decimal("0.00001")
DEFAULT_SCHEDULE_TIMEZONE = "Asia/Jerusalem"
WEEKDAY_LABELS_HE = {
    0: "שני",
    1: "שלישי",
    2: "רביעי",
    3: "חמישי",
    4: "שישי",
    5: "שבת",
    6: "ראשון",
}

active_market: Optional[dict[str, Any]] = None
active_market_lock = asyncio.Lock()
coinbase_volume_state: dict[str, Optional[str]] = {
    "last_sample_at": None,
    "last_success_at": None,
    "status": "starting",
    "last_error": None,
}
export_state_lock = threading.Lock()
export_state: dict[str, Optional[str]] = {
    "status": "idle",
    "started_at": None,
    "finished_at": None,
    "filename": None,
    "path": None,
    "error": None,
    "row_counts": None,
}

EXPORT_SHEETS: list[tuple[str, list[str], str]] = [
    (
        "events",
        [
            "local_event_id",
            "polymarket_event_id",
            "polymarket_market_id",
            "condition_id",
            "event_slug",
            "market_slug",
            "title",
            "question",
            "event_url",
            "start_time",
            "start_time_local",
            "end_time",
            "end_time_local",
            "yes_token_id",
            "no_token_id",
            "outcomes",
            "outcome_prices",
            "active",
            "closed",
            "enable_order_book",
            "accepting_orders",
            "fees_enabled",
            "fee_rate",
            "fee_calculation_source",
            "fee_calculation_version",
            "created_at_poly",
            "created_at_poly_local",
            "discovered_at",
            "discovered_at_local",
            "last_seen_at",
            "last_seen_at_local",
            "status",
            "notes",
        ],
        """
        SELECT
            local_event_id,
            polymarket_event_id,
            polymarket_market_id,
            condition_id,
            event_slug,
            market_slug,
            title,
            question,
            event_url,
            start_time,
            start_time_local,
            end_time,
            end_time_local,
            yes_token_id,
            no_token_id,
            outcomes,
            outcome_prices,
            active,
            closed,
            enable_order_book,
            accepting_orders,
            fees_enabled,
            fee_rate,
            fee_calculation_source,
            fee_calculation_version,
            created_at_poly,
            created_at_poly_local,
            discovered_at,
            discovered_at_local,
            last_seen_at,
            last_seen_at_local,
            status,
            notes
        FROM events
        ORDER BY local_event_id ASC
        """,
    ),
    (
        "orderbook_log",
        [
            "sampled_at",
            "sampled_at_local",
            "event_slug",
            "condition_id",
            "up_token_id",
            "down_token_id",
            "up_best_ask",
            "up_best_bid",
            "down_best_ask",
            "down_best_bid",
            "up_last_trade_price",
            "down_last_trade_price",
            "up_spread",
            "down_spread",
            "up_midpoint",
            "down_midpoint",
            "raw_up_timestamp",
            "raw_down_timestamp",
            "up_volume_shares_10s",
            "down_volume_shares_10s",
            "up_volume_usdc_10s",
            "down_volume_usdc_10s",
            "trades_count_10s",
            "trades_window_start",
            "trades_window_start_local",
            "trades_window_end",
            "trades_window_end_local",
            "trades_error",
            "status",
            "error",
        ],
        """
        SELECT
            sampled_at,
            sampled_at_local,
            event_slug,
            condition_id,
            up_token_id,
            down_token_id,
            up_best_ask,
            up_best_bid,
            down_best_ask,
            down_best_bid,
            up_last_trade_price,
            down_last_trade_price,
            up_spread,
            down_spread,
            up_midpoint,
            down_midpoint,
            raw_up_timestamp,
            raw_down_timestamp,
            up_volume_shares_10s,
            down_volume_shares_10s,
            up_volume_usdc_10s,
            down_volume_usdc_10s,
            trades_count_10s,
            trades_window_start,
            trades_window_start_local,
            trades_window_end,
            trades_window_end_local,
            trades_error,
            status,
            error
        FROM orderbook_log
        ORDER BY id ASC
        """,
    ),
    (
        "btc_volume_log",
        [
            "id",
            "sampled_at",
            "sample_bucket_at",
            "candle_start_at",
            "product_id",
            "granularity_seconds",
            "volume_btc_cumulative",
            "volume_btc_delta",
            "seconds_since_previous_sample",
            "event_slug",
            "condition_id",
            "source",
            "status",
            "error",
        ],
        """
        SELECT
            id,
            sampled_at,
            sample_bucket_at,
            candle_start_at,
            product_id,
            granularity_seconds,
            volume_btc_cumulative,
            volume_btc_delta,
            seconds_since_previous_sample,
            event_slug,
            condition_id,
            source,
            status,
            error
        FROM btc_volume_log
        ORDER BY sampled_at ASC
        """,
    ),
    (
        "rules",
        [
            "id",
            "name",
            "created_at",
            "updated_at",
            "entry_price",
            "stop_loss_price",
            "take_profit_price",
            "max_yes_entries_per_event",
            "max_no_entries_per_event",
            "status",
            "eligible_after_event_id",
            "entry_window_start_seconds_before_end",
            "entry_window_end_seconds_before_end",
            "schedule_timezone",
        ],
        """
        SELECT
            id,
            name,
            created_at,
            updated_at,
            entry_price,
            stop_loss_price,
            take_profit_price,
            max_yes_entries_per_event,
            max_no_entries_per_event,
            status,
            eligible_after_event_id,
            entry_window_start_seconds_before_end,
            entry_window_end_seconds_before_end,
            schedule_timezone
        FROM rules
        ORDER BY id ASC
        """,
    ),
    (
        "deals",
        [
            "id",
            "rule_id",
            "rule_name",
            "event_id",
            "side",
            "result",
            "entry_at",
            "entry_price",
            "entry_orderbook_log_id",
            "exit_at",
            "exit_price",
            "exit_orderbook_log_id",
            "exit_reason",
            "market_result",
            "price_change_points",
            "return_percent",
            "investment_usd",
            "shares",
            "entry_gross_value_usd",
            "entry_liquidity_role",
            "entry_fee_rate",
            "entry_fee_usd",
            "exit_gross_value_usd",
            "exit_liquidity_role",
            "exit_fee_rate",
            "exit_fee_usd",
            "total_fees_usd",
            "gross_pnl_usd",
            "net_pnl_usd",
            "gross_roi_percent",
            "net_roi_percent",
            "entry_btc_volume_log_id",
            "entry_btc_volume_sampled_at",
            "entry_btc_volume_btc_cumulative",
            "entry_btc_volume_btc_delta",
            "entry_btc_volume_status",
            "fee_calculation_source",
            "fee_calculation_version",
            "entry_seconds_before_event_end",
            "created_at",
            "updated_at",
        ],
        """
        SELECT
            id,
            rule_id,
            rule_name,
            event_id,
            side,
            result,
            entry_at,
            entry_price,
            entry_orderbook_log_id,
            exit_at,
            exit_price,
            exit_orderbook_log_id,
            exit_reason,
            market_result,
            price_change_points,
            return_percent,
            investment_usd,
            shares,
            entry_gross_value_usd,
            entry_liquidity_role,
            entry_fee_rate,
            entry_fee_usd,
            exit_gross_value_usd,
            exit_liquidity_role,
            exit_fee_rate,
            exit_fee_usd,
            total_fees_usd,
            gross_pnl_usd,
            net_pnl_usd,
            gross_roi_percent,
            net_roi_percent,
            entry_btc_volume_log_id,
            entry_btc_volume_sampled_at,
            entry_btc_volume_btc_cumulative,
            entry_btc_volume_btc_delta,
            entry_btc_volume_status,
            fee_calculation_source,
            fee_calculation_version,
            entry_seconds_before_event_end,
            created_at,
            updated_at
        FROM deals
        ORDER BY id ASC
        """,
    ),
    (
        "rule_inactive_windows",
        [
            "id",
            "rule_id",
            "day_of_week",
            "start_time",
            "end_time",
            "status",
            "created_at",
            "updated_at",
        ],
        """
        SELECT
            id,
            rule_id,
            day_of_week,
            start_time,
            end_time,
            status,
            created_at,
            updated_at
        FROM rule_inactive_windows
        ORDER BY rule_id ASC, day_of_week ASC, start_time ASC, id ASC
        """,
    ),
]

app = FastAPI(title="Polymarket BTC Collector")

if os.getenv("ENABLE_LEGACY_LIVE_IN_DEMO", "false").strip().lower() in {"1", "true", "yes", "on"}:
    from live.router import configure as configure_live, router as live_router

    app.include_router(live_router)
else:
    configure_live = None


class ClosingConnection(sqlite3.Connection):
    def __exit__(self, exc_type, exc_value, traceback) -> bool:
        result = super().__exit__(exc_type, exc_value, traceback)
        self.close()
        return result


def configure_sqlite_connection(conn: sqlite3.Connection, *, set_journal_mode: bool = True) -> None:
    conn.execute(f"PRAGMA busy_timeout={SQLITE_BUSY_TIMEOUT_MS}")
    conn.execute("PRAGMA synchronous=NORMAL")
    if set_journal_mode:
        conn.execute("PRAGMA journal_mode=WAL")


def connect_sqlite(path: Path, *, set_journal_mode: bool = True) -> sqlite3.Connection:
    conn = sqlite3.connect(path, timeout=SQLITE_TIMEOUT_SECONDS, factory=ClosingConnection)
    configure_sqlite_connection(conn, set_journal_mode=set_journal_mode)
    return conn


def connect_db() -> sqlite3.Connection:
    return connect_sqlite(DB_PATH)


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def now_iso() -> str:
    return now_utc().isoformat()


def format_local_datetime(value: Optional[Any]) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        dt = parse_iso_datetime(str(value))

    if not dt:
        return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt.astimezone(LOCAL_TIMEZONE).strftime("%d/%m/%Y %H:%M:%S")


def now_local_display() -> str:
    return format_local_datetime(now_utc()) or ""


def log_error(scope: str, error: Exception) -> None:
    print(f"[{scope}] error: {type(error).__name__}: {error}", flush=True)


def parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return None


def market_has_ended(market: Optional[dict[str, Any]]) -> bool:
    if not market:
        return True

    end_dt = parse_iso_datetime(market.get("end_time"))
    return bool(end_dt and end_dt <= now_utc())


async def sleep_until_next_tick(interval_seconds: int, started_at: float) -> None:
    elapsed = asyncio.get_running_loop().time() - started_at
    delay = max(0.0, interval_seconds - elapsed)
    await asyncio.sleep(delay)


def init_db() -> None:
    with connect_db() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS events (
            local_event_id INTEGER PRIMARY KEY AUTOINCREMENT,
            polymarket_event_id TEXT,
            polymarket_market_id TEXT,
            condition_id TEXT,
            event_slug TEXT UNIQUE,
            market_slug TEXT,
            title TEXT,
            question TEXT,
            event_url TEXT,
            start_time TEXT,
            start_time_local TEXT,
            end_time TEXT,
            end_time_local TEXT,
            yes_token_id TEXT,
            no_token_id TEXT,
            outcomes TEXT,
            outcome_prices TEXT,
            active INTEGER,
            closed INTEGER,
            enable_order_book INTEGER,
            accepting_orders INTEGER,
            fees_enabled INTEGER,
            fee_rate REAL,
            fee_calculation_source TEXT,
            fee_calculation_version TEXT,
            created_at_poly TEXT,
            created_at_poly_local TEXT,
            discovered_at TEXT,
            discovered_at_local TEXT,
            last_seen_at TEXT,
            last_seen_at_local TEXT,
            status TEXT,
            notes TEXT,
            raw_json TEXT
        )
        """)

        conn.execute("""
        CREATE TABLE IF NOT EXISTS orderbook_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sampled_at TEXT,
            sampled_at_local TEXT,
            event_slug TEXT,
            condition_id TEXT,
            up_token_id TEXT,
            down_token_id TEXT,
            up_best_ask REAL,
            up_best_bid REAL,
            down_best_ask REAL,
            down_best_bid REAL,
            up_last_trade_price REAL,
            down_last_trade_price REAL,
            up_spread REAL,
            down_spread REAL,
            up_midpoint REAL,
            down_midpoint REAL,
            raw_up_timestamp TEXT,
            raw_down_timestamp TEXT,
            up_volume_shares_10s REAL,
            down_volume_shares_10s REAL,
            up_volume_usdc_10s REAL,
            down_volume_usdc_10s REAL,
            trades_count_10s INTEGER,
            trades_window_start TEXT,
            trades_window_start_local TEXT,
            trades_window_end TEXT,
            trades_window_end_local TEXT,
            trades_error TEXT,
            status TEXT,
            error TEXT
        )
        """)

        conn.execute("""
        CREATE TABLE IF NOT EXISTS btc_volume_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sampled_at TEXT NOT NULL,
            sample_bucket_at TEXT NOT NULL,
            candle_start_at TEXT NOT NULL,
            product_id TEXT NOT NULL,
            granularity_seconds INTEGER NOT NULL,
            volume_btc_cumulative REAL,
            volume_btc_delta REAL,
            seconds_since_previous_sample REAL,
            event_slug TEXT,
            condition_id TEXT,
            source TEXT NOT NULL,
            status TEXT NOT NULL,
            error TEXT
        )
        """)

        conn.execute("""
        CREATE TABLE IF NOT EXISTS rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            entry_price REAL NOT NULL,
            stop_loss_price REAL NOT NULL,
            take_profit_price REAL NOT NULL,
            max_yes_entries_per_event INTEGER NOT NULL,
            max_no_entries_per_event INTEGER NOT NULL,
            status TEXT NOT NULL CHECK (status IN ('active', 'inactive')),
            eligible_after_event_id TEXT,
            entry_window_start_seconds_before_end INTEGER,
            entry_window_end_seconds_before_end INTEGER,
            schedule_timezone TEXT NOT NULL DEFAULT 'Asia/Jerusalem'
        )
        """)

        conn.execute("""
        CREATE TABLE IF NOT EXISTS rule_inactive_windows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_id INTEGER NOT NULL,
            day_of_week INTEGER NOT NULL CHECK (day_of_week BETWEEN 0 AND 6),
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            status TEXT NOT NULL CHECK (status IN ('active', 'inactive')),
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (rule_id) REFERENCES rules(id) ON DELETE CASCADE
        )
        """)

        conn.execute("""
        CREATE TABLE IF NOT EXISTS deals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_id INTEGER NOT NULL,
            rule_name TEXT,
            event_id TEXT NOT NULL,
            side TEXT NOT NULL CHECK (side IN ('yes', 'no')),
            result TEXT NOT NULL CHECK (result IN ('open', 'win', 'loss')),
            entry_at TEXT NOT NULL,
            entry_price REAL NOT NULL,
            entry_orderbook_log_id INTEGER NOT NULL,
            exit_at TEXT,
            exit_price REAL,
            exit_orderbook_log_id INTEGER,
            exit_reason TEXT CHECK (exit_reason IS NULL OR exit_reason IN ('take_profit', 'stop_loss', 'event_resolution')),
            market_result TEXT,
            price_change_points REAL,
            return_percent REAL,
            investment_usd REAL,
            shares REAL,
            entry_gross_value_usd REAL,
            entry_liquidity_role TEXT,
            entry_fee_rate REAL,
            entry_fee_usd REAL,
            exit_gross_value_usd REAL,
            exit_liquidity_role TEXT,
            exit_fee_rate REAL,
            exit_fee_usd REAL,
            total_fees_usd REAL,
            gross_pnl_usd REAL,
            net_pnl_usd REAL,
            gross_roi_percent REAL,
            net_roi_percent REAL,
            entry_btc_volume_log_id INTEGER,
            entry_btc_volume_sampled_at TEXT,
            entry_btc_volume_btc_cumulative REAL,
            entry_btc_volume_btc_delta REAL,
            entry_btc_volume_status TEXT,
            fee_calculation_source TEXT,
            fee_calculation_version TEXT,
            entry_seconds_before_event_end INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (rule_id) REFERENCES rules(id),
            FOREIGN KEY (entry_orderbook_log_id) REFERENCES orderbook_log(id),
            FOREIGN KEY (exit_orderbook_log_id) REFERENCES orderbook_log(id),
            FOREIGN KEY (entry_btc_volume_log_id) REFERENCES btc_volume_log(id)
        )
        """)

        ensure_column(conn, "events", "start_time_local", "TEXT")
        ensure_column(conn, "events", "end_time_local", "TEXT")
        ensure_column(conn, "events", "created_at_poly_local", "TEXT")
        ensure_column(conn, "events", "discovered_at_local", "TEXT")
        ensure_column(conn, "events", "last_seen_at_local", "TEXT")
        ensure_column(conn, "events", "fees_enabled", "INTEGER")
        ensure_column(conn, "events", "fee_rate", "REAL")
        ensure_column(conn, "events", "fee_calculation_source", "TEXT")
        ensure_column(conn, "events", "fee_calculation_version", "TEXT")
        ensure_column(conn, "orderbook_log", "sampled_at_local", "TEXT")
        ensure_column(conn, "orderbook_log", "up_volume_shares_10s", "REAL")
        ensure_column(conn, "orderbook_log", "down_volume_shares_10s", "REAL")
        ensure_column(conn, "orderbook_log", "up_volume_usdc_10s", "REAL")
        ensure_column(conn, "orderbook_log", "down_volume_usdc_10s", "REAL")
        ensure_column(conn, "orderbook_log", "trades_count_10s", "INTEGER")
        ensure_column(conn, "orderbook_log", "trades_window_start", "TEXT")
        ensure_column(conn, "orderbook_log", "trades_window_start_local", "TEXT")
        ensure_column(conn, "orderbook_log", "trades_window_end", "TEXT")
        ensure_column(conn, "orderbook_log", "trades_window_end_local", "TEXT")
        ensure_column(conn, "orderbook_log", "trades_error", "TEXT")
        ensure_column(conn, "btc_volume_log", "sampled_at", "TEXT")
        ensure_column(conn, "btc_volume_log", "sample_bucket_at", "TEXT")
        ensure_column(conn, "btc_volume_log", "candle_start_at", "TEXT")
        ensure_column(conn, "btc_volume_log", "product_id", "TEXT")
        ensure_column(conn, "btc_volume_log", "granularity_seconds", "INTEGER")
        ensure_column(conn, "btc_volume_log", "volume_btc_cumulative", "REAL")
        ensure_column(conn, "btc_volume_log", "volume_btc_delta", "REAL")
        ensure_column(conn, "btc_volume_log", "seconds_since_previous_sample", "REAL")
        ensure_column(conn, "btc_volume_log", "event_slug", "TEXT")
        ensure_column(conn, "btc_volume_log", "condition_id", "TEXT")
        ensure_column(conn, "btc_volume_log", "source", "TEXT")
        ensure_column(conn, "btc_volume_log", "status", "TEXT")
        ensure_column(conn, "btc_volume_log", "error", "TEXT")
        ensure_column(conn, "rules", "eligible_after_event_id", "TEXT")
        ensure_column(conn, "rules", "entry_window_start_seconds_before_end", "INTEGER")
        ensure_column(conn, "rules", "entry_window_end_seconds_before_end", "INTEGER")
        ensure_column(conn, "rules", "schedule_timezone", "TEXT NOT NULL DEFAULT 'Asia/Jerusalem'")
        ensure_column(conn, "deals", "rule_name", "TEXT")
        ensure_column(conn, "deals", "investment_usd", "REAL")
        ensure_column(conn, "deals", "shares", "REAL")
        ensure_column(conn, "deals", "entry_gross_value_usd", "REAL")
        ensure_column(conn, "deals", "entry_liquidity_role", "TEXT")
        ensure_column(conn, "deals", "entry_fee_rate", "REAL")
        ensure_column(conn, "deals", "entry_fee_usd", "REAL")
        ensure_column(conn, "deals", "exit_gross_value_usd", "REAL")
        ensure_column(conn, "deals", "exit_liquidity_role", "TEXT")
        ensure_column(conn, "deals", "exit_fee_rate", "REAL")
        ensure_column(conn, "deals", "exit_fee_usd", "REAL")
        ensure_column(conn, "deals", "total_fees_usd", "REAL")
        ensure_column(conn, "deals", "gross_pnl_usd", "REAL")
        ensure_column(conn, "deals", "net_pnl_usd", "REAL")
        ensure_column(conn, "deals", "gross_roi_percent", "REAL")
        ensure_column(conn, "deals", "net_roi_percent", "REAL")
        ensure_column(conn, "deals", "entry_btc_volume_log_id", "INTEGER")
        ensure_column(conn, "deals", "entry_btc_volume_sampled_at", "TEXT")
        ensure_column(conn, "deals", "entry_btc_volume_btc_cumulative", "REAL")
        ensure_column(conn, "deals", "entry_btc_volume_btc_delta", "REAL")
        ensure_column(conn, "deals", "entry_btc_volume_status", "TEXT")
        ensure_column(conn, "deals", "fee_calculation_source", "TEXT")
        ensure_column(conn, "deals", "fee_calculation_version", "TEXT")
        ensure_column(conn, "deals", "entry_seconds_before_event_end", "INTEGER")
        conn.execute("UPDATE rules SET schedule_timezone = ? WHERE schedule_timezone IS NULL OR schedule_timezone = ''", (DEFAULT_SCHEDULE_TIMEZONE,))

        conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_btc_volume_log_unique_bucket
        ON btc_volume_log (product_id, candle_start_at, sample_bucket_at)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_btc_volume_log_sampled_at
        ON btc_volume_log (sampled_at)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_btc_volume_log_candle_start_at
        ON btc_volume_log (candle_start_at)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_btc_volume_log_event_slug
        ON btc_volume_log (event_slug)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_rules_status
        ON rules (status)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_rule_inactive_windows_rule_day_status
        ON rule_inactive_windows (rule_id, day_of_week, status)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_rule_id
        ON deals (rule_id)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_event_id
        ON deals (event_id)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_result
        ON deals (result)
        """)
        conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_deals_one_open_per_rule
        ON deals (rule_id)
        WHERE result = 'open'
        """)
        conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_deals_unique_entry_sample
        ON deals (rule_id, event_id, side, entry_orderbook_log_id)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_rule_event_side
        ON deals (rule_id, event_id, side)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_entry_btc_volume_delta
        ON deals (entry_btc_volume_btc_delta)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_entry_at
        ON deals (entry_at)
        """)
        conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_deals_exit_at
        ON deals (exit_at)
        """)

        backfill_entry_seconds_before_event_end(conn)

        backfill_demo_fee_snapshots(conn)

        conn.commit()
    conn.close()
    if configure_live is not None:
        configure_live(DB_PATH)


def ensure_column(conn: sqlite3.Connection, table: str, column: str, column_type: str) -> None:
    columns = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {column_type}")


def get_conn() -> sqlite3.Connection:
    conn = connect_db()
    conn.row_factory = sqlite3.Row
    return conn


def find_entry_btc_volume_snapshot(
    conn: sqlite3.Connection,
    entry_at: Any,
    event_id: Optional[str] = None,
) -> Optional[sqlite3.Row]:
    if not entry_at:
        return None

    params: tuple[Any, ...]
    if event_id:
        row = conn.execute("""
            SELECT
                id,
                sampled_at,
                volume_btc_cumulative,
                volume_btc_delta,
                status
            FROM btc_volume_log
            WHERE sampled_at <= ?
              AND event_slug = ?
              AND status IN ('success', 'baseline')
            ORDER BY sampled_at DESC, id DESC
            LIMIT 1
        """, (entry_at, event_id)).fetchone()
        if row:
            return row

    params = (entry_at,)
    return conn.execute("""
        SELECT
            id,
            sampled_at,
            volume_btc_cumulative,
            volume_btc_delta,
            status
        FROM btc_volume_log
        WHERE sampled_at <= ?
          AND status IN ('success', 'baseline')
        ORDER BY sampled_at DESC, id DESC
        LIMIT 1
    """, params).fetchone()


def calculate_entry_seconds_before_event_end(
    conn: sqlite3.Connection,
    event_id: Any,
    entry_at: Any,
) -> Optional[int]:
    entry_dt = parse_iso_datetime(str(entry_at)) if entry_at else None
    if entry_dt is None:
        return None
    if entry_dt.tzinfo is None:
        entry_dt = entry_dt.replace(tzinfo=timezone.utc)

    event = conn.execute(
        "SELECT end_time FROM events WHERE event_slug = ? LIMIT 1",
        (str(event_id),),
    ).fetchone()
    if not event or not event["end_time"]:
        return None

    end_dt = parse_iso_datetime(str(event["end_time"]))
    if end_dt is None:
        return None
    if end_dt.tzinfo is None:
        end_dt = end_dt.replace(tzinfo=timezone.utc)

    return int((end_dt - entry_dt).total_seconds())


def backfill_entry_seconds_before_event_end(conn: sqlite3.Connection) -> tuple[int, int]:
    conn.row_factory = sqlite3.Row
    deals = conn.execute("""
        SELECT id, event_id, entry_at
        FROM deals
        WHERE entry_seconds_before_event_end IS NULL
          AND entry_at IS NOT NULL
    """).fetchall()
    updated = 0
    for deal in deals:
        seconds = calculate_entry_seconds_before_event_end(conn, deal["event_id"], deal["entry_at"])
        if seconds is None:
            continue
        conn.execute(
            "UPDATE deals SET entry_seconds_before_event_end = ?, updated_at = ? WHERE id = ?",
            (seconds, now_iso(), deal["id"]),
        )
        updated += 1
    return updated, len(deals) - updated


def backfill_deal_btc_volume_snapshots(conn: sqlite3.Connection) -> None:
    conn.row_factory = sqlite3.Row
    deals = conn.execute("""
        SELECT id, event_id, entry_at
        FROM deals
        WHERE entry_btc_volume_log_id IS NULL
          AND entry_at IS NOT NULL
    """).fetchall()

    for deal in deals:
        snapshot = find_entry_btc_volume_snapshot(conn, deal["entry_at"], deal["event_id"])
        if not snapshot:
            continue
        conn.execute("""
            UPDATE deals SET
                entry_btc_volume_log_id = ?,
                entry_btc_volume_sampled_at = ?,
                entry_btc_volume_btc_cumulative = ?,
                entry_btc_volume_btc_delta = ?,
                entry_btc_volume_status = ?,
                updated_at = COALESCE(updated_at, ?)
            WHERE id = ?
        """, (
            snapshot["id"],
            snapshot["sampled_at"],
            snapshot["volume_btc_cumulative"],
            snapshot["volume_btc_delta"],
            snapshot["status"],
            now_iso(),
            deal["id"],
        ))


def decimal_price(value: Any, field_name: str) -> Decimal:
    try:
        price = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a numeric price")

    if not price.is_finite():
        raise ValueError(f"{field_name} must be a finite numeric price")
    if price <= Decimal("0"):
        raise ValueError(f"{field_name} must be positive")
    if price >= Decimal("1"):
        raise ValueError(f"{field_name} must be below 1")
    return price.quantize(Decimal("0.000001"))


def decimal_from_db(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.000001"))
    except (InvalidOperation, ValueError):
        return None


def prices_equal(left: Any, right: Any) -> bool:
    left_price = decimal_from_db(left)
    right_price = decimal_from_db(right)
    return left_price is not None and right_price is not None and left_price == right_price


def optional_non_negative_int(value: Any, field_name: str) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} must be a non-negative integer")
    if number < 0:
        raise ValueError(f"{field_name} must be non-negative")
    return number


def normalize_schedule_timezone(value: Any) -> str:
    timezone_name = str(value or DEFAULT_SCHEDULE_TIMEZONE).strip() or DEFAULT_SCHEDULE_TIMEZONE
    try:
        ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError:
        raise ValueError("schedule_timezone must be a valid IANA timezone")
    return timezone_name


def normalize_clock_time(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, fmt).time()
            return parsed.strftime("%H:%M:%S")
        except ValueError:
            continue
    raise ValueError(f"{field_name} must use HH:MM or HH:MM:SS")


def validate_inactive_windows(value: Any) -> list[dict[str, Any]]:
    if value in (None, ""):
        return []
    if not isinstance(value, list):
        raise ValueError("inactive_windows must be a list")
    windows: list[dict[str, Any]] = []
    for index, window in enumerate(value):
        if not isinstance(window, dict):
            raise ValueError(f"inactive_windows[{index}] must be an object")
        try:
            day_of_week = int(window.get("day_of_week"))
        except (TypeError, ValueError):
            raise ValueError(f"inactive_windows[{index}].day_of_week must be 0-6")
        if day_of_week < 0 or day_of_week > 6:
            raise ValueError(f"inactive_windows[{index}].day_of_week must be 0-6")
        status = str(window.get("status", "active")).strip().lower()
        if status not in {"active", "inactive"}:
            raise ValueError(f"inactive_windows[{index}].status must be active or inactive")
        windows.append({
            "day_of_week": day_of_week,
            "start_time": normalize_clock_time(window.get("start_time"), f"inactive_windows[{index}].start_time"),
            "end_time": normalize_clock_time(window.get("end_time"), f"inactive_windows[{index}].end_time"),
            "status": status,
        })
    return windows


def validate_rule_payload(payload: dict[str, Any]) -> dict[str, Any]:
    name = str(payload.get("name", "")).strip()
    if not name:
        raise ValueError("name must not be empty")

    entry_price = decimal_price(payload.get("entry_price"), "entry_price")
    stop_loss_price = decimal_price(payload.get("stop_loss_price"), "stop_loss_price")
    take_profit_price = decimal_price(payload.get("take_profit_price"), "take_profit_price")

    if entry_price == Decimal("0.500000"):
        raise ValueError("entry_price must not be 0.5")
    if stop_loss_price >= entry_price:
        raise ValueError("stop_loss_price must be below entry_price")
    if take_profit_price <= entry_price:
        raise ValueError("take_profit_price must be above entry_price")

    try:
        max_yes = int(payload.get("max_yes_entries_per_event"))
        max_no = int(payload.get("max_no_entries_per_event"))
    except (TypeError, ValueError):
        raise ValueError("entry limits must be non-negative integers")

    if max_yes < 0:
        raise ValueError("max_yes_entries_per_event must be non-negative")
    if max_no < 0:
        raise ValueError("max_no_entries_per_event must be non-negative")

    status = str(payload.get("status", "active")).strip().lower()
    if status not in {"active", "inactive"}:
        raise ValueError("status must be active or inactive")

    window_start = optional_non_negative_int(
        payload.get("entry_window_start_seconds_before_end"),
        "entry_window_start_seconds_before_end",
    )
    window_end = optional_non_negative_int(
        payload.get("entry_window_end_seconds_before_end"),
        "entry_window_end_seconds_before_end",
    )
    if (window_start is None) != (window_end is None):
        raise ValueError("entry window start and end must both be empty or both be set")
    if window_start is not None and window_end is not None and window_start < window_end:
        raise ValueError("entry_window_start_seconds_before_end must be >= entry_window_end_seconds_before_end")

    return {
        "name": name,
        "entry_price": float(entry_price),
        "stop_loss_price": float(stop_loss_price),
        "take_profit_price": float(take_profit_price),
        "max_yes_entries_per_event": max_yes,
        "max_no_entries_per_event": max_no,
        "status": status,
        "entry_window_start_seconds_before_end": window_start,
        "entry_window_end_seconds_before_end": window_end,
        "schedule_timezone": normalize_schedule_timezone(payload.get("schedule_timezone")),
        "inactive_windows": validate_inactive_windows(payload.get("inactive_windows")),
    }


def active_event_id_for_rule_gate() -> Optional[str]:
    event_slug, _ = current_active_market_snapshot()
    return event_slug


def create_rule(payload: dict[str, Any]) -> sqlite3.Row:
    values = validate_rule_payload(payload)
    created_at = now_iso()
    eligible_after_event_id = active_event_id_for_rule_gate() if values["status"] == "active" else None

    try:
        with get_conn() as conn:
            cursor = conn.execute("""
                INSERT INTO rules (
                    name,
                    created_at,
                    updated_at,
                    entry_price,
                    stop_loss_price,
                    take_profit_price,
                    max_yes_entries_per_event,
                    max_no_entries_per_event,
                    status,
                    eligible_after_event_id,
                    entry_window_start_seconds_before_end,
                    entry_window_end_seconds_before_end,
                    schedule_timezone
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                values["name"],
                created_at,
                created_at,
                values["entry_price"],
                values["stop_loss_price"],
                values["take_profit_price"],
                values["max_yes_entries_per_event"],
                values["max_no_entries_per_event"],
                values["status"],
                eligible_after_event_id,
                values["entry_window_start_seconds_before_end"],
                values["entry_window_end_seconds_before_end"],
                values["schedule_timezone"],
            ))
            rule_id = cursor.lastrowid
            for window in values["inactive_windows"]:
                conn.execute("""
                    INSERT INTO rule_inactive_windows (
                        rule_id, day_of_week, start_time, end_time, status, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    rule_id,
                    window["day_of_week"],
                    window["start_time"],
                    window["end_time"],
                    window["status"],
                    created_at,
                    created_at,
                ))
            conn.commit()
            row = conn.execute("SELECT * FROM rules WHERE id = ?", (rule_id,)).fetchone()
    except sqlite3.Error as exc:
        log_error("create rule db", exc)
        raise

    print(
        f"[rules] created id={rule_id} name={values['name']} status={values['status']} "
        f"eligible_after_event_id={eligible_after_event_id}",
        flush=True,
    )
    if eligible_after_event_id:
        print(f"[rules] rule id={rule_id} waits for next event after {eligible_after_event_id}", flush=True)
    return row


def row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def load_rule_inactive_windows(conn: sqlite3.Connection, rule_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not rule_ids:
        return {}
    placeholders = ",".join("?" for _ in rule_ids)
    rows = conn.execute(f"""
        SELECT *
        FROM rule_inactive_windows
        WHERE rule_id IN ({placeholders})
        ORDER BY rule_id ASC, day_of_week ASC, start_time ASC, id ASC
    """, tuple(rule_ids)).fetchall()
    grouped: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        item = row_to_dict(row)
        item["day_label_he"] = WEEKDAY_LABELS_HE.get(int(row["day_of_week"]), str(row["day_of_week"]))
        grouped.setdefault(int(row["rule_id"]), []).append(item)
    return grouped


def rules_to_dicts(conn: sqlite3.Connection, rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    rule_ids = [int(row["id"]) for row in rows]
    windows_by_rule = load_rule_inactive_windows(conn, rule_ids)
    result = []
    for row in rows:
        item = row_to_dict(row)
        item["inactive_windows"] = windows_by_rule.get(int(row["id"]), [])
        result.append(item)
    return result


def deactivate_rule(rule_id: int) -> tuple[sqlite3.Row, str]:
    try:
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM rules WHERE id = ?", (rule_id,)).fetchone()
            if not row:
                raise KeyError(f"Rule {rule_id} does not exist")
            if row["status"] == "inactive":
                print(f"[rules] deactivate requested for already inactive id={rule_id}", flush=True)
                return row, "Rule is already inactive"

            updated_at = now_iso()
            conn.execute(
                "UPDATE rules SET status = 'inactive', updated_at = ? WHERE id = ?",
                (updated_at, rule_id),
            )
            conn.commit()
            updated = conn.execute("SELECT * FROM rules WHERE id = ?", (rule_id,)).fetchone()
    except sqlite3.Error as exc:
        log_error("deactivate rule db", exc)
        raise

    print(f"[rules] deactivated id={rule_id}", flush=True)
    return updated, "Rule deactivated"


def calculate_deal_metrics(entry_price: Any, exit_price: Any) -> tuple[float, float]:
    entry = Decimal(str(entry_price))
    exit_value = Decimal(str(exit_price))
    points = abs(exit_value - entry) * Decimal("100")
    return_percent = ((exit_value - entry) / entry) * Decimal("100")
    return float(points), float(return_percent)


def normalize_investment_usd(value: Any) -> Decimal:
    try:
        investment = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("1")

    if not investment.is_finite() or investment <= 0:
        return Decimal("1")
    return investment.quantize(Decimal("0.01"))


def calculate_deal_pnl_usd(entry_price: Any, exit_price: Any, investment_usd: Any = 1) -> tuple[float, float, float]:
    entry = Decimal(str(entry_price))
    exit_value = Decimal(str(exit_price))
    investment = normalize_investment_usd(investment_usd)

    if entry <= 0:
        raise ValueError("entry_price must be positive")

    shares = investment / entry
    pnl = shares * (exit_value - entry)
    roi_percent = (pnl / investment) * Decimal("100")
    return float(pnl), float(roi_percent), float(shares)


def decimal_to_float(value: Optional[Decimal]) -> Optional[float]:
    return None if value is None else float(value)


def decimal_money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT)


def rounded_fee(value: Decimal) -> Decimal:
    rounded = value.quantize(FEE_QUANT, rounding=ROUND_HALF_UP)
    return Decimal("0.00000") if rounded < FEE_QUANT else rounded


def calculate_platform_fee_usd(
    shares: Decimal,
    price: Decimal,
    fee_rate: Decimal,
    liquidity_role: str,
) -> Decimal:
    if liquidity_role.upper() != "TAKER" or fee_rate <= 0:
        return Decimal("0.00000")
    return rounded_fee(shares * fee_rate * price * (Decimal("1") - price))


def normalize_fee_rate(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    try:
        rate = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    if not rate.is_finite() or rate < 0:
        return None
    return rate


def extract_event_fee_snapshot(event_id: Optional[str], conn: Optional[sqlite3.Connection] = None) -> dict[str, Any]:
    snapshot = {
        "fees_enabled": True,
        "fee_rate": POLYMARKET_CRYPTO_TAKER_FEE_RATE,
        "fee_calculation_source": DEMO_FEE_SOURCE_FALLBACK,
        "fee_calculation_version": DEMO_FEE_CALCULATION_VERSION,
    }
    if not event_id:
        return snapshot

    owns_conn = conn is None
    active_conn = conn or get_conn()
    try:
        row = active_conn.execute("""
            SELECT fees_enabled, fee_rate, fee_calculation_source, fee_calculation_version
            FROM events
            WHERE event_slug = ?
            LIMIT 1
        """, (event_id,)).fetchone()
    finally:
        if owns_conn:
            active_conn.close()

    if not row:
        return snapshot

    fees_enabled = row["fees_enabled"]
    fee_rate = normalize_fee_rate(row["fee_rate"])
    source = row["fee_calculation_source"] or DEMO_FEE_SOURCE_FALLBACK
    if fees_enabled is not None and int(fees_enabled) == 0 and source == CLOB_FEE_SOURCE:
        return {
            "fees_enabled": False,
            "fee_rate": Decimal("0"),
            "fee_calculation_source": source,
            "fee_calculation_version": row["fee_calculation_version"] or DEMO_FEE_CALCULATION_VERSION,
        }
    if fee_rate is not None and (fee_rate > 0 or source == CLOB_FEE_SOURCE):
        return {
            "fees_enabled": True,
            "fee_rate": fee_rate,
            "fee_calculation_source": source,
            "fee_calculation_version": row["fee_calculation_version"] or DEMO_FEE_CALCULATION_VERSION,
        }
    return snapshot


def calculate_demo_deal_financials(
    entry_price: Any,
    exit_price: Any = None,
    investment_usd: Any = DEMO_INVESTMENT_USD,
    entry_liquidity_role: str = DEMO_ENTRY_LIQUIDITY_ROLE,
    exit_liquidity_role: Optional[str] = None,
    fee_rate: Any = POLYMARKET_CRYPTO_TAKER_FEE_RATE,
    fee_source: str = DEMO_FEE_SOURCE_FALLBACK,
    fee_version: str = DEMO_FEE_CALCULATION_VERSION,
) -> dict[str, Any]:
    entry = Decimal(str(entry_price))
    if entry <= 0:
        raise ValueError("entry_price must be positive")

    investment = normalize_investment_usd(investment_usd)
    rate = normalize_fee_rate(fee_rate) or Decimal("0")
    shares = investment / entry
    entry_fee = calculate_platform_fee_usd(shares, entry, rate, entry_liquidity_role)
    result = {
        "investment_usd": decimal_money(investment),
        "shares": decimal_money(shares),
        "entry_gross_value_usd": decimal_money(investment),
        "entry_liquidity_role": entry_liquidity_role.upper(),
        "entry_fee_rate": rate,
        "entry_fee_usd": entry_fee,
        "fee_calculation_source": fee_source,
        "fee_calculation_version": fee_version,
    }

    if exit_price is None:
        return result

    exit_value = Decimal(str(exit_price))
    exit_role = (exit_liquidity_role or DEMO_EXIT_LIQUIDITY_ROLE_BY_REASON["event_resolution"]).upper()
    exit_gross = shares * exit_value
    gross_pnl = exit_gross - investment
    exit_fee = calculate_platform_fee_usd(shares, exit_value, rate, exit_role)
    total_fees = entry_fee + exit_fee
    net_pnl = gross_pnl - total_fees
    result.update({
        "exit_gross_value_usd": decimal_money(exit_gross),
        "exit_liquidity_role": exit_role,
        "exit_fee_rate": rate,
        "exit_fee_usd": exit_fee,
        "total_fees_usd": total_fees,
        "gross_pnl_usd": decimal_money(gross_pnl),
        "net_pnl_usd": decimal_money(net_pnl),
        "gross_roi_percent": (gross_pnl / investment) * Decimal("100"),
        "net_roi_percent": (net_pnl / investment) * Decimal("100"),
    })
    return result


def deal_financials_from_row(deal: sqlite3.Row, investment_usd: Any = 1) -> dict[str, Any]:
    if "net_pnl_usd" in deal.keys() and deal["net_pnl_usd"] is not None:
        return {
            "gross_pnl_usd": float(deal["gross_pnl_usd"] or 0),
            "net_pnl_usd": float(deal["net_pnl_usd"]),
            "gross_roi_percent": float(deal["gross_roi_percent"] or 0),
            "net_roi_percent": float(deal["net_roi_percent"] or 0),
            "total_fees_usd": float(deal["total_fees_usd"] or 0),
            "entry_fee_usd": float(deal["entry_fee_usd"] or 0),
            "exit_fee_usd": float(deal["exit_fee_usd"] or 0),
            "shares": float(deal["shares"] or 0),
        }
    pnl_usd, roi_percent, shares = calculate_deal_pnl_usd(deal["entry_price"], deal["exit_price"], investment_usd)
    return {
        "gross_pnl_usd": pnl_usd,
        "net_pnl_usd": pnl_usd,
        "gross_roi_percent": roi_percent,
        "net_roi_percent": roi_percent,
        "total_fees_usd": 0.0,
        "entry_fee_usd": 0.0,
        "exit_fee_usd": 0.0,
        "shares": shares,
    }


def backfill_demo_fee_snapshots(conn: sqlite3.Connection) -> int:
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT *
        FROM deals
        WHERE result IN ('win', 'loss')
          AND exit_price IS NOT NULL
          AND (
            fee_calculation_source IS NULL
            OR total_fees_usd IS NULL
            OR (
                COALESCE(total_fees_usd, 0) = 0
                AND COALESCE(entry_fee_rate, 0) = 0
                AND COALESCE(exit_fee_rate, 0) = 0
            )
          )
        ORDER BY id ASC
    """).fetchall()

    updated = 0
    for deal in rows:
        exit_role = DEMO_EXIT_LIQUIDITY_ROLE_BY_REASON.get(deal["exit_reason"] or "", "TAKER")
        try:
            financials = calculate_demo_deal_financials(
                deal["entry_price"],
                deal["exit_price"],
                deal["investment_usd"] if deal["investment_usd"] is not None else DEMO_INVESTMENT_USD,
                deal["entry_liquidity_role"] or DEMO_ENTRY_LIQUIDITY_ROLE,
                deal["exit_liquidity_role"] or exit_role,
                POLYMARKET_CRYPTO_TAKER_FEE_RATE,
                DEMO_FEE_SOURCE_BACKFILL,
                DEMO_FEE_CALCULATION_VERSION,
            )
        except (InvalidOperation, ValueError, TypeError, ZeroDivisionError):
            continue

        conn.execute("""
            UPDATE deals SET
                investment_usd = ?,
                shares = ?,
                entry_gross_value_usd = ?,
                entry_liquidity_role = ?,
                entry_fee_rate = ?,
                entry_fee_usd = ?,
                exit_gross_value_usd = ?,
                exit_liquidity_role = ?,
                exit_fee_rate = ?,
                exit_fee_usd = ?,
                total_fees_usd = ?,
                gross_pnl_usd = ?,
                net_pnl_usd = ?,
                gross_roi_percent = ?,
                net_roi_percent = ?,
                fee_calculation_source = ?,
                fee_calculation_version = ?,
                updated_at = ?
            WHERE id = ?
        """, (
            decimal_to_float(financials["investment_usd"]),
            decimal_to_float(financials["shares"]),
            decimal_to_float(financials["entry_gross_value_usd"]),
            financials["entry_liquidity_role"],
            decimal_to_float(financials["entry_fee_rate"]),
            decimal_to_float(financials["entry_fee_usd"]),
            decimal_to_float(financials["exit_gross_value_usd"]),
            financials["exit_liquidity_role"],
            decimal_to_float(financials["exit_fee_rate"]),
            decimal_to_float(financials["exit_fee_usd"]),
            decimal_to_float(financials["total_fees_usd"]),
            decimal_to_float(financials["gross_pnl_usd"]),
            decimal_to_float(financials["net_pnl_usd"]),
            decimal_to_float(financials["gross_roi_percent"]),
            decimal_to_float(financials["net_roi_percent"]),
            financials["fee_calculation_source"],
            financials["fee_calculation_version"],
            now_iso(),
            deal["id"],
        ))
        updated += 1

    if updated:
        print(f"[fees] simulated backfill updated {updated} closed deals", flush=True)
    return updated


def close_deal(
    conn: sqlite3.Connection,
    deal: sqlite3.Row,
    result: str,
    exit_reason: str,
    exit_price: Any,
    exit_at: str,
    exit_orderbook_log_id: Optional[int],
    market_result: Optional[str] = None,
) -> None:
    price_change_points, return_percent = calculate_deal_metrics(deal["entry_price"], exit_price)
    fee_snapshot = extract_event_fee_snapshot(deal["event_id"], conn)
    fee_rate = normalize_fee_rate(deal["entry_fee_rate"] if "entry_fee_rate" in deal.keys() else None)
    if fee_rate is None:
        fee_rate = fee_snapshot["fee_rate"]
    fee_source = (
        deal["fee_calculation_source"]
        if "fee_calculation_source" in deal.keys() and deal["fee_calculation_source"]
        else fee_snapshot["fee_calculation_source"]
    )
    fee_version = (
        deal["fee_calculation_version"]
        if "fee_calculation_version" in deal.keys() and deal["fee_calculation_version"]
        else fee_snapshot["fee_calculation_version"]
    )
    exit_role = DEMO_EXIT_LIQUIDITY_ROLE_BY_REASON.get(exit_reason, "TAKER")
    financials = calculate_demo_deal_financials(
        deal["entry_price"],
        exit_price,
        deal["investment_usd"] if "investment_usd" in deal.keys() and deal["investment_usd"] is not None else DEMO_INVESTMENT_USD,
        deal["entry_liquidity_role"] if "entry_liquidity_role" in deal.keys() and deal["entry_liquidity_role"] else DEMO_ENTRY_LIQUIDITY_ROLE,
        exit_role,
        fee_rate,
        fee_source,
        fee_version,
    )
    conn.execute("""
        UPDATE deals SET
            result = ?,
            exit_at = ?,
            exit_price = ?,
            exit_orderbook_log_id = ?,
            exit_reason = ?,
            market_result = ?,
            price_change_points = ?,
            return_percent = ?,
            investment_usd = COALESCE(investment_usd, ?),
            shares = COALESCE(shares, ?),
            entry_gross_value_usd = COALESCE(entry_gross_value_usd, ?),
            entry_liquidity_role = COALESCE(entry_liquidity_role, ?),
            entry_fee_rate = COALESCE(entry_fee_rate, ?),
            entry_fee_usd = COALESCE(entry_fee_usd, ?),
            exit_gross_value_usd = ?,
            exit_liquidity_role = ?,
            exit_fee_rate = ?,
            exit_fee_usd = ?,
            total_fees_usd = ?,
            gross_pnl_usd = ?,
            net_pnl_usd = ?,
            gross_roi_percent = ?,
            net_roi_percent = ?,
            fee_calculation_source = COALESCE(fee_calculation_source, ?),
            fee_calculation_version = COALESCE(fee_calculation_version, ?),
            updated_at = ?
        WHERE id = ? AND result = 'open'
    """, (
        result,
        exit_at,
        float(Decimal(str(exit_price))),
        exit_orderbook_log_id,
        exit_reason,
        market_result,
        price_change_points,
        return_percent,
        decimal_to_float(financials["investment_usd"]),
        decimal_to_float(financials["shares"]),
        decimal_to_float(financials["entry_gross_value_usd"]),
        financials["entry_liquidity_role"],
        decimal_to_float(financials["entry_fee_rate"]),
        decimal_to_float(financials["entry_fee_usd"]),
        decimal_to_float(financials["exit_gross_value_usd"]),
        financials["exit_liquidity_role"],
        decimal_to_float(financials["exit_fee_rate"]),
        decimal_to_float(financials["exit_fee_usd"]),
        decimal_to_float(financials["total_fees_usd"]),
        decimal_to_float(financials["gross_pnl_usd"]),
        decimal_to_float(financials["net_pnl_usd"]),
        decimal_to_float(financials["gross_roi_percent"]),
        decimal_to_float(financials["net_roi_percent"]),
        financials["fee_calculation_source"],
        financials["fee_calculation_version"],
        now_iso(),
        deal["id"],
    ))


def resolve_market_result(event_row: sqlite3.Row) -> Optional[str]:
    if int(event_row["closed"] or 0) != 1 and event_row["status"] != "closed":
        return None

    prices = safe_json_loads(event_row["outcome_prices"])
    if len(prices) < 2:
        return None

    parsed: list[Decimal] = []
    for value in prices[:2]:
        try:
            parsed.append(Decimal(str(value)))
        except (InvalidOperation, ValueError):
            return None

    if parsed[0] == Decimal("1") and parsed[1] == Decimal("0"):
        return "yes"
    if parsed[1] == Decimal("1") and parsed[0] == Decimal("0"):
        return "no"
    return None


def close_deals_for_event_resolution(conn: sqlite3.Connection, event_id: str) -> None:
    conn.row_factory = sqlite3.Row
    event_row = conn.execute("""
        SELECT *
        FROM events
        WHERE event_slug = ?
        LIMIT 1
    """, (event_id,)).fetchone()
    if not event_row:
        return

    market_result = resolve_market_result(event_row)
    if not market_result:
        return

    open_deals = conn.execute("""
        SELECT *
        FROM deals
        WHERE event_id = ? AND result = 'open'
    """, (event_id,)).fetchall()

    for deal in open_deals:
        result = "win" if deal["side"] == market_result else "loss"
        exit_price = 1 if result == "win" else 0
        close_deal(
            conn,
            deal,
            result,
            "event_resolution",
            exit_price,
            now_iso(),
            None,
            market_result,
        )
        print(
            f"[deals] closed id={deal['id']} rule_id={deal['rule_id']} "
            f"reason=event_resolution market_result={market_result} result={result}",
            flush=True,
        )


def floor_to_5m_epoch(dt: datetime) -> int:
    ts = int(dt.timestamp())
    return (ts // 300) * 300


def floor_to_epoch(dt: datetime, seconds: int) -> int:
    ts = int(dt.timestamp())
    return (ts // seconds) * seconds


def iso_from_epoch(epoch: int | float) -> str:
    return datetime.fromtimestamp(float(epoch), timezone.utc).isoformat()


def iso_z(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def truncate_text(value: Optional[str], limit: int = 240) -> Optional[str]:
    if not value:
        return None
    return value if len(value) <= limit else f"{value[:limit - 3]}..."


def candidate_slugs() -> list[str]:
    """
    בודק כמה חלונות סביב הזמן הנוכחי.
    זה חשוב כי לפעמים market עתידי כבר פתוח להזמנות כמה דקות לפני תחילת החלון.
    """
    base = floor_to_5m_epoch(now_utc())
    candidates = [
        base - 300,
        base,
        base + 300,
        base + 600,
    ]
    return [f"btc-updown-5m-{ts}" for ts in candidates]


def safe_json_loads(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            loaded = json.loads(value)
            return loaded if isinstance(loaded, list) else []
        except Exception:
            return []
    return []


def is_market_open(event_data: dict[str, Any], market: dict[str, Any]) -> bool:
    end_dt = parse_iso_datetime(market.get("endDate") or event_data.get("endDate"))
    if end_dt and end_dt <= now_utc():
        return False

    return (
        event_data.get("active") is True
        and event_data.get("closed") is False
        and market.get("active") is True
        and market.get("closed") is False
        and market.get("acceptingOrders") is True
        and market.get("enableOrderBook") is True
    )


def extract_market_fee_config(event_data: dict[str, Any], market: dict[str, Any]) -> dict[str, Any]:
    raw_enabled = market.get("feesEnabled", event_data.get("feesEnabled"))
    fee_schedule = market.get("feeSchedule") or event_data.get("feeSchedule") or {}
    fd = market.get("fd") or fee_schedule.get("fd") or {}
    raw_rate = (
        fee_schedule.get("feeRate")
        or fee_schedule.get("takerFeeRate")
        or fd.get("r")
        or market.get("feeRate")
        or event_data.get("feeRate")
    )
    fee_rate = normalize_fee_rate(raw_rate)

    if raw_enabled is True and fee_rate is not None:
        return {
            "fees_enabled": 1,
            "fee_rate": float(fee_rate),
            "fee_calculation_source": DEMO_FEE_SOURCE_MARKET,
            "fee_calculation_version": DEMO_FEE_CALCULATION_VERSION,
        }
    return {
        "fees_enabled": None,
        "fee_rate": None,
        "fee_calculation_source": None,
        "fee_calculation_version": None,
    }


def extract_market(event_data: dict[str, Any]) -> Optional[dict[str, Any]]:
    markets = event_data.get("markets") or []
    if not markets:
        return None

    market = markets[0]

    outcomes = safe_json_loads(market.get("outcomes"))
    token_ids = safe_json_loads(market.get("clobTokenIds"))

    if len(outcomes) < 2 or len(token_ids) < 2:
        return None

    outcome_to_token = {
        str(outcome).strip().lower(): str(token_id)
        for outcome, token_id in zip(outcomes, token_ids)
    }

    up_token_id = outcome_to_token.get("up") or outcome_to_token.get("yes") or str(token_ids[0])
    down_token_id = outcome_to_token.get("down") or outcome_to_token.get("no") or str(token_ids[1])

    event_slug = event_data.get("slug")
    market_slug = market.get("slug")
    start_time = event_data.get("startTime") or market.get("eventStartTime") or market.get("startDate")
    end_time = event_data.get("endDate") or market.get("endDate")
    created_at_poly = market.get("createdAt") or event_data.get("createdAt")
    fee_config = extract_market_fee_config(event_data, market)

    return {
        "polymarket_event_id": str(event_data.get("id", "")),
        "polymarket_market_id": str(market.get("id", "")),
        "condition_id": market.get("conditionId"),
        "event_slug": event_slug,
        "market_slug": market_slug,
        "title": event_data.get("title"),
        "question": market.get("question"),
        "event_url": f"https://polymarket.com/event/{event_slug}",
        "start_time": start_time,
        "start_time_local": format_local_datetime(start_time),
        "end_time": end_time,
        "end_time_local": format_local_datetime(end_time),
        "yes_token_id": up_token_id,
        "no_token_id": down_token_id,
        "outcomes": json.dumps(outcomes, ensure_ascii=False),
        "outcome_prices": market.get("outcomePrices"),
        "active": 1 if market.get("active") is True else 0,
        "closed": 1 if market.get("closed") is True else 0,
        "enable_order_book": 1 if market.get("enableOrderBook") is True else 0,
        "accepting_orders": 1 if market.get("acceptingOrders") is True else 0,
        "fees_enabled": fee_config["fees_enabled"],
        "fee_rate": fee_config["fee_rate"],
        "fee_calculation_source": fee_config["fee_calculation_source"],
        "fee_calculation_version": fee_config["fee_calculation_version"],
        "created_at_poly": created_at_poly,
        "created_at_poly_local": format_local_datetime(created_at_poly),
        "status": "open" if is_market_open(event_data, market) else "closed",
        "notes": "",
        "raw_json": json.dumps(event_data, ensure_ascii=False),
        "_is_open": is_market_open(event_data, market),
    }


def upsert_event(market_row: dict[str, Any]) -> None:
    discovered_at = now_iso()
    discovered_at_local = now_local_display()
    last_seen_at = now_iso()
    last_seen_at_local = now_local_display()

    with connect_db() as conn:
        existing = conn.execute(
            "SELECT local_event_id, discovered_at FROM events WHERE event_slug = ?",
            (market_row["event_slug"],),
        ).fetchone()

        if existing:
            conn.execute("""
                UPDATE events SET
                    polymarket_event_id = ?,
                    polymarket_market_id = ?,
                    condition_id = ?,
                    market_slug = ?,
                    title = ?,
                    question = ?,
                    event_url = ?,
                    start_time = ?,
                    start_time_local = ?,
                    end_time = ?,
                    end_time_local = ?,
                    yes_token_id = ?,
                    no_token_id = ?,
                    outcomes = ?,
                    outcome_prices = ?,
                    active = ?,
                    closed = ?,
                    enable_order_book = ?,
                    accepting_orders = ?,
                    fees_enabled = ?,
                    fee_rate = ?,
                    fee_calculation_source = ?,
                    fee_calculation_version = ?,
                    created_at_poly = ?,
                    created_at_poly_local = ?,
                    last_seen_at = ?,
                    last_seen_at_local = ?,
                    status = ?,
                    notes = ?,
                    raw_json = ?
                WHERE event_slug = ?
            """, (
                market_row["polymarket_event_id"],
                market_row["polymarket_market_id"],
                market_row["condition_id"],
                market_row["market_slug"],
                market_row["title"],
                market_row["question"],
                market_row["event_url"],
                market_row["start_time"],
                market_row["start_time_local"],
                market_row["end_time"],
                market_row["end_time_local"],
                market_row["yes_token_id"],
                market_row["no_token_id"],
                market_row["outcomes"],
                market_row["outcome_prices"],
                market_row["active"],
                market_row["closed"],
                market_row["enable_order_book"],
                market_row["accepting_orders"],
                market_row.get("fees_enabled"),
                market_row.get("fee_rate"),
                market_row.get("fee_calculation_source"),
                market_row.get("fee_calculation_version"),
                market_row["created_at_poly"],
                market_row["created_at_poly_local"],
                last_seen_at,
                last_seen_at_local,
                market_row["status"],
                market_row["notes"],
                market_row["raw_json"],
                market_row["event_slug"],
            ))
        else:
            conn.execute("""
                INSERT INTO events (
                    polymarket_event_id,
                    polymarket_market_id,
                    condition_id,
                    event_slug,
                    market_slug,
                    title,
                    question,
                    event_url,
                    start_time,
                    start_time_local,
                    end_time,
                    end_time_local,
                    yes_token_id,
                    no_token_id,
                    outcomes,
                    outcome_prices,
                    active,
                    closed,
                    enable_order_book,
                    accepting_orders,
                    fees_enabled,
                    fee_rate,
                    fee_calculation_source,
                    fee_calculation_version,
                    created_at_poly,
                    created_at_poly_local,
                    discovered_at,
                    discovered_at_local,
                    last_seen_at,
                    last_seen_at_local,
                    status,
                    notes,
                    raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                market_row["polymarket_event_id"],
                market_row["polymarket_market_id"],
                market_row["condition_id"],
                market_row["event_slug"],
                market_row["market_slug"],
                market_row["title"],
                market_row["question"],
                market_row["event_url"],
                market_row["start_time"],
                market_row["start_time_local"],
                market_row["end_time"],
                market_row["end_time_local"],
                market_row["yes_token_id"],
                market_row["no_token_id"],
                market_row["outcomes"],
                market_row["outcome_prices"],
                market_row["active"],
                market_row["closed"],
                market_row["enable_order_book"],
                market_row["accepting_orders"],
                market_row.get("fees_enabled"),
                market_row.get("fee_rate"),
                market_row.get("fee_calculation_source"),
                market_row.get("fee_calculation_version"),
                market_row["created_at_poly"],
                market_row["created_at_poly_local"],
                discovered_at,
                discovered_at_local,
                last_seen_at,
                last_seen_at_local,
                market_row["status"],
                market_row["notes"],
                market_row["raw_json"],
            ))

        close_deals_for_event_resolution(conn, market_row["event_slug"])
        conn.commit()


async def discover_open_market() -> Optional[dict[str, Any]]:
    async with httpx.AsyncClient(timeout=10) as client:
        found_open: list[dict[str, Any]] = []

        for slug in candidate_slugs():
            try:
                response = await client.get(GAMMA_URL.format(slug=slug))
                if response.status_code != 200:
                    continue

                event_data = response.json()
                market_row = extract_market(event_data)

                if not market_row:
                    continue

                upsert_event(market_row)

                if market_row["_is_open"]:
                    found_open.append(market_row)

            except Exception as e:
                log_error(f"discover slug={slug}", e)
                continue

        if not found_open:
            return None

        def sort_key(row: dict[str, Any]) -> str:
            return row.get("start_time") or row.get("end_time") or ""

        found_open.sort(key=sort_key)
        return found_open[0]


def best_bid(book: dict[str, Any]) -> Optional[float]:
    bids = book.get("bids") or []
    prices = []
    for item in bids:
        try:
            prices.append(float(item["price"]))
        except Exception:
            pass
    return max(prices) if prices else None


def best_ask(book: dict[str, Any]) -> Optional[float]:
    asks = book.get("asks") or []
    prices = []
    for item in asks:
        try:
            prices.append(float(item["price"]))
        except Exception:
            pass
    return min(prices) if prices else None


def to_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        return float(value)
    except Exception:
        return None


def calc_spread(best_ask_value: Optional[float], best_bid_value: Optional[float]) -> Optional[float]:
    if best_ask_value is None or best_bid_value is None:
        return None
    return round(best_ask_value - best_bid_value, 6)


def calc_midpoint(best_ask_value: Optional[float], best_bid_value: Optional[float]) -> Optional[float]:
    if best_ask_value is None or best_bid_value is None:
        return None
    return round((best_ask_value + best_bid_value) / 2, 6)


async def fetch_book(client: httpx.AsyncClient, token_id: str) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    try:
        response = await client.get(CLOB_BOOK_URL.format(token_id=token_id))
        if response.status_code != 200:
            return None, f"HTTP {response.status_code}: {response.text[:300]}"
        return response.json(), None
    except Exception as e:
        return None, str(e)


def coinbase_candles_url() -> str:
    if "{product_id}" in COINBASE_CANDLES_URL:
        return COINBASE_CANDLES_URL.format(product_id=COINBASE_PRODUCT_ID)
    return COINBASE_CANDLES_URL


def coinbase_candle_params(now_dt: datetime) -> dict[str, Any]:
    bucket_epoch = floor_to_epoch(now_dt, COINBASE_CANDLE_GRANULARITY_SECONDS)
    bucket_start = datetime.fromtimestamp(bucket_epoch, timezone.utc)
    return {
        "granularity": COINBASE_CANDLE_GRANULARITY_SECONDS,
        "start": iso_z(bucket_start),
        "end": iso_z(now_dt),
    }


async def fetch_coinbase_candles(
    client: httpx.AsyncClient,
    params: Optional[dict[str, Any]] = None,
) -> tuple[Optional[list[Any]], Optional[str]]:
    retry_statuses = {429, 500, 502, 503, 504}
    last_error = None
    request_params = params or {"granularity": COINBASE_CANDLE_GRANULARITY_SECONDS}

    for attempt in range(3):
        try:
            response = await client.get(
                coinbase_candles_url(),
                params=request_params,
                headers={"User-Agent": "polymarket-btc-collector/1.0"},
            )
            if response.status_code == 200:
                payload = response.json()
                if not isinstance(payload, list):
                    return None, "Coinbase response is not a list"
                return payload, None

            last_error = f"Coinbase HTTP {response.status_code}: {response.text[:160]}"
            if response.status_code not in retry_statuses:
                break
        except (httpx.TimeoutException, httpx.ConnectError, httpx.NetworkError) as exc:
            last_error = f"{type(exc).__name__}: {exc}"

        if attempt < 2:
            await asyncio.sleep(1)

    return None, last_error or "Coinbase request failed"


async def fetch_current_coinbase_candle(
    client: httpx.AsyncClient,
) -> tuple[Optional[dict[str, Any]], datetime, Optional[str], int, dict[str, Any]]:
    attempts = COINBASE_MISSING_CANDLE_RETRY_COUNT + 1
    last_error = None
    last_now = now_utc()
    last_params = coinbase_candle_params(last_now)

    for attempt in range(attempts):
        last_now = now_utc()
        last_params = coinbase_candle_params(last_now)
        candles, fetch_error = await fetch_coinbase_candles(client, last_params)
        if fetch_error or candles is None:
            last_error = fetch_error or "Coinbase candles missing"
        else:
            candle, candle_error = select_current_coinbase_candle(candles, last_now)
            if candle:
                return candle, last_now, None, attempt + 1, last_params
            last_error = candle_error or "Coinbase candle selection failed"

        if last_error == "Current Coinbase candle not found" and attempt < attempts - 1:
            await asyncio.sleep(COINBASE_MISSING_CANDLE_RETRY_DELAY_SECONDS)
            continue
        break

    return None, last_now, last_error, min(attempts, attempt + 1), last_params


def select_current_coinbase_candle(
    candles: list[Any],
    now_dt: datetime,
) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    current_bucket_epoch = floor_to_epoch(now_dt, COINBASE_CANDLE_GRANULARITY_SECONDS)

    for candle in candles:
        if not isinstance(candle, list) or len(candle) < 6:
            continue

        try:
            candle_time = int(float(candle[0]))
            volume = float(candle[5])
        except (TypeError, ValueError):
            continue

        if volume < 0:
            return None, "Coinbase candle volume is negative"

        if candle_time == current_bucket_epoch:
            return {
                "candle_start_epoch": candle_time,
                "candle_start_at": iso_from_epoch(candle_time),
                "volume_btc_cumulative": volume,
            }, None

    return None, "Current Coinbase candle not found"


def get_latest_valid_coinbase_sample(
    conn: sqlite3.Connection,
    product_id: str,
) -> Optional[sqlite3.Row]:
    conn.row_factory = sqlite3.Row
    return conn.execute("""
        SELECT
            sampled_at,
            candle_start_at,
            volume_btc_cumulative,
            status
        FROM btc_volume_log
        WHERE
            product_id = ?
            AND status IN ('success', 'baseline')
            AND volume_btc_cumulative IS NOT NULL
        ORDER BY sampled_at DESC
        LIMIT 1
    """, (product_id,)).fetchone()


def calculate_coinbase_delta(
    previous_sample: Optional[sqlite3.Row],
    candle_start_at: str,
    sampled_at_dt: datetime,
    volume_btc_cumulative: float,
) -> tuple[Optional[float], Optional[float], str, Optional[str]]:
    if not previous_sample:
        return None, None, "baseline", "no previous valid sample"

    previous_sampled_at = parse_iso_datetime(previous_sample["sampled_at"])
    previous_candle_start_at = previous_sample["candle_start_at"]
    previous_volume = to_float(previous_sample["volume_btc_cumulative"])

    if not previous_sampled_at or previous_volume is None:
        return None, None, "baseline", "previous sample is incomplete"

    seconds_since_previous = (sampled_at_dt - previous_sampled_at).total_seconds()

    if previous_candle_start_at != candle_start_at:
        return None, seconds_since_previous, "baseline", "new candle"

    if seconds_since_previous > COINBASE_MAX_DELTA_GAP_SECONDS:
        return None, seconds_since_previous, "baseline", "gap exceeded max delta threshold"

    delta = volume_btc_cumulative - previous_volume
    if delta < 0:
        return None, seconds_since_previous, "baseline", "cumulative volume decreased within same candle"

    return round(delta, 8), seconds_since_previous, "success", None


def current_active_market_snapshot() -> tuple[Optional[str], Optional[str]]:
    market = active_market
    if not market:
        return None, None
    return market.get("event_slug"), market.get("condition_id")


def insert_btc_volume_log(row: dict[str, Any]) -> bool:
    inserted = False
    with connect_db() as conn:
        cursor = conn.execute("""
            INSERT OR IGNORE INTO btc_volume_log (
                sampled_at,
                sample_bucket_at,
                candle_start_at,
                product_id,
                granularity_seconds,
                volume_btc_cumulative,
                volume_btc_delta,
                seconds_since_previous_sample,
                event_slug,
                condition_id,
                source,
                status,
                error
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row.get("sampled_at"),
            row.get("sample_bucket_at"),
            row.get("candle_start_at"),
            row.get("product_id"),
            row.get("granularity_seconds"),
            row.get("volume_btc_cumulative"),
            row.get("volume_btc_delta"),
            row.get("seconds_since_previous_sample"),
            row.get("event_slug"),
            row.get("condition_id"),
            row.get("source"),
            row.get("status"),
            row.get("error"),
        ))
        inserted = cursor.rowcount > 0
        cursor.close()
        conn.commit()
    conn.close()
    return inserted


def get_latest_coinbase_health() -> dict[str, Optional[str]]:
    with connect_db() as conn:
        conn.row_factory = sqlite3.Row
        latest = conn.execute("""
            SELECT sampled_at, status, error
            FROM btc_volume_log
            ORDER BY sampled_at DESC
            LIMIT 1
        """).fetchone()
        latest_success = conn.execute("""
            SELECT sampled_at
            FROM btc_volume_log
            WHERE status = 'success'
            ORDER BY sampled_at DESC
            LIMIT 1
        """).fetchone()

    if not latest:
        return {
            "last_sample_at": coinbase_volume_state.get("last_sample_at"),
            "last_success_at": coinbase_volume_state.get("last_success_at"),
            "status": coinbase_volume_state.get("status") or "starting",
            "last_error": coinbase_volume_state.get("last_error"),
        }

    status = "ok" if latest["status"] in ("success", "baseline") else "degraded"
    latest_dt = parse_iso_datetime(latest["sampled_at"])
    if latest_dt and (now_utc() - latest_dt).total_seconds() > COINBASE_VOLUME_POLL_INTERVAL_SECONDS * 3:
        status = "stale"

    return {
        "last_sample_at": latest["sampled_at"],
        "last_success_at": latest_success["sampled_at"] if latest_success else None,
        "status": status,
        "last_error": truncate_text(latest["error"]),
    }


async def collect_coinbase_volume_sample(client: Optional[httpx.AsyncClient] = None) -> dict[str, Any]:
    owns_client = client is None
    event_slug, condition_id = current_active_market_snapshot()

    def build_error_row(sampled_at_dt: datetime, error: str) -> dict[str, Any]:
        sample_bucket_epoch = floor_to_epoch(sampled_at_dt, COINBASE_VOLUME_POLL_INTERVAL_SECONDS)
        return {
            "sampled_at": sampled_at_dt.isoformat(),
            "sample_bucket_at": iso_from_epoch(sample_bucket_epoch),
            "candle_start_at": iso_from_epoch(floor_to_epoch(sampled_at_dt, COINBASE_CANDLE_GRANULARITY_SECONDS)),
            "product_id": COINBASE_PRODUCT_ID,
            "granularity_seconds": COINBASE_CANDLE_GRANULARITY_SECONDS,
            "volume_btc_cumulative": None,
            "volume_btc_delta": None,
            "seconds_since_previous_sample": None,
            "event_slug": event_slug,
            "condition_id": condition_id,
            "source": "coinbase_exchange",
            "status": "error",
            "error": truncate_text(error),
        }

    try:
        if owns_client:
            client = httpx.AsyncClient(timeout=COINBASE_REQUEST_TIMEOUT_SECONDS)

        assert client is not None
        candle, sampled_at_dt, fetch_error, attempts, request_params = await fetch_current_coinbase_candle(client)
        sample_bucket_epoch = floor_to_epoch(sampled_at_dt, COINBASE_VOLUME_POLL_INTERVAL_SECONDS)
        if fetch_error or candle is None:
            row = build_error_row(sampled_at_dt, f"{fetch_error or 'Coinbase candle selection failed'}; attempts={attempts}")
        else:
            with connect_db() as conn:
                previous_sample = get_latest_valid_coinbase_sample(conn, COINBASE_PRODUCT_ID)

            delta, seconds_since_previous, status, delta_error = calculate_coinbase_delta(
                previous_sample,
                candle["candle_start_at"],
                sampled_at_dt,
                candle["volume_btc_cumulative"],
            )

            row = {
                "sampled_at": sampled_at_dt.isoformat(),
                "sample_bucket_at": iso_from_epoch(sample_bucket_epoch),
                "candle_start_at": candle["candle_start_at"],
                "product_id": COINBASE_PRODUCT_ID,
                "granularity_seconds": COINBASE_CANDLE_GRANULARITY_SECONDS,
                "volume_btc_cumulative": candle["volume_btc_cumulative"],
                "volume_btc_delta": delta,
                "seconds_since_previous_sample": seconds_since_previous,
                "event_slug": event_slug,
                "condition_id": condition_id,
                "source": "coinbase_exchange",
                "status": status,
                "error": delta_error,
            }

        inserted = insert_btc_volume_log(row)
        coinbase_volume_state["last_sample_at"] = row["sampled_at"]
        coinbase_volume_state["status"] = "ok" if row["status"] in ("success", "baseline") else "degraded"
        coinbase_volume_state["last_error"] = row.get("error")
        if row["status"] == "success":
            coinbase_volume_state["last_success_at"] = row["sampled_at"]

        log_prefix = (
            "Coinbase BTC volume sample saved"
            if row["status"] == "success"
            else "Coinbase BTC volume baseline saved"
            if row["status"] == "baseline"
            else "Coinbase BTC volume collection failed"
        )
        print(
            f"{log_prefix}: candle_start={row['candle_start_at']} "
            f"cumulative_volume={row['volume_btc_cumulative']} "
            f"delta={row['volume_btc_delta']} event_slug={row['event_slug']} "
            f"inserted={inserted}",
            flush=True,
        )
        return row
    finally:
        if owns_client and client is not None:
            await client.aclose()


def empty_volume_metrics() -> dict[str, Any]:
    return {
        "up_volume_shares_10s": 0.0,
        "down_volume_shares_10s": 0.0,
        "up_volume_usdc_10s": 0.0,
        "down_volume_usdc_10s": 0.0,
        "trades_count_10s": 0,
    }


def side_ask(row: dict[str, Any] | sqlite3.Row, side: str) -> Any:
    return row["up_best_ask"] if side == "yes" else row["down_best_ask"]


def side_bid(row: dict[str, Any] | sqlite3.Row, side: str) -> Any:
    return row["up_best_bid"] if side == "yes" else row["down_best_bid"]


def count_entries(conn: sqlite3.Connection, rule_id: int, event_id: str, side: str) -> int:
    return int(conn.execute("""
        SELECT COUNT(*)
        FROM deals
        WHERE rule_id = ? AND event_id = ? AND side = ?
    """, (rule_id, event_id, side)).fetchone()[0])


def has_open_deal(conn: sqlite3.Connection, rule_id: int) -> bool:
    row = conn.execute("""
        SELECT id
        FROM deals
        WHERE rule_id = ? AND result = 'open'
        LIMIT 1
    """, (rule_id,)).fetchone()
    return row is not None


def clock_seconds(value: str) -> int:
    hour, minute, second = (int(part) for part in value.split(":"))
    return hour * 3600 + minute * 60 + second


def local_datetime_for_rule(rule: sqlite3.Row, sampled_at: Any) -> Optional[datetime]:
    sampled = parse_iso_datetime(str(sampled_at)) if sampled_at else None
    if sampled is None:
        return None
    if sampled.tzinfo is None:
        sampled = sampled.replace(tzinfo=timezone.utc)
    try:
        rule_timezone = ZoneInfo(rule["schedule_timezone"] or DEFAULT_SCHEDULE_TIMEZONE)
    except ZoneInfoNotFoundError:
        rule_timezone = ZoneInfo(DEFAULT_SCHEDULE_TIMEZONE)
    return sampled.astimezone(rule_timezone)


def matching_inactive_window(
    conn: sqlite3.Connection,
    rule: sqlite3.Row,
    current_time: Any,
) -> Optional[dict[str, Any]]:
    local_dt = local_datetime_for_rule(rule, current_time)
    if local_dt is None:
        return None
    current_day = local_dt.weekday()
    current_second = local_dt.hour * 3600 + local_dt.minute * 60 + local_dt.second
    rows = conn.execute("""
        SELECT day_of_week, start_time, end_time
        FROM rule_inactive_windows
        WHERE rule_id = ? AND status = 'active'
    """, (rule["id"],)).fetchall()
    for window in rows:
        start_second = clock_seconds(window["start_time"])
        end_second = clock_seconds(window["end_time"])
        window_day = int(window["day_of_week"])
        if start_second == end_second and current_day == window_day:
            return row_to_dict(window)
        if start_second < end_second:
            if current_day == window_day and start_second <= current_second < end_second:
                return row_to_dict(window)
            continue
        if current_day == window_day and current_second >= start_second:
            return row_to_dict(window)
        if current_day == ((window_day + 1) % 7) and current_second < end_second:
            return row_to_dict(window)
    return None


def can_rule_open_new_deal(
    conn: sqlite3.Connection,
    rule_id: int,
    current_time: Any,
) -> tuple[bool, Optional[sqlite3.Row], str, Optional[dict[str, Any]]]:
    """Authoritative DB-backed gate for every new DEMO deal."""
    rule = conn.execute("SELECT * FROM rules WHERE id = ?", (rule_id,)).fetchone()
    if rule is None:
        return False, None, "rule_not_found", None
    if rule["status"] != "active":
        return False, rule, "rule_inactive", None
    matched_window = matching_inactive_window(conn, rule, current_time)
    if matched_window is not None:
        return False, rule, "rule_in_inactive_schedule", matched_window
    if has_open_deal(conn, rule_id):
        return False, rule, "open_deal_exists", None
    return True, rule, "", None


def entry_window_allows_rule(
    conn: sqlite3.Connection,
    rule: sqlite3.Row,
    event_id: str,
    sampled_at: Any,
) -> tuple[bool, Optional[int], str]:
    start_seconds = rule["entry_window_start_seconds_before_end"]
    end_seconds = rule["entry_window_end_seconds_before_end"]
    if start_seconds is None and end_seconds is None:
        return True, None, ""

    remaining = calculate_entry_seconds_before_event_end(conn, event_id, sampled_at)
    if remaining is None:
        return False, None, "missing_event_end_time"
    if remaining <= 0:
        return False, remaining, "event_ended"
    if remaining > int(start_seconds):
        return False, remaining, "before_entry_window"
    if remaining < int(end_seconds):
        return False, remaining, "after_entry_window"
    return True, remaining, ""


def process_demo_exits(
    conn: sqlite3.Connection,
    orderbook_row: dict[str, Any],
    orderbook_log_id: int,
) -> set[int]:
    closed_rule_ids: set[int] = set()
    open_deals = conn.execute("""
        SELECT
            deals.*,
            rules.stop_loss_price,
            rules.take_profit_price
        FROM deals
        JOIN rules ON rules.id = deals.rule_id
        WHERE deals.result = 'open'
    """).fetchall()

    for deal in open_deals:
        bid = side_bid(orderbook_row, deal["side"])
        bid_price = decimal_from_db(bid)
        if bid_price is None:
            continue

        stop_loss = decimal_from_db(deal["stop_loss_price"])
        take_profit = decimal_from_db(deal["take_profit_price"])
        if stop_loss is None or take_profit is None:
            continue

        # Stop loss wins if both thresholds are considered hit in the same processing pass.
        if bid_price <= stop_loss:
            close_deal(
                conn,
                deal,
                "loss",
                "stop_loss",
                stop_loss,
                orderbook_row["sampled_at"],
                orderbook_log_id,
            )
            closed_rule_ids.add(int(deal["rule_id"]))
            print(
                f"[deals] closed id={deal['id']} rule_id={deal['rule_id']} "
                f"reason=stop_loss bid={bid} exit_price={stop_loss}",
                flush=True,
            )
            continue

        if bid_price >= take_profit:
            close_deal(
                conn,
                deal,
                "win",
                "take_profit",
                take_profit,
                orderbook_row["sampled_at"],
                orderbook_log_id,
            )
            closed_rule_ids.add(int(deal["rule_id"]))
            print(
                f"[deals] closed id={deal['id']} rule_id={deal['rule_id']} "
                f"reason=take_profit bid={bid} exit_price={take_profit}",
                flush=True,
            )

    return closed_rule_ids


def process_demo_entries(
    conn: sqlite3.Connection,
    orderbook_row: dict[str, Any],
    orderbook_log_id: int,
    closed_rule_ids: set[int],
) -> None:
    event_id = orderbook_row.get("event_slug")
    if not event_id:
        return

    rules = conn.execute("""
        SELECT *
        FROM rules
        WHERE status = 'active'
        ORDER BY id ASC
    """).fetchall()

    for rule in rules:
        rule_id = int(rule["id"])
        if rule_id in closed_rule_ids:
            continue

        if rule["eligible_after_event_id"] and rule["eligible_after_event_id"] == event_id:
            print(f"[rules] rule id={rule_id} waits for next event after {event_id}", flush=True)
            continue

        matched_window = matching_inactive_window(conn, rule, orderbook_row.get("sampled_at"))
        if matched_window is not None:
            print(
                f"[rules] entry skipped reason=rule_in_inactive_schedule rule_id={rule_id} "
                f"rule_name={rule['name']!r} event_id={event_id} timezone={rule['schedule_timezone']} "
                f"window_day={matched_window['day_of_week']} "
                f"window={matched_window['start_time']}-{matched_window['end_time']}", flush=True,
            )
            continue

        yes_match = prices_equal(side_ask(orderbook_row, "yes"), rule["entry_price"])
        no_match = prices_equal(side_ask(orderbook_row, "no"), rule["entry_price"])

        if yes_match and no_match:
            print(
                f"[deals] both sides match entry price; no deal opened "
                f"rule_id={rule_id} event_id={event_id} orderbook_log_id={orderbook_log_id}",
                flush=True,
            )
            continue

        side = "yes" if yes_match else "no" if no_match else None
        if side is None:
            continue

        if has_open_deal(conn, rule_id):
            print(f"[deals] open deal already exists for rule_id={rule_id}; entry skipped", flush=True)
            continue

        max_entries = (
            int(rule["max_yes_entries_per_event"])
            if side == "yes"
            else int(rule["max_no_entries_per_event"])
        )
        current_entries = count_entries(conn, rule_id, event_id, side)
        if current_entries >= max_entries:
            print(
                f"[deals] entry quota reached rule_id={rule_id} event_id={event_id} "
                f"side={side} quota={max_entries}",
                flush=True,
            )
            continue

        entry_allowed, entry_seconds_before_end, reason = entry_window_allows_rule(
            conn,
            rule,
            event_id,
            orderbook_row.get("sampled_at"),
        )
        if not entry_allowed:
            print(
                f"[rules] rule id={rule_id} entry window blocks entry event_id={event_id} "
                f"remaining_seconds={entry_seconds_before_end} reason={reason}",
                flush=True,
            )
            continue

        created_at = now_iso()
        fee_snapshot = extract_event_fee_snapshot(event_id, conn)
        volume_snapshot = find_entry_btc_volume_snapshot(conn, orderbook_row["sampled_at"], event_id)
        financials = calculate_demo_deal_financials(
            rule["entry_price"],
            None,
            DEMO_INVESTMENT_USD,
            DEMO_ENTRY_LIQUIDITY_ROLE,
            None,
            fee_snapshot["fee_rate"],
            fee_snapshot["fee_calculation_source"],
            fee_snapshot["fee_calculation_version"],
        )
        can_open, fresh_rule, gate_reason, final_window = can_rule_open_new_deal(
            conn, rule_id, orderbook_row.get("sampled_at")
        )
        if not can_open:
            if gate_reason == "rule_in_inactive_schedule" and final_window is not None:
                print(
                    f"[rules] final entry gate blocked reason={gate_reason} rule_id={rule_id} "
                    f"rule_name={fresh_rule['name']!r} event_id={event_id} "
                    f"timezone={fresh_rule['schedule_timezone']} "
                    f"window_day={final_window['day_of_week']} "
                    f"window={final_window['start_time']}-{final_window['end_time']}", flush=True,
                )
            else:
                print(f"[rules] final entry gate blocked rule_id={rule_id} reason={gate_reason}", flush=True)
            continue

        try:
            cursor = conn.execute("""
                INSERT INTO deals (
                    rule_id,
                    rule_name,
                    event_id,
                    side,
                    result,
                    entry_at,
                    entry_price,
                    entry_orderbook_log_id,
                    investment_usd,
                    shares,
                    entry_gross_value_usd,
                    entry_liquidity_role,
                    entry_fee_rate,
                    entry_fee_usd,
                    entry_btc_volume_log_id,
                    entry_btc_volume_sampled_at,
                    entry_btc_volume_btc_cumulative,
                    entry_btc_volume_btc_delta,
                    entry_btc_volume_status,
                    fee_calculation_source,
                    fee_calculation_version,
                    entry_seconds_before_event_end,
                    created_at,
                    updated_at
                ) VALUES (?, ?, ?, ?, 'open', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rule_id,
                rule["name"],
                event_id,
                side,
                orderbook_row["sampled_at"],
                rule["entry_price"],
                orderbook_log_id,
                decimal_to_float(financials["investment_usd"]),
                decimal_to_float(financials["shares"]),
                decimal_to_float(financials["entry_gross_value_usd"]),
                financials["entry_liquidity_role"],
                decimal_to_float(financials["entry_fee_rate"]),
                decimal_to_float(financials["entry_fee_usd"]),
                volume_snapshot["id"] if volume_snapshot else None,
                volume_snapshot["sampled_at"] if volume_snapshot else None,
                volume_snapshot["volume_btc_cumulative"] if volume_snapshot else None,
                volume_snapshot["volume_btc_delta"] if volume_snapshot else None,
                volume_snapshot["status"] if volume_snapshot else None,
                financials["fee_calculation_source"],
                financials["fee_calculation_version"],
                entry_seconds_before_end,
                created_at,
                created_at,
            ))
        except sqlite3.IntegrityError as exc:
            print(
                f"[deals] duplicate or concurrent entry prevented rule_id={rule_id} "
                f"event_id={event_id} side={side} orderbook_log_id={orderbook_log_id}: {exc}",
                flush=True,
            )
            continue

        print(
            f"[deals] opened id={cursor.lastrowid} rule_id={rule_id} "
            f"event_id={event_id} side={side} entry_price={rule['entry_price']} "
            f"orderbook_log_id={orderbook_log_id}",
            flush=True,
        )


def process_demo_trading_for_orderbook(
    conn: sqlite3.Connection,
    orderbook_row: dict[str, Any],
    orderbook_log_id: int,
) -> None:
    closed_rule_ids = process_demo_exits(conn, orderbook_row, orderbook_log_id)
    process_demo_entries(conn, orderbook_row, orderbook_log_id, closed_rule_ids)


def insert_orderbook_log(row: dict[str, Any]) -> int:
    with connect_db() as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("BEGIN IMMEDIATE")
        cursor = conn.execute("""
            INSERT INTO orderbook_log (
                sampled_at,
                sampled_at_local,
                event_slug,
                condition_id,
                up_token_id,
                down_token_id,
                up_best_ask,
                up_best_bid,
                down_best_ask,
                down_best_bid,
                up_last_trade_price,
                down_last_trade_price,
                up_spread,
                down_spread,
                up_midpoint,
                down_midpoint,
                raw_up_timestamp,
                raw_down_timestamp,
                up_volume_shares_10s,
                down_volume_shares_10s,
                up_volume_usdc_10s,
                down_volume_usdc_10s,
                trades_count_10s,
                trades_window_start,
                trades_window_start_local,
                trades_window_end,
                trades_window_end_local,
                trades_error,
                status,
                error
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row.get("sampled_at"),
            row.get("sampled_at_local"),
            row.get("event_slug"),
            row.get("condition_id"),
            row.get("up_token_id"),
            row.get("down_token_id"),
            row.get("up_best_ask"),
            row.get("up_best_bid"),
            row.get("down_best_ask"),
            row.get("down_best_bid"),
            row.get("up_last_trade_price"),
            row.get("down_last_trade_price"),
            row.get("up_spread"),
            row.get("down_spread"),
            row.get("up_midpoint"),
            row.get("down_midpoint"),
            row.get("raw_up_timestamp"),
            row.get("raw_down_timestamp"),
            row.get("up_volume_shares_10s"),
            row.get("down_volume_shares_10s"),
            row.get("up_volume_usdc_10s"),
            row.get("down_volume_usdc_10s"),
            row.get("trades_count_10s"),
            row.get("trades_window_start"),
            row.get("trades_window_start_local"),
            row.get("trades_window_end"),
            row.get("trades_window_end_local"),
            row.get("trades_error"),
            row.get("status"),
            row.get("error"),
        ))
        orderbook_log_id = int(cursor.lastrowid)
        process_demo_trading_for_orderbook(conn, row, orderbook_log_id)
        conn.commit()
        return orderbook_log_id


async def event_collector_loop() -> None:
    global active_market

    while True:
        tick_started_at = asyncio.get_running_loop().time()

        try:
            market = await discover_open_market()

            async with active_market_lock:
                if market:
                    active_market = market
                    print(f"[event] active market: {market['event_slug']}", flush=True)
                else:
                    if market_has_ended(active_market):
                        active_market = None
                    print("[event] no open market found", flush=True)
        except Exception as e:
            log_error("event loop", e)

        finally:
            await sleep_until_next_tick(EVENT_CHECK_INTERVAL_SECONDS, tick_started_at)


async def orderbook_collector_loop() -> None:
    global active_market

    while True:
        tick_started_at = asyncio.get_running_loop().time()

        try:
            async with active_market_lock:
                market = dict(active_market) if active_market else None

            if not market:
                continue

            if market_has_ended(market):
                async with active_market_lock:
                    active_market = None
                print(f"[book] market ended: {market.get('event_slug')}", flush=True)
                continue

            up_token_id = market["yes_token_id"]
            down_token_id = market["no_token_id"]
            condition_id = market.get("condition_id")

            async with httpx.AsyncClient(timeout=10) as client:
                up_book, up_error = await fetch_book(client, up_token_id)
                down_book, down_error = await fetch_book(client, down_token_id)

            errors = []
            if up_error:
                errors.append(f"up_error={up_error}")
            if down_error:
                errors.append(f"down_error={down_error}")

            up_best_bid = best_bid(up_book) if up_book else None
            up_best_ask = best_ask(up_book) if up_book else None
            down_best_bid = best_bid(down_book) if down_book else None
            down_best_ask = best_ask(down_book) if down_book else None
            volume_metrics = empty_volume_metrics()

            if not errors:
                status = "success"
            elif up_book or down_book:
                status = "partial_error"
            else:
                status = "error"

            row = {
                "sampled_at": now_iso(),
                "sampled_at_local": now_local_display(),
                "event_slug": market.get("event_slug"),
                "condition_id": condition_id,
                "up_token_id": up_token_id,
                "down_token_id": down_token_id,
                "up_best_ask": up_best_ask,
                "up_best_bid": up_best_bid,
                "down_best_ask": down_best_ask,
                "down_best_bid": down_best_bid,
                "up_last_trade_price": to_float(up_book.get("last_trade_price")) if up_book else None,
                "down_last_trade_price": to_float(down_book.get("last_trade_price")) if down_book else None,
                "up_spread": calc_spread(up_best_ask, up_best_bid),
                "down_spread": calc_spread(down_best_ask, down_best_bid),
                "up_midpoint": calc_midpoint(up_best_ask, up_best_bid),
                "down_midpoint": calc_midpoint(down_best_ask, down_best_bid),
                "raw_up_timestamp": str(up_book.get("timestamp")) if up_book else None,
                "raw_down_timestamp": str(down_book.get("timestamp")) if down_book else None,
                "up_volume_shares_10s": volume_metrics["up_volume_shares_10s"],
                "down_volume_shares_10s": volume_metrics["down_volume_shares_10s"],
                "up_volume_usdc_10s": volume_metrics["up_volume_usdc_10s"],
                "down_volume_usdc_10s": volume_metrics["down_volume_usdc_10s"],
                "trades_count_10s": volume_metrics["trades_count_10s"],
                "trades_window_start": None,
                "trades_window_start_local": None,
                "trades_window_end": None,
                "trades_window_end_local": None,
                "trades_error": None,
                "status": status,
                "error": " | ".join(errors) if errors else None,
            }

            insert_orderbook_log(row)

            print(
                f"[book] {row['event_slug']} status={status} "
                f"up={up_best_bid}/{up_best_ask} down={down_best_bid}/{down_best_ask} "
                "volume_collection=disabled",
                flush=True,
            )
        except Exception as e:
            log_error("book loop", e)

        finally:
            await sleep_until_next_tick(BOOK_CHECK_INTERVAL_SECONDS, tick_started_at)


async def coinbase_volume_collector_loop() -> None:
    while True:
        tick_started_at = asyncio.get_running_loop().time()

        try:
            await collect_coinbase_volume_sample()
        except Exception as e:
            coinbase_volume_state["status"] = "degraded"
            coinbase_volume_state["last_error"] = truncate_text(f"{type(e).__name__}: {e}")
            log_error("coinbase volume loop", e)
        finally:
            await sleep_until_next_tick(COINBASE_VOLUME_POLL_INTERVAL_SECONDS, tick_started_at)


@app.on_event("startup")
async def startup() -> None:
    init_db()
    asyncio.create_task(event_collector_loop())
    asyncio.create_task(orderbook_collector_loop())
    asyncio.create_task(coinbase_volume_collector_loop())


def render_table(rows: list[sqlite3.Row]) -> str:
    if not rows:
        return "<p>No data yet.</p>"

    headers = rows[0].keys()

    thead = "".join(f"<th>{html.escape(str(h))}</th>" for h in headers)

    body = ""
    for row in rows:
        cells = ""
        for h in headers:
            value = row[h]
            text = "" if value is None else str(value)
            cells += f"<td>{html.escape(text)}</td>"
        body += f"<tr>{cells}</tr>"

    return f"""
    <div class="table-wrap">
      <table>
        <thead><tr>{thead}</tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
    """


def display_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, float):
        return f"{value:.8f}".rstrip("0").rstrip(".")
    return str(value)


def format_storage_size(size_bytes: int) -> str:
    size = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024 or unit == "TB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} {unit}"
        size /= 1024
    return f"{size:.1f} TB"


def render_storage_status() -> str:
    usage = shutil.disk_usage(APP_DIR)
    used_percent = (usage.used / usage.total * 100) if usage.total else 0
    return (
        f"Storage: {format_storage_size(usage.free)} free of "
        f"{format_storage_size(usage.total)} ({used_percent:.1f}% used)"
    )


def render_btc_volume_table(rows: list[sqlite3.Row]) -> str:
    if not rows:
        return "<p>No Coinbase BTC volume data yet.</p>"

    headers = [
        "Sampled At",
        "Candle Start",
        "Product",
        "Cumulative Volume BTC",
        "Volume Delta BTC",
        "Seconds Since Previous Sample",
        "Event Slug",
        "Status",
        "Error",
    ]
    fields = [
        "sampled_at",
        "candle_start_at",
        "product_id",
        "volume_btc_cumulative",
        "volume_btc_delta",
        "seconds_since_previous_sample",
        "event_slug",
        "status",
        "error",
    ]

    thead = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body = ""
    for row in rows:
        cells = ""
        for field in fields:
            value = display_value(row[field])
            if field == "error" and len(value) > 140:
                value = f"{value[:137]}..."
            cells += f"<td>{html.escape(value)}</td>"
        body += f"<tr>{cells}</tr>"

    return f"""
    <div class="table-wrap">
      <table>
        <thead><tr>{thead}</tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
    """


def render_btc_volume_summary(summary: Optional[sqlite3.Row], health_row: dict[str, Optional[str]]) -> str:
    latest_cumulative = display_value(summary["volume_btc_cumulative"]) if summary else "-"
    latest_delta = display_value(summary["volume_btc_delta"]) if summary else "-"
    last_success = display_value(health_row.get("last_success_at"))
    status = display_value(health_row.get("status"))

    return f"""
    <p class="muted">
        Latest cumulative volume: {html.escape(latest_cumulative)} BTC |
        Latest delta: {html.escape(latest_delta)} BTC |
        Last successful sample: {html.escape(last_success)} |
        Collector status: {html.escape(status)}
    </p>
    """


def format_money(value: Optional[float]) -> str:
    if value is None:
        return "-"
    sign = "-" if value < 0 else ""
    return f"{sign}${abs(value):.2f}"


def format_percent(value: Optional[float]) -> str:
    if value is None:
        return "-"
    return f"{value:.2f}%"


def format_factor(value: Optional[float], infinite: bool = False) -> str:
    if infinite:
        return "∞"
    if value is None:
        return "-"
    return f"{value:.2f}"


def max_drawdown(values: list[float]) -> float:
    peak = 0.0
    worst = 0.0
    equity = 0.0
    for value in values:
        equity += value
        peak = max(peak, equity)
        worst = max(worst, peak - equity)
    return worst


def longest_result_streak(results: list[str], target: str) -> int:
    current = 0
    longest = 0
    for result in results:
        if result == target:
            current += 1
            longest = max(longest, current)
        else:
            current = 0
    return longest


CUSTOM_DATETIME_FORMAT = "%d/%m/%Y %H:%M"


def normalize_dashboard_range(value: Any) -> str:
    selected = str(value or "all").strip().lower()
    return selected if selected in {"today", "7d", "30d", "custom", "all"} else "all"


def dashboard_range_options() -> list[tuple[str, str]]:
    return [
        ("all", "כל התקופה"),
        ("today", "היום"),
        ("7d", "7 ימים אחרונים"),
        ("30d", "30 ימים אחרונים"),
        ("custom", "טווח מותאם"),
    ]


def parse_custom_dashboard_datetime(value: Any) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.strptime(text, CUSTOM_DATETIME_FORMAT)
    except ValueError:
        return None
    return parsed.replace(tzinfo=LOCAL_TIMEZONE).astimezone(timezone.utc)


def dashboard_datetime_input_value(value: Any) -> str:
    parsed = parse_custom_dashboard_datetime(value)
    if parsed:
        return parsed.astimezone(LOCAL_TIMEZONE).strftime(CUSTOM_DATETIME_FORMAT)
    return str(value or "").strip()


def dashboard_range_label(value: Any, custom_from: Any = None, custom_to: Any = None) -> str:
    selected = normalize_dashboard_range(value)
    if selected == "custom":
        start_text = dashboard_datetime_input_value(custom_from)
        end_text = dashboard_datetime_input_value(custom_to)
        if start_text and end_text:
            return f"{start_text} עד {end_text}"
        if start_text:
            return f"מ־{start_text}"
        if end_text:
            return f"עד {end_text}"
        return "טווח מותאם"
    return dict(dashboard_range_options()).get(selected, "כל התקופה")


def dashboard_range_bounds(value: Any, custom_from: Any = None, custom_to: Any = None) -> tuple[Optional[str], Optional[str]]:
    selected = normalize_dashboard_range(value)
    now = now_utc()
    if selected == "today":
        return now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(), None
    if selected == "7d":
        return (now - timedelta(days=7)).isoformat(), None
    if selected == "30d":
        return (now - timedelta(days=30)).isoformat(), None
    if selected == "custom":
        start = parse_custom_dashboard_datetime(custom_from)
        end = parse_custom_dashboard_datetime(custom_to)
        return (
            start.isoformat() if start else None,
            end.isoformat() if end else None,
        )
    return None, None


def dashboard_range_start(value: Any, custom_from: Any = None) -> Optional[str]:
    start, _ = dashboard_range_bounds(value, custom_from, None)
    return start


def time_filter_sql(
    column: str,
    dashboard_range: Any,
    custom_from: Any = None,
    custom_to: Any = None,
) -> tuple[str, tuple[Any, ...]]:
    start, end = dashboard_range_bounds(dashboard_range, custom_from, custom_to)
    clauses: list[str] = []
    params: list[Any] = []
    if start:
        clauses.append(f"{column} >= ?")
        params.append(start)
    if end:
        clauses.append(f"{column} <= ?")
        params.append(end)
    if not clauses:
        return "1 = 1", ()
    return " AND ".join(clauses), tuple(params)


def parse_rule_filter(value: Any) -> list[int]:
    text = str(value or "all").strip().lower()
    if not text or text == "all":
        return []
    rule_ids: list[int] = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            rule_id = int(part)
        except ValueError:
            continue
        if rule_id > 0 and rule_id not in rule_ids:
            rule_ids.append(rule_id)
    return rule_ids


def normalize_rule_filter(value: Any) -> str:
    rule_ids = parse_rule_filter(value)
    return ",".join(str(rule_id) for rule_id in rule_ids) if rule_ids else "all"


def rule_filter_sql(column: str, rule_filter: Any) -> tuple[str, tuple[Any, ...]]:
    rule_ids = parse_rule_filter(rule_filter)
    if not rule_ids:
        return "1 = 1", ()
    placeholders = ",".join("?" for _ in rule_ids)
    return f"{column} IN ({placeholders})", tuple(rule_ids)


def dashboard_rule_label(rule_filter: Any) -> str:
    rule_ids = parse_rule_filter(rule_filter)
    if not rule_ids:
        return "כל החוקים"
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT id, name FROM rules WHERE id IN ({','.join('?' for _ in rule_ids)}) ORDER BY id ASC",
            tuple(rule_ids),
        ).fetchall()
    names = [f"{row['name']} (#{row['id']})" for row in rows]
    return ", ".join(names) if names else ", ".join(f"Rule #{rule_id}" for rule_id in rule_ids)


def deal_matches_range(deal: sqlite3.Row, dashboard_range: Any, custom_from: Any = None, custom_to: Any = None) -> bool:
    start, end = dashboard_range_bounds(dashboard_range, custom_from, custom_to)
    exit_at = deal["exit_at"] if "exit_at" in deal.keys() else None
    exit_dt = parse_iso_datetime(exit_at)
    start_dt = parse_iso_datetime(start)
    end_dt = parse_iso_datetime(end)
    if not exit_dt:
        return False
    if start_dt and exit_dt < start_dt:
        return False
    if end_dt and exit_dt > end_dt:
        return False
    return True


def load_dashboard_overview(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
    conn: Optional[sqlite3.Connection] = None,
) -> dict[str, Any]:
    investment = normalize_investment_usd(investment_usd)
    entry_filter, entry_params = time_filter_sql("entry_at", dashboard_range, custom_from, custom_to)
    exit_filter, exit_params = time_filter_sql("exit_at", dashboard_range, custom_from, custom_to)
    deal_rule_filter, deal_rule_params = rule_filter_sql("rule_id", rule_filter)
    active_rule_filter, active_rule_params = rule_filter_sql("id", rule_filter)
    owns_conn = conn is None
    if conn is None:
        conn = get_conn()
    try:
        total_deals = int(conn.execute(
            f"SELECT COUNT(*) FROM deals WHERE {entry_filter} AND {deal_rule_filter}",
            entry_params + deal_rule_params,
        ).fetchone()[0])
        open_deals = int(conn.execute(
            f"SELECT COUNT(*) FROM deals WHERE result = 'open' AND {deal_rule_filter}",
            deal_rule_params,
        ).fetchone()[0])
        btc_volume_gt_6_deals = int(conn.execute(f"""
            SELECT COUNT(*)
            FROM deals
            WHERE entry_btc_volume_btc_delta > 6
              AND {entry_filter}
              AND {deal_rule_filter}
        """, entry_params + deal_rule_params).fetchone()[0])
        missing_btc_volume_snapshot_deals = int(conn.execute(f"""
            SELECT COUNT(*)
            FROM deals
            WHERE entry_btc_volume_log_id IS NULL
              AND {entry_filter}
              AND {deal_rule_filter}
        """, entry_params + deal_rule_params).fetchone()[0])
        active_rules = int(conn.execute(
            f"SELECT COUNT(*) FROM rules WHERE status = 'active' AND {active_rule_filter}",
            active_rule_params,
        ).fetchone()[0])
        last_orderbook_sample = conn.execute("SELECT MAX(sampled_at) FROM orderbook_log").fetchone()[0]
        orderbook_errors = int(conn.execute("""
            SELECT COUNT(*)
            FROM orderbook_log
            WHERE status IS NOT NULL AND status != 'success'
        """).fetchone()[0])
        closed_deals = conn.execute("""
            SELECT *
            FROM deals
            WHERE result IN ('win', 'loss') AND exit_price IS NOT NULL
                AND """ + exit_filter + """
                AND """ + deal_rule_filter + """
            ORDER BY exit_at ASC, id ASC
        """, exit_params + deal_rule_params).fetchall()
    finally:
        if owns_conn:
            conn.close()

    pnl_values: list[float] = []
    roi_values: list[float] = []
    gross_pnl_values: list[float] = []
    total_fees = 0.0
    entry_fees = 0.0
    exit_fees = 0.0
    fee_charged_deals = 0
    taker_fills = 0
    maker_fills = 0
    results: list[str] = []
    per_rule: dict[int, dict[str, Any]] = {}

    for deal in closed_deals:
        try:
            financials = deal_financials_from_row(deal, investment)
        except (InvalidOperation, ValueError, TypeError, ZeroDivisionError):
            continue

        pnl_usd = financials["net_pnl_usd"]
        roi_percent = financials["net_roi_percent"]
        pnl_values.append(pnl_usd)
        roi_values.append(roi_percent)
        gross_pnl_values.append(financials["gross_pnl_usd"])
        total_fees += financials["total_fees_usd"]
        fee_charged_deals += 1 if financials["total_fees_usd"] > 0 else 0
        entry_fees += financials["entry_fee_usd"]
        exit_fees += financials["exit_fee_usd"]
        taker_fills += 1 if deal["entry_liquidity_role"] == "TAKER" else 0
        taker_fills += 1 if deal["exit_liquidity_role"] == "TAKER" else 0
        maker_fills += 1 if deal["entry_liquidity_role"] == "MAKER" else 0
        maker_fills += 1 if deal["exit_liquidity_role"] == "MAKER" else 0
        results.append(str(deal["result"]))

        rule_id = int(deal["rule_id"])
        rule_summary = per_rule.setdefault(rule_id, {
            "rule_id": rule_id,
            "rule_name": deal["rule_name"] or f"Rule {rule_id}",
            "deals": 0,
            "wins": 0,
            "losses": 0,
            "pnl_usd": 0.0,
        })
        rule_summary["deals"] += 1
        rule_summary["wins"] += 1 if deal["result"] == "win" else 0
        rule_summary["losses"] += 1 if deal["result"] == "loss" else 0
        rule_summary["pnl_usd"] += pnl_usd

    closed_count = len(pnl_values)
    wins = results.count("win")
    losses = results.count("loss")
    net_pnl = sum(pnl_values)
    gross_pnl = sum(gross_pnl_values)
    gross_profit = sum(value for value in pnl_values if value > 0)
    gross_loss_abs = abs(sum(value for value in pnl_values if value < 0))
    profit_factor = None if gross_loss_abs == 0 else gross_profit / gross_loss_abs
    profit_factor_infinite = gross_loss_abs == 0 and gross_profit > 0
    best_rule = max(per_rule.values(), key=lambda item: item["pnl_usd"], default=None)
    worst_rule = min(per_rule.values(), key=lambda item: item["pnl_usd"], default=None)

    return {
        "investment_usd": float(investment),
        "range": normalize_dashboard_range(dashboard_range),
        "range_label": dashboard_range_label(dashboard_range, custom_from, custom_to),
        "rule_filter": normalize_rule_filter(rule_filter),
        "rule_label": dashboard_rule_label(rule_filter),
        "custom_from": dashboard_datetime_input_value(custom_from),
        "custom_to": dashboard_datetime_input_value(custom_to),
        "total_deals": total_deals,
        "closed_deals": closed_count,
        "open_deals": open_deals,
        "btc_volume_gt_6_deals": btc_volume_gt_6_deals,
        "missing_btc_volume_snapshot_deals": missing_btc_volume_snapshot_deals,
        "active_rules": active_rules,
        "wins": wins,
        "losses": losses,
        "win_rate": (wins / closed_count * 100) if closed_count else None,
        "gross_pnl_usd": gross_pnl,
        "net_pnl_usd": net_pnl,
        "total_fees_usd": total_fees,
        "entry_fees_usd": entry_fees,
        "exit_fees_usd": exit_fees,
        "avg_fee_usd": (total_fees / closed_count) if closed_count else None,
        "fees_to_investment_percent": (total_fees / (float(investment) * closed_count) * 100) if closed_count else None,
        "fees_to_gross_profit_percent": (total_fees / gross_pnl * 100) if gross_pnl > 0 else None,
        "fee_charged_deals": fee_charged_deals,
        "maker_fills": maker_fills,
        "taker_fills": taker_fills,
        "estimated_maker_savings_usd": 0.0,
        "avg_pnl_usd": (net_pnl / closed_count) if closed_count else None,
        "avg_roi_percent": (sum(roi_values) / closed_count) if closed_count else None,
        "profit_factor": profit_factor,
        "profit_factor_infinite": profit_factor_infinite,
        "max_drawdown_usd": max_drawdown(pnl_values),
        "longest_loss_streak": longest_result_streak(results, "loss"),
        "longest_win_streak": longest_result_streak(results, "win"),
        "best_rule": best_rule,
        "worst_rule": worst_rule,
        "last_orderbook_sample": last_orderbook_sample,
        "orderbook_errors": orderbook_errors,
    }


def load_rules_performance(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> list[dict[str, Any]]:
    investment = normalize_investment_usd(investment_usd)
    exit_filter, exit_params = time_filter_sql("exit_at", dashboard_range, custom_from, custom_to)
    rules_filter, rules_params = rule_filter_sql("id", rule_filter)
    deals_rule_filter, deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        rules = conn.execute(f"""
            SELECT
                id,
                name,
                status,
                entry_price,
                stop_loss_price,
                take_profit_price
            FROM rules
            WHERE {rules_filter}
            ORDER BY id ASC
        """, rules_params).fetchall()
        deals = conn.execute("""
            SELECT *
            FROM deals
            WHERE """ + deals_rule_filter + """
              AND (result = 'open' OR (result IN ('win', 'loss') AND exit_price IS NOT NULL AND """ + exit_filter + """))
            ORDER BY rule_id ASC, exit_at ASC, id ASC
        """, deals_rule_params + exit_params).fetchall()

    performance: dict[int, dict[str, Any]] = {}
    for rule in rules:
        rule_id = int(rule["id"])
        performance[rule_id] = {
            "rule_id": rule_id,
            "rule_name": rule["name"],
            "status": rule["status"],
            "entry_price": rule["entry_price"],
            "stop_loss_price": rule["stop_loss_price"],
            "take_profit_price": rule["take_profit_price"],
            "closed_deals": 0,
            "open_deals": 0,
            "wins": 0,
            "losses": 0,
            "net_pnl_usd": 0.0,
            "avg_pnl_usd": None,
            "avg_roi_percent": None,
            "win_rate": None,
            "profit_factor": None,
            "profit_factor_infinite": False,
            "max_drawdown_usd": 0.0,
            "expectancy_usd": None,
            "_pnl_values": [],
            "_roi_values": [],
        }

    for deal in deals:
        rule_id = int(deal["rule_id"])
        row = performance.get(rule_id)
        if row is None:
            row = {
                "rule_id": rule_id,
                "rule_name": f"Missing rule {rule_id}",
                "status": "missing",
                "entry_price": None,
                "stop_loss_price": None,
                "take_profit_price": None,
                "closed_deals": 0,
                "open_deals": 0,
                "wins": 0,
                "losses": 0,
                "net_pnl_usd": 0.0,
                "avg_pnl_usd": None,
                "avg_roi_percent": None,
                "win_rate": None,
                "profit_factor": None,
                "profit_factor_infinite": False,
                "max_drawdown_usd": 0.0,
                "expectancy_usd": None,
                "_pnl_values": [],
                "_roi_values": [],
            }
            performance[rule_id] = row

        if deal["result"] == "open":
            row["open_deals"] += 1
            continue

        if deal["result"] not in ("win", "loss") or deal["exit_price"] is None:
            continue

        try:
            financials = deal_financials_from_row(deal, investment)
        except (InvalidOperation, ValueError, TypeError, ZeroDivisionError):
            continue
        pnl_usd = financials["net_pnl_usd"]
        roi_percent = financials["net_roi_percent"]

        row["closed_deals"] += 1
        row["wins"] += 1 if deal["result"] == "win" else 0
        row["losses"] += 1 if deal["result"] == "loss" else 0
        row["net_pnl_usd"] += pnl_usd
        row["_pnl_values"].append(pnl_usd)
        row["_roi_values"].append(roi_percent)

    rows = list(performance.values())
    for row in rows:
        closed_deals = row["closed_deals"]
        pnl_values = row.pop("_pnl_values")
        roi_values = row.pop("_roi_values")
        gross_profit = sum(value for value in pnl_values if value > 0)
        gross_loss_abs = abs(sum(value for value in pnl_values if value < 0))

        if closed_deals:
            row["win_rate"] = row["wins"] / closed_deals * 100
            row["avg_pnl_usd"] = row["net_pnl_usd"] / closed_deals
            row["avg_roi_percent"] = sum(roi_values) / closed_deals
            row["expectancy_usd"] = row["avg_pnl_usd"]
            row["max_drawdown_usd"] = max_drawdown(pnl_values)

        if gross_loss_abs == 0:
            row["profit_factor_infinite"] = gross_profit > 0
            row["profit_factor"] = None
        else:
            row["profit_factor"] = gross_profit / gross_loss_abs

    rows.sort(key=lambda item: (item["net_pnl_usd"], item["closed_deals"], -item["rule_id"]), reverse=True)
    return rows


def load_risk_snapshot(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> dict[str, Any]:
    investment = normalize_investment_usd(investment_usd)
    exit_filter, exit_params = time_filter_sql("exit_at", dashboard_range, custom_from, custom_to)
    deals_rule_filter, deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        closed_deals = conn.execute("""
            SELECT *
            FROM deals
            WHERE result IN ('win', 'loss') AND exit_price IS NOT NULL
                AND """ + exit_filter + """
                AND """ + deals_rule_filter + """
            ORDER BY exit_at ASC, id ASC
        """, exit_params + deals_rule_params).fetchall()

    equity = 0.0
    peak = 0.0
    worst_drawdown = 0.0
    worst_drawdown_after = None
    best_deal = None
    worst_deal = None
    pnl_values: list[float] = []
    results: list[str] = []
    exit_reasons: dict[str, dict[str, Any]] = {}

    for deal in closed_deals:
        try:
            financials = deal_financials_from_row(deal, investment)
        except (InvalidOperation, ValueError, TypeError, ZeroDivisionError):
            continue
        pnl_usd = financials["net_pnl_usd"]
        roi_percent = financials["net_roi_percent"]

        pnl_values.append(pnl_usd)
        results.append(str(deal["result"]))
        equity += pnl_usd
        peak = max(peak, equity)
        drawdown = peak - equity
        if drawdown > worst_drawdown:
            worst_drawdown = drawdown
            worst_drawdown_after = deal["exit_at"]

        deal_summary = {
            "deal_id": deal["id"],
            "rule_name": deal["rule_name"] or f"Rule {deal['rule_id']}",
            "event_id": deal["event_id"],
            "side": deal["side"],
            "result": deal["result"],
            "exit_reason": deal["exit_reason"] or "unknown",
            "exit_at": deal["exit_at"],
            "pnl_usd": pnl_usd,
            "roi_percent": roi_percent,
        }
        if best_deal is None or pnl_usd > best_deal["pnl_usd"]:
            best_deal = deal_summary
        if worst_deal is None or pnl_usd < worst_deal["pnl_usd"]:
            worst_deal = deal_summary

        reason = deal_summary["exit_reason"]
        reason_summary = exit_reasons.setdefault(reason, {
            "exit_reason": reason,
            "deals": 0,
            "wins": 0,
            "losses": 0,
            "pnl_usd": 0.0,
        })
        reason_summary["deals"] += 1
        reason_summary["wins"] += 1 if deal["result"] == "win" else 0
        reason_summary["losses"] += 1 if deal["result"] == "loss" else 0
        reason_summary["pnl_usd"] += pnl_usd

    closed_count = len(pnl_values)
    negative_deals = [value for value in pnl_values if value < 0]
    positive_deals = [value for value in pnl_values if value > 0]
    exit_reason_rows = sorted(
        exit_reasons.values(),
        key=lambda item: (item["deals"], item["pnl_usd"]),
        reverse=True,
    )
    for row in exit_reason_rows:
        row["win_rate"] = (row["wins"] / row["deals"] * 100) if row["deals"] else None

    return {
        "closed_deals": closed_count,
        "ending_equity_usd": equity,
        "peak_equity_usd": peak,
        "max_drawdown_usd": worst_drawdown,
        "max_drawdown_after": worst_drawdown_after,
        "longest_loss_streak": longest_result_streak(results, "loss"),
        "longest_win_streak": longest_result_streak(results, "win"),
        "best_deal": best_deal,
        "worst_deal": worst_deal,
        "avg_loss_usd": (sum(negative_deals) / len(negative_deals)) if negative_deals else None,
        "avg_win_usd": (sum(positive_deals) / len(positive_deals)) if positive_deals else None,
        "loss_deals": len(negative_deals),
        "win_deals": len(positive_deals),
        "exit_reasons": exit_reason_rows,
    }


def price_bucket_label(entry_price: Any) -> str:
    price = decimal_from_db(entry_price)
    if price is None:
        return "unknown"
    if price < Decimal("0.60"):
        return "< 0.60"
    if price < Decimal("0.70"):
        return "0.60-0.69"
    if price < Decimal("0.80"):
        return "0.70-0.79"
    if price < Decimal("0.90"):
        return "0.80-0.89"
    return "0.90+"


def add_condition_result(group: dict[str, Any], result: str, pnl_usd: float, roi_percent: float) -> None:
    group["closed_deals"] += 1
    group["wins"] += 1 if result == "win" else 0
    group["losses"] += 1 if result == "loss" else 0
    group["net_pnl_usd"] += pnl_usd
    group["_roi_values"].append(roi_percent)


def finalize_condition_group(group: dict[str, Any]) -> dict[str, Any]:
    closed_deals = group["closed_deals"]
    roi_values = group.pop("_roi_values")
    group["win_rate"] = (group["wins"] / closed_deals * 100) if closed_deals else None
    group["avg_roi_percent"] = (sum(roi_values) / closed_deals) if closed_deals else None
    group["avg_pnl_usd"] = (group["net_pnl_usd"] / closed_deals) if closed_deals else None
    return group


def load_market_conditions(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> dict[str, list[dict[str, Any]]]:
    investment = normalize_investment_usd(investment_usd)
    exit_filter, exit_params = time_filter_sql("exit_at", dashboard_range, custom_from, custom_to)
    deals_rule_filter, deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        closed_deals = conn.execute("""
            SELECT *
            FROM deals
            WHERE result IN ('win', 'loss') AND exit_price IS NOT NULL
                AND """ + exit_filter + """
                AND """ + deals_rule_filter + """
            ORDER BY id ASC
        """, exit_params + deals_rule_params).fetchall()

    side_groups: dict[str, dict[str, Any]] = {}
    price_groups: dict[str, dict[str, Any]] = {}

    for deal in closed_deals:
        try:
            financials = deal_financials_from_row(deal, investment)
        except (InvalidOperation, ValueError, TypeError, ZeroDivisionError):
            continue
        pnl_usd = financials["net_pnl_usd"]
        roi_percent = financials["net_roi_percent"]

        side = str(deal["side"] or "unknown").upper()
        side_group = side_groups.setdefault(side, {
            "label": side,
            "closed_deals": 0,
            "wins": 0,
            "losses": 0,
            "win_rate": None,
            "net_pnl_usd": 0.0,
            "avg_pnl_usd": None,
            "avg_roi_percent": None,
            "_roi_values": [],
        })
        add_condition_result(side_group, str(deal["result"]), pnl_usd, roi_percent)

        bucket = price_bucket_label(deal["entry_price"])
        price_group = price_groups.setdefault(bucket, {
            "label": bucket,
            "closed_deals": 0,
            "wins": 0,
            "losses": 0,
            "win_rate": None,
            "net_pnl_usd": 0.0,
            "avg_pnl_usd": None,
            "avg_roi_percent": None,
            "_roi_values": [],
        })
        add_condition_result(price_group, str(deal["result"]), pnl_usd, roi_percent)

    price_order = {
        "< 0.60": 0,
        "0.60-0.69": 1,
        "0.70-0.79": 2,
        "0.80-0.89": 3,
        "0.90+": 4,
        "unknown": 5,
    }
    return {
        "by_side": sorted(
            [finalize_condition_group(group) for group in side_groups.values()],
            key=lambda item: item["label"],
        ),
        "by_entry_price": sorted(
            [finalize_condition_group(group) for group in price_groups.values()],
            key=lambda item: price_order.get(item["label"], 99),
        ),
    }


def load_system_health_snapshot() -> dict[str, Any]:
    with get_conn() as conn:
        events_count = int(conn.execute("SELECT COUNT(*) FROM events").fetchone()[0])
        orderbook_count = int(conn.execute("SELECT COUNT(*) FROM orderbook_log").fetchone()[0])
        btc_volume_count = int(conn.execute("SELECT COUNT(*) FROM btc_volume_log").fetchone()[0])
        rules_count = int(conn.execute("SELECT COUNT(*) FROM rules").fetchone()[0])
        active_rules = int(conn.execute("SELECT COUNT(*) FROM rules WHERE status = 'active'").fetchone()[0])
        deals_count = int(conn.execute("SELECT COUNT(*) FROM deals").fetchone()[0])
        open_deals = int(conn.execute("SELECT COUNT(*) FROM deals WHERE result = 'open'").fetchone()[0])
        latest_event = conn.execute("SELECT MAX(last_seen_at) FROM events").fetchone()[0]
        latest_orderbook = conn.execute("SELECT MAX(sampled_at) FROM orderbook_log").fetchone()[0]
        latest_btc_volume = conn.execute("SELECT MAX(sampled_at) FROM btc_volume_log").fetchone()[0]
        orderbook_issues = int(conn.execute("""
            SELECT COUNT(*)
            FROM orderbook_log
            WHERE status IS NOT NULL AND status != 'success'
        """).fetchone()[0])
        btc_volume_issues = int(conn.execute("""
            SELECT COUNT(*)
            FROM btc_volume_log
            WHERE status = 'error'
        """).fetchone()[0])

    db_exists = DB_PATH.exists()
    db_size_bytes = DB_PATH.stat().st_size if db_exists else 0
    coinbase_health = get_latest_coinbase_health()
    issue_count = orderbook_issues + btc_volume_issues
    status = "ok"
    if issue_count:
        status = "degraded"
    if not db_exists:
        status = "db_missing"

    return {
        "status": status,
        "db_path": str(DB_PATH),
        "db_exists": db_exists,
        "db_size": format_storage_size(db_size_bytes),
        "events_count": events_count,
        "orderbook_count": orderbook_count,
        "btc_volume_count": btc_volume_count,
        "rules_count": rules_count,
        "active_rules": active_rules,
        "deals_count": deals_count,
        "open_deals": open_deals,
        "latest_event": latest_event,
        "latest_orderbook": latest_orderbook,
        "latest_btc_volume": latest_btc_volume,
        "orderbook_issues": orderbook_issues,
        "btc_volume_issues": btc_volume_issues,
        "coinbase_status": coinbase_health.get("status"),
        "coinbase_last_success": coinbase_health.get("last_success_at"),
        "coinbase_last_error": coinbase_health.get("last_error"),
    }


def local_date_key(value: Any) -> str:
    dt = parse_iso_datetime(str(value)) if value else None
    if not dt:
        return "unknown"
    return dt.astimezone(LOCAL_TIMEZONE).strftime("%Y-%m-%d")


def local_time_label(value: Any) -> str:
    dt = parse_iso_datetime(str(value)) if value else None
    if not dt:
        return "unknown"
    return dt.astimezone(LOCAL_TIMEZONE).strftime("%d/%m %H:%M")


def load_time_trends(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> list[dict[str, Any]]:
    investment = normalize_investment_usd(investment_usd)
    exit_filter, exit_params = time_filter_sql("exit_at", dashboard_range, custom_from, custom_to)
    deals_rule_filter, deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        closed_deals = conn.execute("""
            SELECT *
            FROM deals
            WHERE result IN ('win', 'loss') AND exit_price IS NOT NULL
                AND """ + exit_filter + """
                AND """ + deals_rule_filter + """
            ORDER BY exit_at ASC, id ASC
        """, exit_params + deals_rule_params).fetchall()

    by_day: dict[str, dict[str, Any]] = {}
    for deal in closed_deals:
        try:
            financials = deal_financials_from_row(deal, investment)
        except (InvalidOperation, ValueError, TypeError, ZeroDivisionError):
            continue
        pnl_usd = financials["net_pnl_usd"]
        roi_percent = financials["net_roi_percent"]

        day = local_date_key(deal["exit_at"])
        row = by_day.setdefault(day, {
            "day": day,
            "closed_deals": 0,
            "wins": 0,
            "losses": 0,
            "net_pnl_usd": 0.0,
            "avg_roi_percent": None,
            "win_rate": None,
            "_roi_values": [],
        })
        row["closed_deals"] += 1
        row["wins"] += 1 if deal["result"] == "win" else 0
        row["losses"] += 1 if deal["result"] == "loss" else 0
        row["net_pnl_usd"] += pnl_usd
        row["_roi_values"].append(roi_percent)

    rows = sorted(by_day.values(), key=lambda item: item["day"])
    for row in rows:
        roi_values = row.pop("_roi_values")
        row["win_rate"] = (row["wins"] / row["closed_deals"] * 100) if row["closed_deals"] else None
        row["avg_roi_percent"] = (sum(roi_values) / len(roi_values)) if roi_values else None
    return rows


def load_btc_volume_trends(
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    limit: int = 120,
) -> list[dict[str, Any]]:
    btc_filter, btc_params = time_filter_sql("sampled_at", dashboard_range, custom_from, custom_to)
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT
                sampled_at,
                volume_btc_cumulative,
                volume_btc_delta,
                event_slug,
                status
            FROM btc_volume_log
            WHERE status IN ('success', 'baseline')
              AND """ + btc_filter + """
            ORDER BY sampled_at DESC, id DESC
            LIMIT ?
        """, (*btc_params, limit)).fetchall()

    result = []
    for row in reversed(rows):
        result.append({
            "sampled_at": row["sampled_at"],
            "label": local_time_label(row["sampled_at"]),
            "volume_btc_cumulative": float(row["volume_btc_cumulative"] or 0),
            "volume_btc_delta": float(row["volume_btc_delta"] or 0),
            "event_slug": row["event_slug"],
            "status": row["status"],
        })
    return result


def load_btc_volume_deal_snapshot(
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> dict[str, Any]:
    entry_filter, entry_params = time_filter_sql("entry_at", dashboard_range, custom_from, custom_to)
    deals_rule_filter, deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        summary = conn.execute("""
            SELECT
                COUNT(*) AS total_deals,
                SUM(CASE WHEN entry_btc_volume_btc_delta > 6 THEN 1 ELSE 0 END) AS deals_over_6_delta,
                SUM(CASE WHEN entry_btc_volume_log_id IS NULL THEN 1 ELSE 0 END) AS missing_snapshot_deals,
                AVG(entry_btc_volume_btc_delta) AS avg_delta,
                MAX(entry_btc_volume_btc_delta) AS max_delta
            FROM deals
            WHERE """ + entry_filter + """
              AND """ + deals_rule_filter, entry_params + deals_rule_params).fetchone()
        rows = conn.execute("""
            SELECT
                id,
                rule_name,
                event_id,
                side,
                result,
                entry_at,
                entry_btc_volume_sampled_at,
                entry_btc_volume_btc_cumulative,
                entry_btc_volume_btc_delta
            FROM deals
            WHERE entry_btc_volume_btc_delta > 6
              AND """ + entry_filter + """
              AND """ + deals_rule_filter + """
            ORDER BY entry_btc_volume_btc_delta DESC, entry_at DESC
            LIMIT 25
        """, entry_params + deals_rule_params).fetchall()

    return {
        "total_deals": int(summary["total_deals"] or 0),
        "deals_over_6_delta": int(summary["deals_over_6_delta"] or 0),
        "missing_snapshot_deals": int(summary["missing_snapshot_deals"] or 0),
        "avg_delta": summary["avg_delta"],
        "max_delta": summary["max_delta"],
        "rows_over_6": [row_to_dict(row) for row in rows],
    }


def load_data_quality_snapshot(
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> dict[str, Any]:
    orderbook_filter, orderbook_params = time_filter_sql("sampled_at", dashboard_range, custom_from, custom_to)
    btc_filter, btc_params = time_filter_sql("sampled_at", dashboard_range, custom_from, custom_to)
    deals_rule_filter, deals_rule_params = rule_filter_sql("d.rule_id", rule_filter)
    simple_deals_rule_filter, simple_deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        missing_rule_deals = int(conn.execute("""
            SELECT COUNT(*)
            FROM deals d
            LEFT JOIN rules r ON r.id = d.rule_id
            WHERE r.id IS NULL AND """ + deals_rule_filter, deals_rule_params).fetchone()[0])
        missing_event_deals = int(conn.execute("""
            SELECT COUNT(*)
            FROM deals d
            LEFT JOIN events e ON e.event_slug = d.event_id
            WHERE d.event_id IS NOT NULL AND e.event_slug IS NULL AND """ + deals_rule_filter, deals_rule_params).fetchone()[0])
        stale_open_deals = int(conn.execute("""
            SELECT COUNT(*)
            FROM deals
            WHERE result = 'open' AND entry_at < ? AND """ + simple_deals_rule_filter,
            ((now_utc() - timedelta(minutes=30)).isoformat(),) + simple_deals_rule_params,
        ).fetchone()[0])
        closed_deals_missing_fee_snapshot = int(conn.execute("""
            SELECT COUNT(*)
            FROM deals
            WHERE result IN ('win', 'loss')
              AND exit_price IS NOT NULL
              AND (fee_calculation_source IS NULL OR net_pnl_usd IS NULL)
              AND """ + simple_deals_rule_filter, simple_deals_rule_params).fetchone()[0])
        deals_missing_btc_volume_snapshot = int(conn.execute("""
            SELECT COUNT(*)
            FROM deals
            WHERE entry_btc_volume_log_id IS NULL
              AND """ + simple_deals_rule_filter, simple_deals_rule_params).fetchone()[0])
        event_status_mismatch = int(conn.execute("""
            SELECT COUNT(*)
            FROM events
            WHERE
                (status = 'closed' AND COALESCE(closed, 0) = 0)
                OR (status = 'open' AND COALESCE(closed, 0) = 1)
        """).fetchone()[0])
        orderbook_issues = int(conn.execute("""
            SELECT COUNT(*)
            FROM orderbook_log
            WHERE status IS NOT NULL AND status != 'success'
                AND """ + orderbook_filter, orderbook_params).fetchone()[0])
        invalid_orderbook_prices = int(conn.execute("""
            SELECT COUNT(*)
            FROM orderbook_log
            WHERE (
                up_best_ask NOT BETWEEN 0 AND 1
                OR up_best_bid NOT BETWEEN 0 AND 1
                OR down_best_ask NOT BETWEEN 0 AND 1
                OR down_best_bid NOT BETWEEN 0 AND 1
            )
            AND """ + orderbook_filter, orderbook_params).fetchone()[0])
        btc_volume_errors = int(conn.execute("""
            SELECT COUNT(*)
            FROM btc_volume_log
            WHERE status = 'error'
                AND """ + btc_filter, btc_params).fetchone()[0])
        negative_btc_delta = int(conn.execute("""
            SELECT COUNT(*)
            FROM btc_volume_log
            WHERE volume_btc_delta < 0
                AND """ + btc_filter, btc_params).fetchone()[0])

    checks = [
        ("עסקאות ללא חוק", missing_rule_deals, "error"),
        ("עסקאות ללא אירוע", missing_event_deals, "warning"),
        ("עסקאות פתוחות מעל 30 דקות", stale_open_deals, "warning"),
        ("חוסר התאמה בסטטוס אירוע", event_status_mismatch, "warning"),
        ("דגימות Orderbook לא תקינות", orderbook_issues, "warning"),
        ("מחירי Orderbook לא תקינים", invalid_orderbook_prices, "error"),
        ("שגיאות נפח BTC", btc_volume_errors, "warning"),
        ("דלתא נפח BTC שלילית", negative_btc_delta, "error"),
    ]
    checks.append(("Closed deals missing fee snapshot", closed_deals_missing_fee_snapshot, "warning"))
    checks.append(("Deals missing BTC volume snapshot", deals_missing_btc_volume_snapshot, "warning"))
    issue_count = sum(count for _, count, _ in checks)
    return {
        "status": "ok" if issue_count == 0 else "needs_review",
        "issue_count": issue_count,
        "checks": [
            {"name": name, "count": count, "severity": severity}
            for name, count, severity in checks
        ],
    }


def render_metric(label: str, value: str, note: str = "") -> str:
    note_html = f"<div class=\"metric-note\">{html.escape(note)}</div>" if note else ""
    return f"""
    <div class="metric">
        <div class="metric-label">{html.escape(label)}</div>
        <div class="metric-value">{html.escape(value)}</div>
        {note_html}
    </div>
    """


def render_dashboard_overview(overview: dict[str, Any]) -> str:
    best_rule = overview.get("best_rule")
    worst_rule = overview.get("worst_rule")
    best_rule_label = "-" if not best_rule else f"{best_rule['rule_name']} ({format_money(best_rule['pnl_usd'])})"
    worst_rule_label = "-" if not worst_rule else f"{worst_rule['rule_name']} ({format_money(worst_rule['pnl_usd'])})"
    data_health = "תקין" if overview["orderbook_errors"] == 0 else f"{overview['orderbook_errors']} בעיות Orderbook"

    metrics = [
        render_metric("רווח/הפסד נטו", format_money(overview["net_pnl_usd"]), "עסקאות סגורות בלבד"),
        render_metric("Gross P&L", format_money(overview["gross_pnl_usd"]), "לפני עמלות"),
        render_metric("Total fees", format_money(overview["total_fees_usd"]), f"Entry {format_money(overview['entry_fees_usd'])} / Exit {format_money(overview['exit_fees_usd'])}"),
        render_metric("Deals volume > 6 BTC", str(overview["btc_volume_gt_6_deals"]), f"{overview['missing_btc_volume_snapshot_deals']} without BTC snapshot"),
        render_metric("עסקאות סגורות", str(overview["closed_deals"]), f"{overview['open_deals']} פתוחות"),
        render_metric("Avg fee / deal", format_money(overview["avg_fee_usd"]), f"{format_percent(overview['fees_to_investment_percent'])} מההשקעה"),
        render_metric("אחוז הצלחה", format_percent(overview["win_rate"]), f"{overview['wins']} רווחיות / {overview['losses']} הפסדיות"),
        render_metric("ROI ממוצע לעסקה", format_percent(overview["avg_roi_percent"]), f"{format_money(overview['avg_pnl_usd'])} רווח/הפסד ממוצע"),
        render_metric(
            "יחס רווח להפסד",
            format_factor(overview["profit_factor"], overview["profit_factor_infinite"]),
            "רווח גולמי / הפסד גולמי",
        ),
        render_metric("ירידה מקסימלית", format_money(overview["max_drawdown_usd"]), "עקומת הון של עסקאות סגורות"),
        render_metric("החוק המוביל", best_rule_label),
        render_metric("תקינות נתונים", data_health, f"דגימת Orderbook אחרונה: {display_value(overview['last_orderbook_sample'])}"),
    ]

    investment_value = display_value(overview["investment_usd"])
    selected_range = normalize_dashboard_range(overview.get("range"))
    selected_rule_filter = normalize_rule_filter(overview.get("rule_filter"))
    custom_from_value = html.escape(str(overview.get("custom_from") or ""))
    custom_to_value = html.escape(str(overview.get("custom_to") or ""))
    range_options = "".join(
        f"<option value=\"{html.escape(value)}\"{' selected' if value == selected_range else ''}>{html.escape(label)}</option>"
        for value, label in dashboard_range_options()
    )
    with get_conn() as conn:
        rules = conn.execute("SELECT id, name FROM rules ORDER BY id ASC").fetchall()
    rule_options = [f"<option value=\"all\"{' selected' if selected_rule_filter == 'all' else ''}>כל החוקים</option>"]
    for rule in rules:
        value = str(rule["id"])
        selected = " selected" if selected_rule_filter == value else ""
        rule_options.append(f"<option value=\"{value}\"{selected}>{html.escape(rule['name'])} (#{rule['id']})</option>")
    return f"""
    <div class="card overview-card">
        <div class="overview-header">
            <div>
                <h2>סקירה ניהולית</h2>
                <p class="muted">מדדים פיננסיים מחושבים מעסקאות סגורות בלבד. טווח נוכחי: {html.escape(overview['range_label'])}.</p>
            </div>
            <form method="get" action="/" class="investment-form">
                <label>השקעה לעסקה
                    <input name="investment_usd" type="number" step="0.01" min="0.01" value="{html.escape(investment_value)}">
                </label>
                <label>טווח תאריכים
                    <select name="range_filter" id="range-filter" onchange="toggleCustomRangeInputs()">{range_options}</select>
                </label>
                <label>Rule filter
                    <select name="rule_filter" id="rule-filter" onchange="rememberRuleFilterDefault()">{''.join(rule_options)}</select>
                </label>
                <label class="custom-range-field">מתאריך ושעה
                    <input name="custom_from" type="text" inputmode="numeric" placeholder="DD/MM/YYYY HH:mm" value="{custom_from_value}">
                </label>
                <label class="custom-range-field">עד תאריך ושעה
                    <input name="custom_to" type="text" inputmode="numeric" placeholder="DD/MM/YYYY HH:mm" value="{custom_to_value}">
                </label>
                <button class="button" type="submit">החל</button>
            </form>
        </div>
        <div class="metric-grid">
            {''.join(metrics)}
        </div>
        <p class="muted">
            Gross P&L - Entry fees - Exit fees = Net P&L |
            Maker fills: {overview['maker_fills']} |
            Taker fills: {overview['taker_fills']} |
            Deals charged fees: {overview['fee_charged_deals']}
        </p>
        <p class="muted">
            חוקים פעילים: {overview['active_rules']} |
            סך עסקאות: {overview['total_deals']} |
            רצף רווחים ארוך ביותר: {overview['longest_win_streak']} |
            רצף הפסדים ארוך ביותר: {overview['longest_loss_streak']} |
            החוק החלש ביותר: {html.escape(worst_rule_label)}
        </p>
    </div>
    """


def render_rules_performance(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return """
        <div class="card">
            <h2>ביצועי חוקים</h2>
            <p>אין חוקים עדיין.</p>
        </div>
        """

    headers = [
        "חוק",
        "סטטוס",
        "סגורות",
        "פתוחות",
        "רווחיות",
        "הפסדיות",
        "אחוז הצלחה",
        "רווח/הפסד נטו",
        "ROI ממוצע",
        "יחס רווח להפסד",
        "תוחלת",
        "ירידה מקסימלית",
        "כניסה / סטופ / יעד",
    ]
    thead = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body = ""
    for row in rows:
        price_summary = (
            f"{display_value(row['entry_price'])} / "
            f"{display_value(row['stop_loss_price'])} / "
            f"{display_value(row['take_profit_price'])}"
        )
        values = [
            row["rule_name"],
            row["status"],
            row["closed_deals"],
            row["open_deals"],
            row["wins"],
            row["losses"],
            format_percent(row["win_rate"]),
            format_money(row["net_pnl_usd"]),
            format_percent(row["avg_roi_percent"]),
            format_factor(row["profit_factor"], row["profit_factor_infinite"]),
            format_money(row["expectancy_usd"]),
            format_money(row["max_drawdown_usd"]),
            price_summary,
        ]
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in values)
        body += f"<tr>{cells}</tr>"

    return f"""
    <div class="card">
        <h2>ביצועי חוקים</h2>
        <p class="muted">מדורג לפי רווח/הפסד נטו בעסקאות סגורות. ההשקעה היא פילטר תצוגה בלבד.</p>
        <div class="table-wrap compact-table">
          <table>
            <thead><tr>{thead}</tr></thead>
            <tbody>{body}</tbody>
          </table>
        </div>
    </div>
    """


def render_deal_summary(deal: Optional[dict[str, Any]]) -> str:
    if not deal:
        return "-"
    return (
        f"#{deal['deal_id']} {deal['rule_name']} "
        f"{format_money(deal['pnl_usd'])} / {format_percent(deal['roi_percent'])}"
    )


def render_exit_reason_breakdown(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "<p>No closed deals yet.</p>"

    headers = ["סיבת יציאה", "עסקאות", "רווחיות", "הפסדיות", "אחוז הצלחה", "רווח/הפסד נטו"]
    thead = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body = ""
    for row in rows:
        values = [
            row["exit_reason"],
            row["deals"],
            row["wins"],
            row["losses"],
            format_percent(row["win_rate"]),
            format_money(row["pnl_usd"]),
        ]
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in values)
        body += f"<tr>{cells}</tr>"

    return f"""
    <div class="table-wrap risk-table">
      <table>
        <thead><tr>{thead}</tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
    """


def render_risk_snapshot(risk: dict[str, Any]) -> str:
    metrics = [
        render_metric("הון מסכם", format_money(risk["ending_equity_usd"]), "סכום רווח/הפסד מעסקאות סגורות"),
        render_metric("שיא הון", format_money(risk["peak_equity_usd"]), "נקודת ההון הגבוהה ביותר"),
        render_metric("ירידה מקסימלית", format_money(risk["max_drawdown_usd"]), f"אחרי: {display_value(risk['max_drawdown_after'])}"),
        render_metric("רצף הפסדים", str(risk["longest_loss_streak"]), f"{risk['loss_deals']} עסקאות הפסדיות"),
        render_metric("רווח ממוצע", format_money(risk["avg_win_usd"]), f"{risk['win_deals']} עסקאות רווחיות"),
        render_metric("הפסד ממוצע", format_money(risk["avg_loss_usd"]), "ממוצע לעסקה הפסדית"),
        render_metric("העסקה הטובה ביותר", render_deal_summary(risk["best_deal"])),
        render_metric("העסקה החלשה ביותר", render_deal_summary(risk["worst_deal"])),
    ]

    return f"""
    <div class="card">
        <h2>תמונת סיכון</h2>
        <p class="muted">מדדי הסיכון מחושבים מעסקאות סגורות בלבד ולפי פילטר ההשקעה הנוכחי.</p>
        <div class="metric-grid">
            {''.join(metrics)}
        </div>
        <h3>סיבות יציאה</h3>
        {render_exit_reason_breakdown(risk["exit_reasons"])}
    </div>
    """


def render_condition_table(rows: list[dict[str, Any]], empty_text: str) -> str:
    if not rows:
        return f"<p>{html.escape(empty_text)}</p>"

    headers = ["קבוצה", "סגורות", "רווחיות", "הפסדיות", "אחוז הצלחה", "רווח/הפסד נטו", "רווח/הפסד ממוצע", "ROI ממוצע"]
    thead = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body = ""
    for row in rows:
        values = [
            row["label"],
            row["closed_deals"],
            row["wins"],
            row["losses"],
            format_percent(row["win_rate"]),
            format_money(row["net_pnl_usd"]),
            format_money(row["avg_pnl_usd"]),
            format_percent(row["avg_roi_percent"]),
        ]
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in values)
        body += f"<tr>{cells}</tr>"

    return f"""
    <div class="table-wrap condition-table">
      <table>
        <thead><tr>{thead}</tr></thead>
        <tbody>{body}</tbody>
      </table>
    </div>
    """


def render_market_conditions(conditions: dict[str, list[dict[str, Any]]]) -> str:
    return f"""
    <div class="card">
        <h2>תנאי שוק</h2>
        <p class="muted">עסקאות סגורות בלבד. בשלב הנוכחי משתמשים בשדות אמינים מהעסקה: צד וטווח מחיר כניסה.</p>
        <h3>ביצועים לפי צד</h3>
        {render_condition_table(conditions["by_side"], "אין עדיין עסקאות סגורות לפי צד.")}
        <h3>ביצועים לפי מחיר כניסה</h3>
        {render_condition_table(conditions["by_entry_price"], "אין עדיין עסקאות סגורות לפי מחיר כניסה.")}
    </div>
    """


def render_system_health(health: dict[str, Any]) -> str:
    metrics = [
        render_metric("סטטוס מערכת", display_value(health["status"]), f"גודל DB: {health['db_size']}"),
        render_metric("אירוע אחרון", display_value(health["latest_event"])),
        render_metric("Orderbook אחרון", display_value(health["latest_orderbook"]), f"{health['orderbook_issues']} בעיות"),
        render_metric("נפח BTC אחרון", display_value(health["latest_btc_volume"]), f"{health['btc_volume_issues']} שגיאות"),
        render_metric("איסוף Coinbase", display_value(health["coinbase_status"]), f"הצלחה: {display_value(health['coinbase_last_success'])}"),
        render_metric("עסקאות פתוחות", str(health["open_deals"]), f"{health['deals_count']} עסקאות בסך הכל"),
    ]

    rows = [
        ("אירועים", health["events_count"]),
        ("דגימות Orderbook", health["orderbook_count"]),
        ("דגימות נפח BTC", health["btc_volume_count"]),
        ("חוקים", health["rules_count"]),
        ("חוקים פעילים", health["active_rules"]),
        ("עסקאות", health["deals_count"]),
    ]
    body = "".join(
        f"<tr><td>{html.escape(label)}</td><td>{html.escape(str(value))}</td></tr>"
        for label, value in rows
    )
    last_error = display_value(health["coinbase_last_error"])

    return f"""
    <div class="card">
        <h2>בריאות מערכת</h2>
        <p class="muted">תמונת מצב תפעולית של ה־collector המקומי ב־FastAPI/SQLite.</p>
        <div class="metric-grid">
            {''.join(metrics)}
        </div>
        <div class="table-wrap health-table">
          <table>
            <thead><tr><th>ישות</th><th>כמות</th></tr></thead>
            <tbody>{body}</tbody>
          </table>
        </div>
        <p class="muted">DB: {html.escape(health["db_path"])} | שגיאת Coinbase אחרונה: {html.escape(last_error)}</p>
    </div>
    """


def chart_bar_width(value: Any, max_abs: float) -> int:
    try:
        numeric = abs(float(value))
    except (TypeError, ValueError):
        return 0
    if max_abs <= 0:
        return 0
    return max(2, min(100, int(round(numeric / max_abs * 100))))


def render_time_trends(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return """
        <div class="card">
            <h2>מגמות לאורך זמן</h2>
            <p>אין עסקאות סגורות בטווח שנבחר.</p>
        </div>
        """

    max_abs_pnl = max(abs(float(row["net_pnl_usd"])) for row in rows) or 1
    headers = ["יום", "סגורות", "רווחיות", "הפסדיות", "אחוז הצלחה", "רווח/הפסד נטו", "ROI ממוצע", "רווח/הפסד"]
    thead = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body = ""
    for row in rows:
        pnl = float(row["net_pnl_usd"])
        bar_class = "positive" if pnl >= 0 else "negative"
        bar_width = chart_bar_width(pnl, max_abs_pnl)
        values = [
            row["day"],
            row["closed_deals"],
            row["wins"],
            row["losses"],
            format_percent(row["win_rate"]),
            format_money(pnl),
            format_percent(row["avg_roi_percent"]),
            f"<div class=\"bar-track\"><div class=\"bar {bar_class}\" style=\"width:{bar_width}%\"></div></div>",
        ]
        cells = "".join(
            f"<td>{value}</td>" if str(value).startswith("<div") else f"<td>{html.escape(str(value))}</td>"
            for value in values
        )
        body += f"<tr>{cells}</tr>"

    return f"""
    <div class="card">
        <h2>מגמות לאורך זמן</h2>
        <p class="muted">ביצועי עסקאות סגורות לפי תאריך מקומי.</p>
        <div class="table-wrap trends-table">
          <table>
            <thead><tr>{thead}</tr></thead>
            <tbody>{body}</tbody>
          </table>
        </div>
    </div>
    """


def render_data_quality(quality: dict[str, Any]) -> str:
    status_class = "ok" if quality["status"] == "ok" else "warning"
    metrics = [
        render_metric("סטטוס איכות", display_value(quality["status"]), f"{quality['issue_count']} בעיות בסך הכל"),
    ]
    rows = ""
    for check in quality["checks"]:
        badge_class = "ok" if check["count"] == 0 else check["severity"]
        rows += (
            "<tr>"
            f"<td>{html.escape(check['name'])}</td>"
            f"<td><span class=\"badge {badge_class}\">{html.escape(str(check['count']))}</span></td>"
            f"<td>{html.escape(check['severity'])}</td>"
            "</tr>"
        )

    return f"""
    <div class="card quality-card {status_class}">
        <h2>איכות נתונים</h2>
        <p class="muted">בדיקות פנימיות על עסקאות, אירועים, דגימות Orderbook ולוג נפח Coinbase.</p>
        <div class="metric-grid">{''.join(metrics)}</div>
        <div class="table-wrap quality-table">
          <table>
            <thead><tr><th>בדיקה</th><th>כמות</th><th>חומרה</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>
    </div>
    """


def render_dashboard_charts(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return """
        <div class="card">
            <h2>גרפים</h2>
            <p>אין נתונים לגרפים בטווח שנבחר.</p>
        </div>
        """

    max_abs_pnl = max(abs(float(row["net_pnl_usd"])) for row in rows) or 1
    max_deals = max(int(row["closed_deals"]) for row in rows) or 1
    pnl_items = ""
    deal_items = ""
    for row in rows[-30:]:
        pnl = float(row["net_pnl_usd"])
        pnl_class = "positive" if pnl >= 0 else "negative"
        pnl_items += f"""
            <div class="chart-row">
                <div class="chart-label">{html.escape(row['day'])}</div>
                <div class="bar-track"><div class="bar {pnl_class}" style="width:{chart_bar_width(pnl, max_abs_pnl)}%"></div></div>
                <div class="chart-value">{html.escape(format_money(pnl))}</div>
            </div>
        """
        deal_items += f"""
            <div class="chart-row">
                <div class="chart-label">{html.escape(row['day'])}</div>
                <div class="bar-track"><div class="bar neutral" style="width:{chart_bar_width(row['closed_deals'], max_deals)}%"></div></div>
                <div class="chart-value">{html.escape(str(row['closed_deals']))}</div>
            </div>
        """

    return f"""
    <div class="card">
        <h2>גרפים</h2>
        <p class="muted">עד 30 הימים האחרונים שקובצו בטווח שנבחר.</p>
        <div class="charts-grid">
            <section>
                <h3>רווח/הפסד יומי נטו</h3>
                {pnl_items}
            </section>
            <section>
                <h3>עסקאות סגורות ליום</h3>
                {deal_items}
            </section>
        </div>
    </div>
    """


def chart_json_payload(time_trends: list[dict[str, Any]], btc_volume_trends: list[dict[str, Any]]) -> str:
    payload = {
        "timeTrendLabels": [row["day"] for row in time_trends[-30:]],
        "dailyPnl": [round(float(row["net_pnl_usd"]), 8) for row in time_trends[-30:]],
        "dailyDeals": [int(row["closed_deals"]) for row in time_trends[-30:]],
        "btcVolumeLabels": [row["label"] for row in btc_volume_trends],
        "btcVolumeDelta": [round(float(row["volume_btc_delta"]), 8) for row in btc_volume_trends],
        "btcVolumeCumulative": [round(float(row["volume_btc_cumulative"]), 8) for row in btc_volume_trends],
    }
    return (
        json.dumps(payload, ensure_ascii=False)
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )


def render_chartjs_charts(time_trends: list[dict[str, Any]], btc_volume_trends: list[dict[str, Any]]) -> str:
    if not time_trends and not btc_volume_trends:
        return """
        <div class="card chartjs-card">
            <h2>Interactive charts</h2>
            <p>No chart data yet.</p>
        </div>
        """

    return f"""
    <div class="card chartjs-card">
        <h2>Interactive charts</h2>
        <p class="muted">Line trend, bar chart, and BTC volume over time from the local SQLite data.</p>
        <script type="application/json" id="dashboard-chart-data">{chart_json_payload(time_trends, btc_volume_trends)}</script>
        <div class="canvas-grid">
            <section class="canvas-panel">
                <h3>Net P&L trend</h3>
                <canvas id="pnlTrendChart" height="220"></canvas>
            </section>
            <section class="canvas-panel">
                <h3>Closed deals by day</h3>
                <canvas id="dealsBarChart" height="220"></canvas>
            </section>
            <section class="canvas-panel wide">
                <h3>BTC volume vs time</h3>
                <canvas id="btcVolumeChart" height="240"></canvas>
            </section>
        </div>
    </div>
    """


def render_btc_volume_deal_snapshot(snapshot: dict[str, Any]) -> str:
    metrics = [
        render_metric("Deals volume > 6 BTC", str(snapshot["deals_over_6_delta"]), "entry volume_btc_delta threshold"),
        render_metric("Avg entry BTC delta", display_value(snapshot["avg_delta"]), f"{snapshot['total_deals']} deals in range"),
        render_metric("Max entry BTC delta", display_value(snapshot["max_delta"]), f"{snapshot['missing_snapshot_deals']} without snapshot"),
    ]

    if snapshot["rows_over_6"]:
        headers = [
            "deal",
            "rule",
            "event",
            "side",
            "result",
            "entry_at",
            "volume_sampled_at",
            "volume_delta",
            "volume_cumulative",
        ]
        thead = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
        body = ""
        for row in snapshot["rows_over_6"]:
            values = [
                row["id"],
                row["rule_name"] or "-",
                row["event_id"],
                row["side"],
                row["result"],
                row["entry_at"],
                row["entry_btc_volume_sampled_at"],
                row["entry_btc_volume_btc_delta"],
                row["entry_btc_volume_btc_cumulative"],
            ]
            body += "<tr>" + "".join(f"<td>{html.escape(str(display_value(value)))}</td>" for value in values) + "</tr>"
        table = f"""
        <div class="table-wrap btc-volume-deals-table">
          <table>
            <thead><tr>{thead}</tr></thead>
            <tbody>{body}</tbody>
          </table>
        </div>
        """
    else:
        table = "<p>No deals above 6 BTC volume delta in the selected range.</p>"

    return f"""
    <div class="card">
        <h2>BTC volume on deals</h2>
        <p class="muted">Each new deal stores the latest Coinbase BTC volume sample available at entry time. The threshold here is volume_btc_delta &gt; 6.</p>
        <div class="metric-grid">{''.join(metrics)}</div>
        {table}
    </div>
    """


def latest_export_path() -> Optional[Path]:
    state_path = export_state.get("path")
    if state_path:
        path = Path(state_path)
        if path.exists():
            return path

    if not EXPORT_DIR.exists():
        return None

    exports = sorted(
        EXPORT_DIR.glob(f"{EXPORT_FILE_PREFIX}*{EXPORT_FILE_SUFFIX}"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    return exports[0] if exports else None


def create_sqlite_snapshot(snapshot_path: Path) -> None:
    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    source_conn = connect_db()
    snapshot_conn = connect_sqlite(snapshot_path, set_journal_mode=False)
    try:
        source_conn.backup(snapshot_conn)
        snapshot_conn.commit()
    finally:
        snapshot_conn.close()
        source_conn.close()


def cleanup_export_temp_file(path: Path) -> bool:
    removed = False
    for candidate in (path, Path(f"{path}-wal"), Path(f"{path}-shm"), Path(f"{path}-journal")):
        try:
            if candidate.exists():
                candidate.unlink()
                removed = True
        except OSError as exc:
            print(f"[export] cleanup warning path={candidate}: {type(exc).__name__}: {exc}", flush=True)
    return removed


def write_xlsx_export() -> tuple[Path, dict[str, int]]:
    init_db()
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = now_utc().strftime("%Y%m%d_%H%M%S")
    final_path = EXPORT_DIR / f"{EXPORT_FILE_PREFIX}{timestamp}{EXPORT_FILE_SUFFIX}"
    temp_path = EXPORT_DIR / f".{final_path.stem}.tmp{EXPORT_FILE_SUFFIX}"
    snapshot_path = EXPORT_DIR / f"{EXPORT_SNAPSHOT_PREFIX}{timestamp}.sqlite3"
    row_counts: dict[str, int] = {}
    started_at = now_utc()

    print(f"[export] started path={final_path}", flush=True)
    create_sqlite_snapshot(snapshot_path)
    print(
        f"[export] snapshot_created path={snapshot_path} "
        f"duration_seconds={(now_utc() - started_at).total_seconds():.3f}",
        flush=True,
    )

    workbook: Optional[Workbook] = None
    conn: Optional[sqlite3.Connection] = None
    try:
        workbook = Workbook(write_only=True)
        conn = connect_sqlite(snapshot_path, set_journal_mode=False)
        conn.row_factory = sqlite3.Row
        for sheet_name, headers, query in EXPORT_SHEETS:
            sheet_started_at = now_utc()
            print(f"[export] sheet_started sheet={sheet_name}", flush=True)
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(headers)
            cursor = conn.execute(query)
            try:
                count = 0
                while True:
                    rows = cursor.fetchmany(EXPORT_FETCH_CHUNK_SIZE)
                    if not rows:
                        break
                    for row in rows:
                        worksheet.append(list(row))
                    count += len(rows)
                row_counts[sheet_name] = count
                print(
                    f"[export] sheet_completed sheet={sheet_name} rows={count} "
                    f"duration_seconds={(now_utc() - sheet_started_at).total_seconds():.3f}",
                    flush=True,
                )
            finally:
                cursor.close()
        overview = load_dashboard_overview(DEMO_INVESTMENT_USD, conn=conn)
        fee_sheet = workbook.create_sheet("fee_summary")
        fee_sheet.append(["metric", "value"])
        fee_metrics = [
            ("investment_usd", overview["investment_usd"]),
            ("closed_deals", overview["closed_deals"]),
            ("gross_pnl_usd", overview["gross_pnl_usd"]),
            ("entry_fees_usd", overview["entry_fees_usd"]),
            ("exit_fees_usd", overview["exit_fees_usd"]),
            ("total_fees_usd", overview["total_fees_usd"]),
            ("net_pnl_usd", overview["net_pnl_usd"]),
            ("avg_fee_usd", overview["avg_fee_usd"]),
            ("fees_to_investment_percent", overview["fees_to_investment_percent"]),
            ("fees_to_gross_profit_percent", overview["fees_to_gross_profit_percent"]),
            ("fee_charged_deals", overview["fee_charged_deals"]),
            ("maker_fills", overview["maker_fills"]),
            ("taker_fills", overview["taker_fills"]),
            ("fee_calculation_version", DEMO_FEE_CALCULATION_VERSION),
        ]
        for metric, value in fee_metrics:
            fee_sheet.append([metric, value])
        row_counts["fee_summary"] = len(fee_metrics)
        workbook.save(temp_path)
        temp_path.replace(final_path)
        print(
            f"[export] completed path={final_path} size_bytes={final_path.stat().st_size} "
            f"duration_seconds={(now_utc() - started_at).total_seconds():.3f} rows={row_counts}",
            flush=True,
        )
        return final_path, row_counts
    except Exception as exc:
        cleanup_export_temp_file(temp_path)
        print(f"[export] failed error={type(exc).__name__}: {exc}", flush=True)
        raise
    finally:
        if conn is not None:
            conn.close()
        snapshot_removed = cleanup_export_temp_file(snapshot_path)
        print(f"[export] cleanup_completed snapshot_removed={snapshot_removed}", flush=True)


def run_xlsx_export() -> None:
    try:
        export_path, row_counts = write_xlsx_export()
        with export_state_lock:
            export_state.update({
                "status": "ready",
                "finished_at": now_iso(),
                "filename": export_path.name,
                "path": str(export_path),
                "error": None,
                "row_counts": json.dumps(row_counts, ensure_ascii=False),
            })
        print(f"[export] Excel ready: {export_path} rows={row_counts}", flush=True)
    except Exception as e:
        with export_state_lock:
            export_state.update({
                "status": "error",
                "finished_at": now_iso(),
                "error": truncate_text(f"{type(e).__name__}: {e}", 500),
            })
        log_error("excel export", e)


def render_export_actions() -> str:
    with export_state_lock:
        state = dict(export_state)

    latest_path = latest_export_path()
    status = state.get("status") or "idle"
    started_at = display_value(state.get("started_at"))
    finished_at = display_value(state.get("finished_at"))
    filename = state.get("filename") or (latest_path.name if latest_path else None)
    error = state.get("error")
    row_counts = state.get("row_counts")
    generate_disabled = " disabled" if status == "running" else ""
    download_disabled = "" if latest_path else " disabled"
    download_href = "/download.xlsx" if latest_path else "#"
    summary_parts = [f"Status: {status}"]
    if status == "running":
        summary_parts.append(f"Started: {started_at}")
    if status == "ready":
        summary_parts.append(f"Ready: {finished_at}")
    if filename:
        summary_parts.append(f"File: {filename}")
    if row_counts:
        summary_parts.append(f"Rows: {row_counts}")
    if error:
        summary_parts.append(f"Error: {error}")
    summary = " | ".join(summary_parts)

    return f"""
    <form method="post" action="/generate.xlsx" class="inline-form">
        <button class="button" type="submit"{generate_disabled}>Generate Excel</button>
    </form>
    <a class="button{download_disabled}" href="{download_href}">Download Excel</a>
    <div class="muted export-status">{html.escape(summary)}</div>
    """


def render_rule_actions() -> str:
    return """
    <button class="button" type="button" onclick="openRuleModal()">Create Rule / יצירת חוק</button>
    <button class="button secondary" type="button" onclick="openDeactivateModal()">Deactivate Rule / השבתת חוק</button>

    <div id="rule-modal" class="modal" hidden>
      <div class="modal-panel">
        <h2>Create Rule / יצירת חוק</h2>
        <label>Name <input id="rule-name" type="text"></label>
        <label>Entry price <input id="rule-entry" type="number" step="0.01" min="0.01" max="0.99"></label>
        <label>Stop loss <input id="rule-stop" type="number" step="0.01" min="0.01" max="0.99"></label>
        <label>Take profit <input id="rule-take" type="number" step="0.01" min="0.01" max="0.99"></label>
        <label>Max YES entries <input id="rule-max-yes" type="number" step="1" min="0" value="1"></label>
        <label>Max NO entries <input id="rule-max-no" type="number" step="1" min="0" value="1"></label>
        <label>Entry window start, seconds before event end <input id="rule-window-start" type="number" step="1" min="0" placeholder="120"></label>
        <label>Entry window end, seconds before event end <input id="rule-window-end" type="number" step="1" min="0" placeholder="0"></label>
        <label>Schedule timezone <input id="rule-timezone" type="text" value="Asia/Jerusalem"></label>
        <h3>Inactive windows</h3>
        <div id="inactive-windows"></div>
        <button class="button secondary" type="button" onclick="addInactiveWindow()">Add inactive window</button>
        <label>Status
          <select id="rule-status">
            <option value="active">active</option>
            <option value="inactive">inactive</option>
          </select>
        </label>
        <div id="rule-error" class="error"></div>
        <div class="modal-actions">
          <button class="button" type="button" onclick="submitRule()">Save</button>
          <button class="button secondary" type="button" onclick="closeRuleModal()">Cancel</button>
        </div>
      </div>
    </div>

    <div id="deactivate-modal" class="modal" hidden>
      <div class="modal-panel small">
        <h2>Deactivate Rule / השבתת חוק</h2>
        <label>Rule ID <input id="deactivate-rule-id" type="number" step="1" min="1"></label>
        <div id="deactivate-error" class="error"></div>
        <div class="modal-actions">
          <button class="button" type="button" onclick="submitDeactivate()">Deactivate</button>
          <button class="button secondary" type="button" onclick="closeDeactivateModal()">Cancel</button>
        </div>
      </div>
    </div>
    """


def render_dashboard_scripts() -> str:
    return """
    <script>
      let dashboardRefreshInFlight = false;
      let dashboardChartInstances = [];

      function modalIsOpen() {
        const ruleModal = document.getElementById("rule-modal");
        const deactivateModal = document.getElementById("deactivate-modal");
        return Boolean((ruleModal && !ruleModal.hidden) ||
          (deactivateModal && !deactivateModal.hidden));
      }

      function userIsEditing() {
        const element = document.activeElement;
        if (!element) {
          return false;
        }
        return ["INPUT", "SELECT", "TEXTAREA"].includes(element.tagName);
      }

      function autoRefreshIsPaused() {
        return modalIsOpen() || userIsEditing();
      }

      function toggleCustomRangeInputs() {
        const rangeFilter = document.getElementById("range-filter");
        const visible = rangeFilter && rangeFilter.value === "custom";
        document.querySelectorAll(".custom-range-field").forEach((element) => {
          element.hidden = !visible;
        });
      }

      function rememberRuleFilterDefault() {
        const ruleFilter = document.getElementById("rule-filter");
        if (ruleFilter) {
          window.localStorage.setItem("polymarket.dashboard.defaultRuleFilter", ruleFilter.value || "all");
        }
      }

      function applyDefaultRuleFilter() {
        const params = new URLSearchParams(window.location.search);
        const ruleFilter = document.getElementById("rule-filter");
        if (!ruleFilter || params.has("rule_filter")) {
          return;
        }
        const saved = window.localStorage.getItem("polymarket.dashboard.defaultRuleFilter");
        if (saved && [...ruleFilter.options].some((option) => option.value === saved)) {
          ruleFilter.value = saved;
          params.set("rule_filter", saved);
          window.history.replaceState({}, "", `${window.location.pathname}?${params.toString()}`);
          refreshDashboardContent(true);
        }
      }

      function readDashboardChartData() {
        const element = document.getElementById("dashboard-chart-data");
        if (!element) {
          return null;
        }
        try {
          return JSON.parse(element.textContent || "{}");
        } catch (error) {
          console.warn("Chart data parse failed", error);
          return null;
        }
      }

      function resetDashboardCharts() {
        dashboardChartInstances.forEach((chart) => chart.destroy());
        dashboardChartInstances = [];
      }

      function createChart(canvasId, config) {
        const canvas = document.getElementById(canvasId);
        if (!canvas || !window.Chart) {
          return;
        }
        dashboardChartInstances.push(new Chart(canvas, config));
      }

      function renderDashboardCharts() {
        const data = readDashboardChartData();
        resetDashboardCharts();
        if (!data || !window.Chart) {
          return;
        }
        const baseOptions = {
          responsive: true,
          maintainAspectRatio: false,
          plugins: {legend: {position: "bottom"}},
          scales: {x: {ticks: {maxRotation: 0, autoSkip: true}}, y: {beginAtZero: false}}
        };
        createChart("pnlTrendChart", {
          type: "line",
          data: {
            labels: data.timeTrendLabels || [],
            datasets: [{
              label: "Net P&L",
              data: data.dailyPnl || [],
              borderColor: "#147a3f",
              backgroundColor: "rgba(20, 122, 63, 0.12)",
              tension: 0.25,
              fill: true
            }]
          },
          options: baseOptions
        });
        createChart("dealsBarChart", {
          type: "bar",
          data: {
            labels: data.timeTrendLabels || [],
            datasets: [{
              label: "Closed deals",
              data: data.dailyDeals || [],
              backgroundColor: "#2f5f98"
            }]
          },
          options: {...baseOptions, scales: {...baseOptions.scales, y: {beginAtZero: true, ticks: {precision: 0}}}}
        });
        createChart("btcVolumeChart", {
          type: "bar",
          data: {
            labels: data.btcVolumeLabels || [],
            datasets: [
              {
                type: "bar",
                label: "BTC delta",
                data: data.btcVolumeDelta || [],
                backgroundColor: "rgba(47, 95, 152, 0.55)",
                yAxisID: "y"
              },
              {
                type: "line",
                label: "BTC cumulative",
                data: data.btcVolumeCumulative || [],
                borderColor: "#b42318",
                backgroundColor: "rgba(180, 35, 24, 0.12)",
                tension: 0.2,
                pointRadius: 1,
                yAxisID: "y1"
              }
            ]
          },
          options: {
            responsive: true,
            maintainAspectRatio: false,
            plugins: {legend: {position: "bottom"}},
            scales: {
              x: {ticks: {maxRotation: 0, autoSkip: true}},
              y: {beginAtZero: true, position: "left"},
              y1: {beginAtZero: true, position: "right", grid: {drawOnChartArea: false}}
            }
          }
        });
      }

      async function refreshDashboardContent(force = false) {
        if (dashboardRefreshInFlight || (!force && autoRefreshIsPaused())) {
          return;
        }
        dashboardRefreshInFlight = true;
        try {
          const contentPath = document.body.dataset.contentPath || "/dashboard-content";
          const response = await fetch(`${contentPath}${window.location.search}`, {headers: {"X-Requested-With": "fetch"}});
          if (response.ok) {
            document.getElementById("dashboard-content").innerHTML = await response.text();
            toggleCustomRangeInputs();
            renderDashboardCharts();
          }
        } finally {
          dashboardRefreshInFlight = false;
        }
      }

      function openRuleModal() {
        if (!document.getElementById("rule-modal")) {
          return;
        }
        document.getElementById("rule-error").textContent = "";
        document.getElementById("rule-modal").hidden = false;
      }
      function closeRuleModal() {
        if (!document.getElementById("rule-modal")) {
          return;
        }
        document.getElementById("rule-modal").hidden = true;
      }
      function addInactiveWindow(windowData = {}) {
        const container = document.getElementById("inactive-windows");
        if (!container) {
          return;
        }
        const row = document.createElement("div");
        row.className = "inactive-window-row";
        row.innerHTML = `
          <label>Day
            <select class="inactive-day">
              <option value="0">Monday / שני</option>
              <option value="1">Tuesday / שלישי</option>
              <option value="2">Wednesday / רביעי</option>
              <option value="3">Thursday / חמישי</option>
              <option value="4">Friday / שישי</option>
              <option value="5">Saturday / שבת</option>
              <option value="6">Sunday / ראשון</option>
            </select>
          </label>
          <label>Start <input class="inactive-start" type="time" step="1" value="${windowData.start_time || "02:00:00"}"></label>
          <label>End <input class="inactive-end" type="time" step="1" value="${windowData.end_time || "05:00:00"}"></label>
          <label>Window state
            <select class="inactive-status">
              <option value="active">Enabled — blocks new entries</option>
              <option value="inactive">Disabled — does not block</option>
            </select>
          </label>
          <button class="button secondary" type="button">Remove</button>
        `;
        row.querySelector(".inactive-day").value = String(windowData.day_of_week ?? 0);
        row.querySelector(".inactive-status").value = windowData.status || "active";
        row.querySelector("button").addEventListener("click", () => row.remove());
        container.appendChild(row);
      }

      function readInactiveWindows() {
        return [...document.querySelectorAll(".inactive-window-row")].map((row) => ({
          day_of_week: row.querySelector(".inactive-day").value,
          start_time: row.querySelector(".inactive-start").value,
          end_time: row.querySelector(".inactive-end").value,
          status: row.querySelector(".inactive-status").value
        }));
      }
      function openDeactivateModal() {
        if (!document.getElementById("deactivate-modal")) {
          return;
        }
        document.getElementById("deactivate-error").textContent = "";
        document.getElementById("deactivate-modal").hidden = false;
      }
      function closeDeactivateModal() {
        if (!document.getElementById("deactivate-modal")) {
          return;
        }
        document.getElementById("deactivate-modal").hidden = true;
      }
      async function submitRule() {
        const payload = {
          name: document.getElementById("rule-name").value,
          entry_price: document.getElementById("rule-entry").value,
          stop_loss_price: document.getElementById("rule-stop").value,
          take_profit_price: document.getElementById("rule-take").value,
          max_yes_entries_per_event: document.getElementById("rule-max-yes").value,
          max_no_entries_per_event: document.getElementById("rule-max-no").value,
          entry_window_start_seconds_before_end: document.getElementById("rule-window-start").value,
          entry_window_end_seconds_before_end: document.getElementById("rule-window-end").value,
          schedule_timezone: document.getElementById("rule-timezone").value,
          inactive_windows: readInactiveWindows(),
          status: document.getElementById("rule-status").value
        };
        const response = await fetch("/rules", {
          method: "POST",
          headers: {"Content-Type": "application/json"},
          body: JSON.stringify(payload)
        });
        const data = await response.json();
        if (!response.ok) {
          document.getElementById("rule-error").textContent = data.detail || "Rule creation failed";
          return;
        }
        closeRuleModal();
        await refreshDashboardContent(true);
      }
      async function submitDeactivate() {
        const ruleId = document.getElementById("deactivate-rule-id").value;
        const response = await fetch(`/rules/${ruleId}/deactivate`, {method: "POST"});
        const data = await response.json();
        if (!response.ok) {
          document.getElementById("deactivate-error").textContent = data.detail || "Deactivate failed";
          return;
        }
        closeDeactivateModal();
        await refreshDashboardContent(true);
      }

      document.addEventListener("DOMContentLoaded", () => {
        toggleCustomRangeInputs();
        applyDefaultRuleFilter();
        renderDashboardCharts();
      });
      window.setInterval(() => refreshDashboardContent(false), 10000);
    </script>
    """


def load_dashboard_rows(rule_filter: Any = "all") -> tuple[
    list[sqlite3.Row],
    list[sqlite3.Row],
    list[sqlite3.Row],
    Optional[sqlite3.Row],
    list[sqlite3.Row],
    list[sqlite3.Row],
    dict[str, Optional[str]],
]:
    deals_rule_filter, deals_rule_params = rule_filter_sql("rule_id", rule_filter)
    with get_conn() as conn:
        events = conn.execute("""
            SELECT
                local_event_id,
                polymarket_event_id,
                polymarket_market_id,
                condition_id,
                event_slug,
                market_slug,
                title,
                question,
                event_url,
                start_time,
                start_time_local,
                end_time,
                end_time_local,
                yes_token_id,
                no_token_id,
                outcomes,
                outcome_prices,
                active,
                closed,
                enable_order_book,
                accepting_orders,
                fees_enabled,
                fee_rate,
                fee_calculation_source,
                fee_calculation_version,
                created_at_poly,
                created_at_poly_local,
                discovered_at,
                discovered_at_local,
                last_seen_at,
                last_seen_at_local,
                status,
                notes
            FROM events
            ORDER BY local_event_id DESC
            LIMIT 50
        """).fetchall()

        logs = conn.execute("""
            SELECT
                sampled_at,
                sampled_at_local,
                event_slug,
                condition_id,
                up_token_id,
                down_token_id,
                up_best_ask,
                up_best_bid,
                down_best_ask,
                down_best_bid,
                up_last_trade_price,
                down_last_trade_price,
                up_spread,
                down_spread,
                up_midpoint,
                down_midpoint,
                raw_up_timestamp,
                raw_down_timestamp,
                up_volume_shares_10s,
                down_volume_shares_10s,
                up_volume_usdc_10s,
                down_volume_usdc_10s,
                trades_count_10s,
                trades_window_start,
                trades_window_start_local,
                trades_window_end,
                trades_window_end_local,
                trades_error,
                status,
                error
            FROM orderbook_log
            ORDER BY id DESC
            LIMIT 300
        """).fetchall()

        btc_volume_rows = conn.execute("""
            SELECT
                sampled_at,
                candle_start_at,
                product_id,
                volume_btc_cumulative,
                volume_btc_delta,
                seconds_since_previous_sample,
                event_slug,
                status,
                error
            FROM btc_volume_log
            ORDER BY sampled_at DESC
            LIMIT 50
        """).fetchall()

        btc_volume_summary = conn.execute("""
            SELECT
                volume_btc_cumulative,
                volume_btc_delta
            FROM btc_volume_log
            WHERE status IN ('success', 'baseline')
            ORDER BY sampled_at DESC
            LIMIT 1
        """).fetchone()

        rules = conn.execute("""
            SELECT
                id,
                name,
                entry_price,
                stop_loss_price,
                take_profit_price,
                max_yes_entries_per_event,
                max_no_entries_per_event,
                status,
                eligible_after_event_id,
                entry_window_start_seconds_before_end,
                entry_window_end_seconds_before_end,
                schedule_timezone,
                created_at,
                updated_at
            FROM rules
            ORDER BY id DESC
            LIMIT 100
        """).fetchall()

        deals = conn.execute("""
            SELECT
                id AS deal_id,
                rule_id,
                rule_name,
                event_id,
                side,
                result,
                entry_at,
                entry_price,
                exit_at,
                exit_price,
                exit_reason,
                market_result,
                price_change_points,
                return_percent,
                investment_usd,
                shares,
                entry_fee_usd,
                exit_fee_usd,
                total_fees_usd,
                gross_pnl_usd,
                net_pnl_usd,
                gross_roi_percent,
                net_roi_percent,
                entry_btc_volume_sampled_at,
                entry_btc_volume_btc_cumulative,
                entry_btc_volume_btc_delta,
                entry_seconds_before_event_end,
                entry_liquidity_role,
                exit_liquidity_role,
                fee_calculation_source
            FROM deals
            WHERE """ + deals_rule_filter + """
            ORDER BY id DESC
            LIMIT 200
        """, deals_rule_params).fetchall()

    btc_health = get_latest_coinbase_health()
    return events, logs, btc_volume_rows, btc_volume_summary, rules, deals, btc_health


def render_dashboard_content(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> str:
    dashboard_range = normalize_dashboard_range(dashboard_range)
    overview = load_dashboard_overview(investment_usd, dashboard_range, custom_from, custom_to, rule_filter)
    time_trends = load_time_trends(investment_usd, dashboard_range, custom_from, custom_to, rule_filter)
    btc_volume_trends = load_btc_volume_trends(dashboard_range, custom_from, custom_to)
    data_quality = load_data_quality_snapshot(dashboard_range, custom_from, custom_to, rule_filter)
    return f"""
        <div class="storage-status">{html.escape(render_storage_status())}</div>
        <div class="muted">רענון אוטומטי כל 10 שניות אלא אם טופס פתוח. זמן שרת: {html.escape(now_iso())}. טווח: {html.escape(dashboard_range_label(dashboard_range, custom_from, custom_to))}</div>

        {render_dashboard_overview(overview)}
        {render_time_trends(time_trends)}
        {render_chartjs_charts(time_trends, btc_volume_trends)}
        {render_data_quality(data_quality)}
    """


def render_rules_page_content(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> str:
    dashboard_range = normalize_dashboard_range(dashboard_range)
    rules_performance = load_rules_performance(investment_usd, dashboard_range, custom_from, custom_to, rule_filter)
    _, _, _, _, rules, _, _ = load_dashboard_rows()
    return f"""
        <div class="storage-status">{html.escape(render_storage_status())}</div>
        <div class="muted">Auto refresh every 10 seconds unless a form is open. Server time: {html.escape(now_iso())}. Range: {html.escape(dashboard_range_label(dashboard_range, custom_from, custom_to))}</div>
        <div class="actions">
            {render_rule_actions()}
        </div>
        {render_rules_performance(rules_performance)}
        <div class="card">
            <h2>Rules</h2>
            {render_table(rules)}
        </div>
    """


def render_deals_page_content(
    investment_usd: Any = 1,
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> str:
    dashboard_range = normalize_dashboard_range(dashboard_range)
    risk_snapshot = load_risk_snapshot(investment_usd, dashboard_range, custom_from, custom_to, rule_filter)
    market_conditions = load_market_conditions(investment_usd, dashboard_range, custom_from, custom_to, rule_filter)
    btc_volume_deal_snapshot = load_btc_volume_deal_snapshot(dashboard_range, custom_from, custom_to, rule_filter)
    _, _, _, _, _, deals, _ = load_dashboard_rows(rule_filter)
    return f"""
        <div class="storage-status">{html.escape(render_storage_status())}</div>
        <div class="muted">Auto refresh every 10 seconds. Server time: {html.escape(now_iso())}. Range: {html.escape(dashboard_range_label(dashboard_range, custom_from, custom_to))}</div>
        {render_risk_snapshot(risk_snapshot)}
        {render_market_conditions(market_conditions)}
        {render_btc_volume_deal_snapshot(btc_volume_deal_snapshot)}
        <div class="card">
            <h2>Deals</h2>
            {render_table(deals)}
        </div>
    """


def render_market_data_page_content() -> str:
    events, logs, btc_volume_rows, btc_volume_summary, _, _, btc_health = load_dashboard_rows()
    return f"""
        <div class="storage-status">{html.escape(render_storage_status())}</div>
        <div class="muted">Auto refresh every 10 seconds. Server time: {html.escape(now_iso())}.</div>
        <div class="card">
            <h2>Events / Markets</h2>
            {render_table(events)}
        </div>
        <div class="card">
            <h2>Coinbase BTC Volume</h2>
            {render_btc_volume_summary(btc_volume_summary, btc_health)}
            {render_btc_volume_table(btc_volume_rows)}
        </div>
        <div class="card">
            <h2>Orderbook Log</h2>
            {render_table(logs)}
        </div>
    """


def render_system_page_content(
    dashboard_range: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> str:
    dashboard_range = normalize_dashboard_range(dashboard_range)
    data_quality = load_data_quality_snapshot(dashboard_range, custom_from, custom_to, rule_filter)
    system_health = load_system_health_snapshot()
    return f"""
        <div class="storage-status">{html.escape(render_storage_status())}</div>
        <div class="muted">Auto refresh every 10 seconds. Server time: {html.escape(now_iso())}. Range: {html.escape(dashboard_range_label(dashboard_range, custom_from, custom_to))}</div>
        <div class="actions">
            {render_export_actions()}
        </div>
        {render_data_quality(data_quality)}
        {render_system_health(system_health)}
    """


PAGE_CONTENT_PATHS = {
    "overview": "/dashboard-content",
    "rules": "/rules-page-content",
    "deals": "/deals-page-content",
    "market": "/market-data-content",
    "system": "/system-page-content",
}


def normalize_dashboard_page(value: Any) -> str:
    selected = str(value or "overview").strip().lower()
    return selected if selected in PAGE_CONTENT_PATHS else "overview"


def page_query(
    page: str,
    investment_usd: Any,
    range_filter: Any,
    custom_from: Any,
    custom_to: Any,
    rule_filter: Any = "all",
) -> str:
    params = {
        "page": page,
        "investment_usd": investment_usd,
        "range_filter": range_filter,
        "rule_filter": normalize_rule_filter(rule_filter),
    }
    if custom_from:
        params["custom_from"] = custom_from
    if custom_to:
        params["custom_to"] = custom_to
    return urlencode(params)


def render_dashboard_nav(
    active_page: str,
    investment_usd: Any,
    range_filter: Any,
    custom_from: Any,
    custom_to: Any,
    rule_filter: Any = "all",
) -> str:
    links = [
        ("overview", "Overview", "/"),
        ("rules", "Rules", "/rules-page"),
        ("deals", "Deals", "/deals-page"),
        ("market", "Market Data", "/market-data"),
        ("system", "System", "/system-page"),
    ]
    items = []
    for page, label, path in links:
        query = page_query(page, investment_usd, range_filter, custom_from, custom_to, rule_filter)
        active_class = " active" if page == active_page else ""
        items.append(
            f"<a class=\"nav-link{active_class}\" href=\"{path}?{html.escape(query)}\">{html.escape(label)}</a>"
        )
    return f"<nav class=\"top-nav\">{''.join(items)}</nav>"


def render_page_content(
    page: Any,
    investment_usd: Any = 1,
    range_filter: Any = "all",
    custom_from: Any = None,
    custom_to: Any = None,
    rule_filter: Any = "all",
) -> str:
    selected = normalize_dashboard_page(page)
    if selected == "rules":
        return render_rules_page_content(investment_usd, range_filter, custom_from, custom_to, rule_filter)
    if selected == "deals":
        return render_deals_page_content(investment_usd, range_filter, custom_from, custom_to, rule_filter)
    if selected == "market":
        return render_market_data_page_content()
    if selected == "system":
        return render_system_page_content(range_filter, custom_from, custom_to, rule_filter)
    return render_dashboard_content(investment_usd, range_filter, custom_from, custom_to, rule_filter)


@app.get("/dashboard-content", response_class=HTMLResponse)
def dashboard_content(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return render_dashboard_content(investment_usd, range_filter, custom_from, custom_to, rule_filter)


@app.get("/rules-page-content", response_class=HTMLResponse)
def rules_page_content(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return render_rules_page_content(investment_usd, range_filter, custom_from, custom_to, rule_filter)


@app.get("/deals-page-content", response_class=HTMLResponse)
def deals_page_content(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return render_deals_page_content(investment_usd, range_filter, custom_from, custom_to, rule_filter)


@app.get("/market-data-content", response_class=HTMLResponse)
def market_data_content() -> str:
    return render_market_data_page_content()


@app.get("/system-page-content", response_class=HTMLResponse)
def system_page_content(
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return render_system_page_content(range_filter, custom_from, custom_to, rule_filter)


@app.get("/", response_class=HTMLResponse)
def dashboard(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
    page: str = "overview",
) -> str:
    active_page = normalize_dashboard_page(page)
    content_path = PAGE_CONTENT_PATHS[active_page]
    page_content = render_page_content(active_page, investment_usd, range_filter, custom_from, custom_to, rule_filter)
    nav = render_dashboard_nav(active_page, investment_usd, range_filter, custom_from, custom_to, rule_filter)
    return f"""
    <!doctype html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Polymarket BTC Collector</title>
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 24px;
                background: #f7f7f7;
                color: #111;
            }}
            h1, h2 {{
                margin-bottom: 8px;
            }}
            .top-nav {{
                display: flex;
                flex-wrap: wrap;
                gap: 8px;
                margin: 12px 0 20px 0;
            }}
            .nav-link {{
                display: inline-flex;
                align-items: center;
                min-height: 36px;
                padding: 0 12px;
                border: 1px solid #d4d4d4;
                border-radius: 6px;
                color: #222;
                background: #fff;
                text-decoration: none;
                font-size: 14px;
                font-weight: 700;
            }}
            .nav-link.active {{
                color: #fff;
                background: #111;
                border-color: #111;
            }}
            .card {{
                background: white;
                border: 1px solid #ddd;
                border-radius: 10px;
                padding: 16px;
                margin-bottom: 24px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            }}
            .overview-header {{
                display: flex;
                gap: 16px;
                align-items: flex-start;
                justify-content: space-between;
                flex-wrap: wrap;
            }}
            .investment-form {{
                display: flex;
                gap: 8px;
                align-items: flex-end;
                flex-wrap: wrap;
            }}
            .investment-form label {{
                margin: 0;
                min-width: 180px;
            }}
            .metric-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
                gap: 10px;
                margin: 12px 0;
            }}
            .metric {{
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 12px;
                background: #fafafa;
                min-height: 78px;
            }}
            .metric-label {{
                font-size: 12px;
                color: #666;
                margin-bottom: 6px;
            }}
            .metric-value {{
                font-size: 22px;
                font-weight: 700;
                line-height: 1.2;
                overflow-wrap: anywhere;
            }}
            .metric-note {{
                margin-top: 6px;
                font-size: 12px;
                color: #777;
            }}
            .actions {{
                margin: 12px 0 20px 0;
            }}
            .inline-form {{
                display: inline-block;
                margin: 0 8px 8px 0;
            }}
            a.button, button.button {{
                display: inline-block;
                padding: 10px 14px;
                background: #111;
                color: white;
                text-decoration: none;
                border-radius: 6px;
                border: 0;
                cursor: pointer;
                font: inherit;
            }}
            a.button.disabled, button.button:disabled {{
                background: #888;
                pointer-events: none;
                cursor: default;
            }}
            button.secondary, a.secondary {{
                background: #555;
            }}
            .export-status {{
                margin-top: 4px;
            }}
            .storage-status {{
                margin: 0 0 6px 0;
                font-size: 14px;
                font-weight: 700;
            }}
            .modal[hidden] {{
                display: none;
            }}
            .modal {{
                position: fixed;
                inset: 0;
                background: rgba(0, 0, 0, 0.45);
                display: flex;
                align-items: center;
                justify-content: center;
                z-index: 10;
            }}
            .modal-panel {{
                background: white;
                border-radius: 8px;
                padding: 18px;
                width: min(520px, calc(100vw - 32px));
                box-shadow: 0 8px 24px rgba(0,0,0,0.18);
            }}
            .modal-panel.small {{
                width: min(360px, calc(100vw - 32px));
            }}
            label {{
                display: block;
                margin: 10px 0;
                font-size: 13px;
            }}
            input, select {{
                box-sizing: border-box;
                width: 100%;
                margin-top: 4px;
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 6px;
                font: inherit;
            }}
            .modal-actions {{
                margin-top: 14px;
            }}
            .error {{
                color: #9b1c1c;
                min-height: 20px;
                font-size: 13px;
            }}
            .table-wrap {{
                overflow-x: auto;
                max-height: 520px;
                border: 1px solid #ddd;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                min-width: 1600px;
                font-size: 13px;
            }}
            .compact-table table {{
                min-width: 1050px;
            }}
            .risk-table table {{
                min-width: 720px;
            }}
            .condition-table table {{
                min-width: 760px;
            }}
            .health-table table {{
                min-width: 360px;
            }}
            .trends-table table {{
                min-width: 780px;
            }}
            .quality-table table {{
                min-width: 520px;
            }}
            .charts-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
                gap: 18px;
            }}
            .canvas-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                gap: 18px;
            }}
            .canvas-panel {{
                min-height: 280px;
            }}
            .canvas-panel.wide {{
                grid-column: 1 / -1;
            }}
            .canvas-panel canvas {{
                width: 100% !important;
                max-height: 260px;
            }}
            .btc-volume-deals-table table {{
                min-width: 980px;
            }}
            .chart-row {{
                display: grid;
                grid-template-columns: 92px 1fr 80px;
                gap: 8px;
                align-items: center;
                margin: 8px 0;
                font-size: 13px;
            }}
            .chart-label, .chart-value {{
                overflow-wrap: anywhere;
            }}
            .bar-track {{
                height: 12px;
                background: #ececec;
                border-radius: 6px;
                overflow: hidden;
                min-width: 90px;
            }}
            .bar {{
                height: 100%;
                border-radius: 6px;
            }}
            .bar.positive {{
                background: #147a3f;
            }}
            .bar.negative {{
                background: #b42318;
            }}
            .bar.neutral {{
                background: #2f5f98;
            }}
            .badge {{
                display: inline-block;
                min-width: 26px;
                padding: 3px 8px;
                border-radius: 6px;
                text-align: center;
                color: white;
                background: #555;
            }}
            .badge.ok {{
                background: #147a3f;
            }}
            .badge.warning {{
                background: #98690c;
            }}
            .badge.error {{
                background: #b42318;
            }}
            h3 {{
                margin: 16px 0 8px 0;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 7px;
                white-space: nowrap;
                vertical-align: top;
            }}
            th {{
                background: #f0f0f0;
                position: sticky;
                top: 0;
                z-index: 1;
            }}
            .muted {{
                color: #666;
                font-size: 13px;
            }}
        </style>
    </head>
    <body data-content-path="{html.escape(content_path)}">
        <h1>Polymarket BTC Collector</h1>
        {nav}
        <div id="dashboard-content">
            {page_content}
        </div>
        {render_dashboard_scripts()}
    </body>
    </html>
    """


@app.get("/rules-page", response_class=HTMLResponse)
def rules_page(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return dashboard(investment_usd, range_filter, custom_from, custom_to, rule_filter, "rules")


@app.get("/deals-page", response_class=HTMLResponse)
def deals_page(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return dashboard(investment_usd, range_filter, custom_from, custom_to, rule_filter, "deals")


@app.get("/market-data", response_class=HTMLResponse)
def market_data_page(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return dashboard(investment_usd, range_filter, custom_from, custom_to, rule_filter, "market")


@app.get("/system-page", response_class=HTMLResponse)
def system_page(
    investment_usd: float = 1.0,
    range_filter: str = "all",
    custom_from: str = "",
    custom_to: str = "",
    rule_filter: str = "all",
) -> str:
    return dashboard(investment_usd, range_filter, custom_from, custom_to, rule_filter, "system")


@app.post("/rules")
def api_create_rule(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    try:
        row = create_rule(payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except sqlite3.Error:
        raise HTTPException(status_code=500, detail="Database error while creating rule")
    with get_conn() as conn:
        return rules_to_dicts(conn, [row])[0]


@app.get("/rules")
def api_get_rules() -> list[dict[str, Any]]:
    try:
        with get_conn() as conn:
            rows = conn.execute("""
                SELECT *
                FROM rules
                ORDER BY id DESC
            """).fetchall()
            return rules_to_dicts(conn, rows)
    except sqlite3.Error:
        raise HTTPException(status_code=500, detail="Database error while loading rules")


@app.post("/rules/{rule_id}/deactivate")
def api_deactivate_rule(rule_id: int) -> dict[str, Any]:
    try:
        row, message = deactivate_rule(rule_id)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Rule {rule_id} does not exist")
    except sqlite3.Error:
        raise HTTPException(status_code=500, detail="Database error while deactivating rule")
    return {"message": message, "rule": row_to_dict(row)}


@app.get("/deals")
def api_get_deals() -> list[dict[str, Any]]:
    try:
        with get_conn() as conn:
            rows = conn.execute("""
                SELECT *
                FROM deals
                ORDER BY id DESC
            """).fetchall()
    except sqlite3.Error:
        raise HTTPException(status_code=500, detail="Database error while loading deals")
    return [row_to_dict(row) for row in rows]


@app.post("/generate.xlsx")
def generate_xlsx(background_tasks: BackgroundTasks):
    with export_state_lock:
        if export_state.get("status") == "running":
            return HTMLResponse("Export already in progress", status_code=409)
        export_state.update({
            "status": "running",
            "started_at": now_iso(),
            "finished_at": None,
            "error": None,
            "row_counts": None,
        })
        background_tasks.add_task(run_xlsx_export)

    return RedirectResponse("/", status_code=303)


@app.get("/download.xlsx")
def download_xlsx():
    export_path = latest_export_path()
    if not export_path:
        return HTMLResponse(
            "<p>No Excel export is ready yet. Generate one first.</p>",
            status_code=404,
        )

    return FileResponse(
        export_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=export_path.name,
    )


@app.get("/health")
def health() -> dict[str, Any]:
    coinbase_health = get_latest_coinbase_health()
    return {
        "ok": True,
        "time": now_iso(),
        "db": str(DB_PATH),
        "active_market": active_market.get("event_slug") if active_market else None,
        "coinbase_volume_last_sample_at": coinbase_health.get("last_sample_at"),
        "coinbase_volume_last_success_at": coinbase_health.get("last_success_at"),
        "coinbase_volume_collector_status": coinbase_health.get("status"),
        "coinbase_volume_last_error": coinbase_health.get("last_error"),
    }
