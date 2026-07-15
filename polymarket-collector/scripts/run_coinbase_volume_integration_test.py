import argparse
import asyncio
import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import app  # noqa: E402


def fetch_dataframes(db_path: Path) -> dict[str, pd.DataFrame]:
    with sqlite3.connect(db_path) as conn:
        return {
            "events": pd.read_sql_query("SELECT * FROM events ORDER BY local_event_id ASC", conn),
            "orderbook_log": pd.read_sql_query("SELECT * FROM orderbook_log ORDER BY id ASC", conn),
            "btc_volume_log": pd.read_sql_query("SELECT * FROM btc_volume_log ORDER BY sampled_at ASC", conn),
        }


def write_excel(db_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"polymarket_coinbase_10min_test_{timestamp}.xlsx"

    dataframes = fetch_dataframes(db_path)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in dataframes.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(xlsx_path, read_only=True)
    workbook.close()
    return xlsx_path


def summarize(db_path: Path, xlsx_path: Path, started_at: datetime, finished_at: datetime) -> dict:
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        coinbase_count = conn.execute("SELECT COUNT(*) FROM btc_volume_log").fetchone()[0]
        polymarket_count = conn.execute("SELECT COUNT(*) FROM orderbook_log").fetchone()[0]
        unique_candles = conn.execute(
            "SELECT COUNT(DISTINCT candle_start_at) FROM btc_volume_log"
        ).fetchone()[0]
        baseline_count = conn.execute(
            "SELECT COUNT(*) FROM btc_volume_log WHERE status = 'baseline'"
        ).fetchone()[0]
        valid_delta_count = conn.execute(
            "SELECT COUNT(*) FROM btc_volume_log WHERE status = 'success' AND volume_btc_delta IS NOT NULL"
        ).fetchone()[0]
        error_count = conn.execute(
            "SELECT COUNT(*) FROM btc_volume_log WHERE status = 'error'"
        ).fetchone()[0]
        duplicate_buckets = conn.execute("""
            SELECT COUNT(*)
            FROM (
                SELECT product_id, candle_start_at, sample_bucket_at, COUNT(*) AS count
                FROM btc_volume_log
                GROUP BY product_id, candle_start_at, sample_bucket_at
                HAVING count > 1
            )
        """).fetchone()[0]
        intervals = [
            row[0]
            for row in conn.execute("""
                SELECT seconds_since_previous_sample
                FROM btc_volume_log
                WHERE seconds_since_previous_sample IS NOT NULL
            """).fetchall()
        ]

    workbook = load_workbook(xlsx_path, read_only=True)
    sheet_names = workbook.sheetnames
    row_counts = {
        sheet_name: workbook[sheet_name].max_row - 1
        for sheet_name in sheet_names
    }
    workbook.close()

    return {
        "start_time": started_at.isoformat(),
        "end_time": finished_at.isoformat(),
        "actual_duration_seconds": (finished_at - started_at).total_seconds(),
        "db_path": str(db_path),
        "coinbase_sample_count": coinbase_count,
        "polymarket_sample_count": polymarket_count,
        "unique_candle_count": unique_candles,
        "baseline_count": baseline_count,
        "valid_delta_count": valid_delta_count,
        "error_count": error_count,
        "duplicate_bucket_count": duplicate_buckets,
        "min_sample_interval": min(intervals) if intervals else None,
        "avg_sample_interval": sum(intervals) / len(intervals) if intervals else None,
        "max_sample_interval": max(intervals) if intervals else None,
        "xlsx_path": str(xlsx_path),
        "xlsx_size_bytes": xlsx_path.stat().st_size,
        "xlsx_sheet_names": sheet_names,
        "xlsx_row_counts": row_counts,
        "xlsx_reopen_success": True,
    }


async def run(duration_seconds: int, db_path: Path, output_dir: Path) -> dict:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    app.DB_PATH = db_path
    app.init_db()

    tasks = [
        asyncio.create_task(app.event_collector_loop()),
        asyncio.create_task(app.orderbook_collector_loop()),
        asyncio.create_task(app.coinbase_volume_collector_loop()),
    ]

    started_at = datetime.now(timezone.utc)
    try:
        await asyncio.sleep(duration_seconds)
    finally:
        for task in tasks:
            task.cancel()
        await asyncio.gather(*tasks, return_exceptions=True)

    finished_at = datetime.now(timezone.utc)
    xlsx_path = write_excel(db_path, output_dir)
    return summarize(db_path, xlsx_path, started_at, finished_at)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--duration-seconds", type=int, default=600)
    parser.add_argument("--db-path", type=Path)
    parser.add_argument("--output-dir", type=Path, default=PROJECT_ROOT / "output")
    args = parser.parse_args()

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    db_path = args.db_path or PROJECT_ROOT / "output" / f"coinbase_volume_test_{timestamp}.sqlite3"
    summary = asyncio.run(run(args.duration_seconds, db_path, args.output_dir))
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
