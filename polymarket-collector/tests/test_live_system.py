import asyncio
import hashlib
import os
import sqlite3
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import app
import live_app
from live.config import LiveConfig, redact_mapping
from live.public_client import MockPublicClobClient
from live.repository import LiveRepository
from live.router import configure, refresh_public_market_metadata

try:
    from argon2 import PasswordHasher
except ImportError:  # pragma: no cover
    PasswordHasher = None


class LiveSystemTests(unittest.TestCase):
    def setUp(self):
        self.original_db_path = app.DB_PATH
        self.original_env = {key: os.environ.get(key) for key in [
            "LIVE_DB_PATH",
            "LIVE_BACKUP_DIR",
            "LIVE_LOGIN_USERNAME",
            "LIVE_LOGIN_PASSWORD_HASH",
            "LIVE_SESSION_SECRET",
            "LIVE_OPERATOR_TOKEN",
            "LIVE_MODULE_ENABLED",
            "LIVE_ADAPTER",
            "LIVE_KILL_SWITCH",
            "LIVE_MARKET_DATA_STALE_AFTER_SECONDS",
            "LIVE_RECONCILIATION_MAX_AGE_SECONDS",
        ]}
        self.temp_dir = tempfile.TemporaryDirectory()
        self.demo_db_path = Path(self.temp_dir.name) / "demo.sqlite3"
        self.db_path = Path(self.temp_dir.name) / "live.sqlite3"
        self.backup_dir = Path(self.temp_dir.name) / "backups"
        app.DB_PATH = self.demo_db_path
        app.init_db()
        os.environ["LIVE_DB_PATH"] = str(self.db_path)
        os.environ["LIVE_BACKUP_DIR"] = str(self.backup_dir)
        os.environ["LIVE_LOGIN_USERNAME"] = "Admin@system.com"
        os.environ["LIVE_LOGIN_PASSWORD_HASH"] = "sha256:" + hashlib.sha256(b"pw").hexdigest()
        os.environ["LIVE_SESSION_SECRET"] = "test-session-secret"
        os.environ["LIVE_OPERATOR_TOKEN"] = "test-token"
        os.environ["LIVE_MODULE_ENABLED"] = "true"
        os.environ["LIVE_ADAPTER"] = "mock"
        os.environ["LIVE_KILL_SWITCH"] = "true"
        os.environ["LIVE_MARKET_DATA_STALE_AFTER_SECONDS"] = "10000"
        os.environ["LIVE_RECONCILIATION_MAX_AGE_SECONDS"] = "10000"
        self.config = LiveConfig(
            live_module_enabled=True,
            live_adapter="mock",
            operator_token="test-token",
            login_username="Admin@system.com",
            login_password_hash="sha256:" + hashlib.sha256(b"pw").hexdigest(),
            session_secret="test-session-secret",
            session_ttl_seconds=0,
            live_kill_switch_default=True,
            max_market_data_age_seconds=10_000,
            max_reconciliation_age_seconds=10_000,
            live_db_path=str(self.db_path),
            backup_dir=str(self.backup_dir),
        )
        configure(self.db_path, self.config)

    def tearDown(self):
        app.DB_PATH = self.original_db_path
        for key, value in self.original_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        self.temp_dir.cleanup()

    def client(self):
        return TestClient(live_app.app, base_url="https://testserver")

    def demo_client(self):
        return TestClient(app.app)

    def login(self, client):
        response = client.post("/live/login", json={"username": "Admin@system.com", "password": "pw"})
        self.assertEqual(response.status_code, 200)
        return response.headers["X-Live-CSRF-Token"]

    def auth_headers(self, csrf: str, reauth: bool = False) -> dict[str, str]:
        headers = {"X-Live-Operator-Token": "test-token", "X-Live-CSRF-Token": csrf}
        if reauth:
            headers["X-Live-Reauth-Password"] = "pw"
        return headers

    def test_config_defaults_are_fail_closed(self):
        defaults = LiveConfig()
        self.assertEqual(defaults.trading_mode, "DEMO")
        self.assertFalse(defaults.live_module_enabled)
        self.assertFalse(defaults.live_trading_enabled)
        self.assertFalse(defaults.live_order_submission_enabled)
        self.assertEqual(defaults.live_adapter, "mock")
        self.assertTrue(defaults.live_kill_switch_default)
        self.assertEqual(defaults.login_username, "Admin@system.com")
        self.assertEqual(defaults.session_ttl_seconds, 0)
        self.assertFalse(defaults.real_submission_armed())

        armed = LiveConfig(
            trading_mode="LIVE",
            live_module_enabled=True,
            live_trading_enabled=True,
            live_order_submission_enabled=True,
            live_adapter="polymarket",
        )
        self.assertTrue(armed.real_submission_armed())

        redacted = redact_mapping({"POLYMARKET_API_SECRET": "abcdef", "safe": "value"})
        self.assertEqual(redacted["POLYMARKET_API_SECRET"], "ab****ef")
        self.assertEqual(redacted["safe"], "value")

    def test_live_migration_is_idempotent_and_separate_from_demo_tables(self):
        demo_conn = sqlite3.connect(self.demo_db_path)
        try:
            demo_tables = {row[0] for row in demo_conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        finally:
            demo_conn.close()
        self.assertIn("events", demo_tables)
        self.assertIn("rules", demo_tables)
        self.assertNotIn("live_orders", demo_tables)

        repo = LiveRepository(self.db_path)
        repo.migrate()
        repo.migrate()
        conn = sqlite3.connect(self.db_path)
        try:
            tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        finally:
            conn.close()
        self.assertNotIn("events", tables)
        self.assertNotIn("rules", tables)
        self.assertNotIn("deals", tables)
        self.assertIn("live_orders", tables)
        self.assertIn("live_order_fills", tables)
        self.assertIn("live_reconciliation_runs", tables)
        self.assertIn("live_system_state", tables)
        self.assertIn("live_backups", tables)

    def test_live_routes_health_and_auth_blocks_writes_without_token(self):
        client = self.client()
        self.assertEqual(client.get("/health").json(), {"status": "ok"})
        dashboard = client.get("/live", follow_redirects=False)
        self.assertEqual(dashboard.status_code, 303)
        self.assertEqual(dashboard.headers["location"], "/live/login")
        unauthenticated = client.get("/live/health")
        self.assertEqual(unauthenticated.status_code, 401)
        csrf = self.login(client)
        self.assertEqual(client.get("/live", follow_redirects=False).status_code, 200)
        logged_in_login = client.get("/live/login", follow_redirects=False)
        self.assertEqual(logged_in_login.status_code, 303)
        self.assertEqual(logged_in_login.headers["location"], "/live")
        response = client.get("/live/health")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["kill_switch_active"])
        self.assertEqual(payload["adapter"], "mock")
        self.assertFalse(payload["config"]["real_submission_armed"])

        blocked = client.post("/live/kill-switch/deactivate")
        self.assertEqual(blocked.status_code, 403)

        missing_reauth = client.post("/live/kill-switch/deactivate", headers=self.auth_headers(csrf), follow_redirects=False)
        self.assertEqual(missing_reauth.status_code, 403)

        ok = client.post("/live/kill-switch/deactivate", headers=self.auth_headers(csrf, reauth=True), follow_redirects=False)
        self.assertEqual(ok.status_code, 303)
        self.assertFalse(client.get("/live/health").json()["kill_switch_active"])

    def test_root_login_alias_and_styled_login_page(self):
        client = self.client()
        alias = client.get("/login", follow_redirects=False)
        self.assertEqual(alias.status_code, 307)
        self.assertEqual(alias.headers["location"], "/live/login")
        page = client.get("/live/login")
        self.assertEqual(page.status_code, 200)
        self.assertIn("Polymarket LIVE", page.text)
        self.assertIn("Control Center", page.text)
        self.assertIn('credentials:"same-origin"', page.text)

    def test_argon2id_password_verification_and_revoke_all_sessions(self):
        if PasswordHasher is None:
            self.skipTest("argon2 dependency missing")
        argon_hash = PasswordHasher().hash("pw")
        os.environ["LIVE_LOGIN_PASSWORD_HASH"] = argon_hash
        configure(self.db_path, LiveConfig(
            live_module_enabled=True,
            live_adapter="mock",
            operator_token="test-token",
            login_username="Admin@system.com",
            login_password_hash=argon_hash,
            session_secret="test-session-secret",
            live_db_path=str(self.db_path),
            backup_dir=str(self.backup_dir),
        ))
        client = self.client()
        csrf = self.login(client)
        self.assertEqual(client.get("/live/health").status_code, 200)
        revoked = client.post("/live/sessions/revoke-all", headers=self.auth_headers(csrf, reauth=True))
        self.assertEqual(revoked.status_code, 200)
        self.assertEqual(client.get("/live/health").status_code, 401)

    def test_mock_order_lifecycle_and_fills(self):
        client = self.client()
        csrf = self.login(client)
        headers = self.auth_headers(csrf)
        client.post("/live/kill-switch/deactivate", headers=self.auth_headers(csrf, reauth=True), follow_redirects=False)
        result = client.post(
            "/live/orders/mock",
            headers=headers,
            json={
                "idempotency_key": "entry-1",
                "condition_id": "condition-1",
                "token_id": "yes-token",
                "requested_price": 0.50,
                "requested_amount_usd": 1,
            },
        )
        self.assertEqual(result.status_code, 200)
        data = result.json()
        self.assertEqual(data["order"]["status"], "filled")
        self.assertEqual(data["order"]["polymarket_order_id"], "mock-entry-1")
        self.assertEqual(len(client.get("/live/fills").json()), 1)

        duplicate = client.post(
            "/live/orders/mock",
            headers=headers,
            json={
                "idempotency_key": "entry-1",
                "condition_id": "condition-1",
                "token_id": "yes-token",
                "requested_price": 0.50,
                "requested_amount_usd": 1,
            },
        )
        self.assertEqual(duplicate.status_code, 200)
        self.assertEqual(duplicate.json()["risk"]["reason_code"], "DUPLICATE_IDEMPOTENCY")

    def test_risk_blocks_kill_switch_and_partial_fill_policy(self):
        client = self.client()
        csrf = self.login(client)
        headers = self.auth_headers(csrf)
        blocked = client.post(
            "/live/orders/mock",
            headers=headers,
            json={
                "idempotency_key": "blocked-1",
                "condition_id": "condition-1",
                "token_id": "yes-token",
                "requested_price": 0.50,
                "requested_amount_usd": 1,
            },
        )
        self.assertEqual(blocked.status_code, 200)
        self.assertEqual(blocked.json()["order"]["status"], "blocked")
        self.assertEqual(blocked.json()["risk"]["reason_code"], "KILL_SWITCH_ACTIVE")

        client.post("/live/kill-switch/deactivate", headers=self.auth_headers(csrf, reauth=True), follow_redirects=False)
        fak = client.post(
            "/live/orders/mock",
            headers=headers,
            json={
                "idempotency_key": "fak-1",
                "condition_id": "condition-1",
                "token_id": "yes-token",
                "order_type": "FAK",
                "requested_price": 0.50,
                "requested_amount_usd": 1,
                "mock_scenario": "partial",
            },
        )
        self.assertEqual(fak.status_code, 200)
        self.assertEqual(fak.json()["order"]["status"], "blocked")
        self.assertEqual(fak.json()["risk"]["reason_code"], "PARTIAL_FILLS_DISABLED")

    def test_public_metadata_and_websocket_resolution_fixture(self):
        metadata = asyncio.run(refresh_public_market_metadata("condition-1", use_mock=True))
        self.assertEqual(metadata["token_mapping_status"], "matched")
        client = self.client()
        csrf = self.login(client)
        headers = self.auth_headers(csrf)
        message = {
            "event_type": "market_resolved",
            "condition_id": "condition-1",
            "winning_asset_id": "yes-token",
            "winning_outcome": "Yes",
        }
        first = client.post("/live/market-ws/fixture", headers=headers, json=message)
        second = client.post("/live/market-ws/fixture", headers=headers, json=message)
        self.assertEqual(first.status_code, 200)
        self.assertTrue(first.json()["stored"])
        self.assertFalse(second.json()["stored"])
        market = client.get("/live/markets").json()[0]
        self.assertEqual(market["market_resolved"], 1)
        self.assertEqual(market["winning_asset_id"], "yes-token")

    def test_user_websocket_fixture_reconciliation_and_export(self):
        client = self.client()
        csrf = self.login(client)
        headers = self.auth_headers(csrf)
        user_msg = {"type": "trade", "condition_id": "condition-1", "trade_id": "trade-1", "status": "MATCHED"}
        self.assertTrue(client.post("/live/user-ws/fixture", headers=headers, json=user_msg).json()["stored"])
        recon = client.post("/live/reconciliation/run", headers=headers, follow_redirects=False)
        self.assertEqual(recon.status_code, 303)
        self.assertGreaterEqual(len(client.get("/live/reconciliation").json()), 1)

        generated = client.post("/live/export/generate", headers=headers, follow_redirects=False)
        self.assertEqual(generated.status_code, 303)
        # BackgroundTasks run before TestClient returns.
        download = client.get("/live/export/download")
        self.assertEqual(download.status_code, 200)
        workbook = load_workbook(BytesIO(download.content), read_only=True)
        try:
            self.assertIn("live_orders", workbook.sheetnames)
            self.assertIn("live_audit_log", workbook.sheetnames)
            self.assertIn("live_dry_runs", workbook.sheetnames)
        finally:
            workbook.close()

    def test_account_identity_secret_readiness_and_dry_run(self):
        client = self.client()
        csrf = self.login(client)
        headers = self.auth_headers(csrf)
        configure(self.db_path, LiveConfig(
            live_module_enabled=True,
            live_adapter="mock",
            operator_token="test-token",
            login_username="Admin@system.com",
            login_password_hash="sha256:" + hashlib.sha256(b"pw").hexdigest(),
            session_secret="test-session-secret",
            profile_address="0xcE075637152167517e1492FcF5ff2D131686ee38",
            live_kill_switch_default=True,
            max_market_data_age_seconds=10_000,
            max_reconciliation_age_seconds=10_000,
            live_db_path=str(self.db_path),
            backup_dir=str(self.backup_dir),
        ))
        csrf = self.login(client)
        headers = self.auth_headers(csrf)
        identity = client.post("/live/account/public-refresh?use_mock=true", headers=headers)
        self.assertEqual(identity.status_code, 200)
        self.assertEqual(identity.json()["account_identity_status"], "UNVERIFIED")
        self.assertNotIn("raw_public_payload", identity.json())

        dry = client.post("/live/dry-run", headers=headers, json={
            "idempotency_key": "dry-1",
            "condition_id": "condition-1",
            "token_id": "yes-token",
            "requested_price": 0.50,
            "requested_amount_usd": 1,
            "purpose": "entry",
        })
        self.assertEqual(dry.status_code, 200)
        self.assertEqual(dry.json()["final_decision"], "BLOCKED")
        self.assertEqual(client.get("/live/secrets/readiness").status_code, 200)

    def test_demo_routes_still_work_with_live_module_enabled(self):
        client = self.demo_client()
        response = client.post("/rules", json={
            "name": "demo rule",
            "entry_price": "0.77",
            "stop_loss_price": "0.60",
            "take_profit_price": "0.90",
            "max_yes_entries_per_event": 1,
            "max_no_entries_per_event": 1,
            "status": "active",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(client.get("/rules").status_code, 200)
        self.assertEqual(client.get("/deals").status_code, 200)
        self.assertEqual(client.get("/health").status_code, 200)
        self.assertIn("Polymarket BTC Collector", client.get("/").text)
        self.assertEqual(client.get("/live").status_code, 404)

    def test_maintenance_drain_readiness_cancel_and_backup(self):
        client = self.client()
        csrf = self.login(client)
        headers = self.auth_headers(csrf)

        drain = client.post("/live/maintenance/drain", headers=headers)
        self.assertEqual(drain.status_code, 200)
        self.assertEqual(drain.json()["mode"], "DRAINING")

        ready = client.post("/live/maintenance/readiness", headers=headers)
        self.assertEqual(ready.status_code, 200)
        self.assertTrue(ready.json()["stop_ready"])

        cancel = client.post("/live/maintenance/cancel", headers=headers)
        self.assertEqual(cancel.status_code, 200)

        backup = client.post("/live/backup/create", headers=headers)
        self.assertEqual(backup.status_code, 200)
        self.assertEqual(backup.json()["status"], "ok")
        self.assertTrue(Path(backup.json()["path"]).exists())
        self.assertGreater(len(backup.json()["sha256"]), 20)

    def test_live_product_ui_screens_render(self):
        client = self.client()
        self.login(client)
        views = ["overview", "operations", "risk", "logs", "market", "account", "dry-run", "reconciliation", "orders", "deployment", "maintenance"]
        for view in views:
            response = client.get(f"/live?view={view}")
            self.assertEqual(response.status_code, 200)
            self.assertIn("Polymarket LIVE", response.text)
            self.assertIn("Control Center", response.text)
            self.assertIn('name="viewport"', response.text)
        self.assertIn("Dry Run Studio", client.get("/live?view=dry-run").text)
        self.assertIn("Deployment Checklist", client.get("/live?view=deployment").text)
        self.assertIn("Safe Maintenance", client.get("/live?view=maintenance").text)
        self.assertIn("attr(data-label)", client.get("/live?view=risk").text)


if __name__ == "__main__":
    unittest.main()
