import sqlite3
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import app
from scripts import backfill_deal_btc_volume_snapshots


def valid_rule(**overrides):
    payload = {
        "name": "demo rule",
        "entry_price": "0.77",
        "stop_loss_price": "0.60",
        "take_profit_price": "0.90",
        "max_yes_entries_per_event": 1,
        "max_no_entries_per_event": 1,
        "status": "active",
    }
    payload.update(overrides)
    return payload


def orderbook(event_id, *, yes_ask=None, yes_bid=None, no_ask=None, no_bid=None, sampled_at=None):
    return {
        "sampled_at": sampled_at or "2026-07-17T09:00:01+00:00",
        "sampled_at_local": "17/07/2026 12:00:01",
        "event_slug": event_id,
        "condition_id": f"condition-{event_id}",
        "up_token_id": "yes-token",
        "down_token_id": "no-token",
        "up_best_ask": yes_ask,
        "up_best_bid": yes_bid,
        "down_best_ask": no_ask,
        "down_best_bid": no_bid,
        "up_last_trade_price": None,
        "down_last_trade_price": None,
        "up_spread": None,
        "down_spread": None,
        "up_midpoint": None,
        "down_midpoint": None,
        "raw_up_timestamp": None,
        "raw_down_timestamp": None,
        "up_volume_shares_10s": 0.0,
        "down_volume_shares_10s": 0.0,
        "up_volume_usdc_10s": 0.0,
        "down_volume_usdc_10s": 0.0,
        "trades_count_10s": 0,
        "trades_window_start": None,
        "trades_window_start_local": None,
        "trades_window_end": None,
        "trades_window_end_local": None,
        "trades_error": None,
        "status": "success",
        "error": None,
    }


def btc_volume(sampled_at, *, event_id="event-a", cumulative=12.5, delta=6.7, status="success"):
    return {
        "sampled_at": sampled_at,
        "sample_bucket_at": sampled_at,
        "candle_start_at": "2026-07-17T08:55:00+00:00",
        "product_id": "BTC-USD",
        "granularity_seconds": 300,
        "volume_btc_cumulative": cumulative,
        "volume_btc_delta": delta,
        "seconds_since_previous_sample": 30,
        "event_slug": event_id,
        "condition_id": f"condition-{event_id}",
        "source": "coinbase",
        "status": status,
        "error": None,
    }


def market_row(event_id, *, start_time="2026-07-17T09:00:00Z", end_time="2026-07-17T09:05:00Z", active=1, closed=0, status="open"):
    return {
        "polymarket_event_id": f"poly-{event_id}",
        "polymarket_market_id": f"market-{event_id}",
        "condition_id": f"condition-{event_id}",
        "event_slug": event_id,
        "market_slug": event_id,
        "title": "title",
        "question": "question",
        "event_url": "url",
        "start_time": start_time,
        "start_time_local": app.format_local_datetime(start_time),
        "end_time": end_time,
        "end_time_local": app.format_local_datetime(end_time),
        "yes_token_id": "yes-token",
        "no_token_id": "no-token",
        "outcomes": '["Up", "Down"]',
        "outcome_prices": '["1", "0"]',
        "active": active,
        "closed": closed,
        "enable_order_book": 1,
        "accepting_orders": 1 if active else 0,
        "created_at_poly": "2026-07-17T08:59:00Z",
        "created_at_poly_local": "17/07/2026 11:59:00",
        "status": status,
        "notes": "",
        "raw_json": "{}",
    }


def expected_net_pnl(entry_price, exit_price):
    return float(app.calculate_demo_deal_financials(entry_price, exit_price)["net_pnl_usd"])


class RuleDealTests(unittest.TestCase):
    def setUp(self):
        self.original_db_path = app.DB_PATH
        self.original_export_dir = app.EXPORT_DIR
        self.original_active_market = app.active_market
        self.temp_dir = tempfile.TemporaryDirectory()
        temp_path = Path(self.temp_dir.name)
        app.DB_PATH = temp_path / "test.sqlite3"
        app.EXPORT_DIR = temp_path / "output"
        app.active_market = None
        app.init_db()

    def tearDown(self):
        app.DB_PATH = self.original_db_path
        app.EXPORT_DIR = self.original_export_dir
        app.active_market = self.original_active_market
        self.temp_dir.cleanup()

    def fetch_deals(self):
        with app.get_conn() as conn:
            return conn.execute("SELECT * FROM deals ORDER BY id").fetchall()

    def test_rule_validation_and_deactivation_api(self):
        client = TestClient(app.app)

        response = client.post("/rules", json=valid_rule())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "active")

        invalid_cases = [
            valid_rule(name=""),
            valid_rule(entry_price="0.5"),
            valid_rule(stop_loss_price="0.80"),
            valid_rule(take_profit_price="0.70"),
            valid_rule(max_yes_entries_per_event=-1),
            valid_rule(status="paused"),
        ]
        for payload in invalid_cases:
            self.assertEqual(client.post("/rules", json=payload).status_code, 400)

        inactive = client.post("/rules", json=valid_rule(name="inactive", status="inactive"))
        self.assertEqual(inactive.status_code, 200)
        inactive_id = inactive.json()["id"]

        first = client.post(f"/rules/{inactive_id}/deactivate")
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.json()["message"], "Rule is already inactive")
        self.assertEqual(client.post("/rules/9999/deactivate").status_code, 404)
        self.assertIn(client.put(f"/rules/{inactive_id}", json={"name": "changed"}).status_code, {404, 405})

    def test_rule_scheduling_validation_and_api_payload(self):
        client = TestClient(app.app)

        valid = valid_rule(
            entry_window_start_seconds_before_end=120,
            entry_window_end_seconds_before_end=10,
            schedule_timezone="Asia/Jerusalem",
            inactive_windows=[
                {"day_of_week": 0, "start_time": "22:00", "end_time": "02:00", "status": "active"},
                {"day_of_week": 5, "start_time": "02:00:00", "end_time": "05:00:00", "status": "inactive"},
            ],
        )
        response = client.post("/rules", json=valid)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["entry_window_start_seconds_before_end"], 120)
        self.assertEqual(payload["entry_window_end_seconds_before_end"], 10)
        self.assertEqual(payload["schedule_timezone"], "Asia/Jerusalem")
        self.assertEqual(len(payload["inactive_windows"]), 2)
        self.assertEqual(payload["inactive_windows"][0]["start_time"], "22:00:00")

        invalid_cases = [
            valid_rule(entry_window_start_seconds_before_end=120),
            valid_rule(entry_window_start_seconds_before_end=5, entry_window_end_seconds_before_end=10),
            valid_rule(schedule_timezone="UTC+3"),
            valid_rule(inactive_windows=[{"day_of_week": 7, "start_time": "02:00", "end_time": "05:00"}]),
        ]
        for payload in invalid_cases:
            self.assertEqual(client.post("/rules", json=payload).status_code, 400)

    def test_entry_window_blocks_and_stores_seconds_before_event_end(self):
        app.upsert_event(market_row("event-a"))
        app.create_rule(valid_rule(
            entry_window_start_seconds_before_end=120,
            entry_window_end_seconds_before_end=10,
        ))

        app.insert_orderbook_log(orderbook(
            "event-a",
            yes_ask=0.77,
            yes_bid=0.76,
            no_ask=0.2,
            no_bid=0.19,
            sampled_at="2026-07-17T09:02:00+00:00",
        ))
        self.assertEqual(len(self.fetch_deals()), 0)

        app.insert_orderbook_log(orderbook(
            "event-a",
            yes_ask=0.77,
            yes_bid=0.76,
            no_ask=0.2,
            no_bid=0.19,
            sampled_at="2026-07-17T09:03:37+00:00",
        ))
        deals = self.fetch_deals()
        self.assertEqual(len(deals), 1)
        self.assertEqual(deals[0]["entry_seconds_before_event_end"], 83)

    def test_inactive_window_crossing_midnight_blocks_only_new_entries(self):
        app.create_rule(valid_rule(
            inactive_windows=[
                {"day_of_week": 0, "start_time": "22:00", "end_time": "02:00", "status": "active"},
            ],
        ))

        app.insert_orderbook_log(orderbook(
            "event-a",
            yes_ask=0.77,
            yes_bid=0.76,
            no_ask=0.2,
            no_bid=0.19,
            sampled_at="2026-07-20T20:30:00+00:00",
        ))
        app.insert_orderbook_log(orderbook(
            "event-a",
            yes_ask=0.77,
            yes_bid=0.76,
            no_ask=0.2,
            no_bid=0.19,
            sampled_at="2026-07-20T22:30:00+00:00",
        ))
        self.assertEqual(len(self.fetch_deals()), 0)

        app.insert_orderbook_log(orderbook(
            "event-a",
            yes_ask=0.77,
            yes_bid=0.76,
            no_ask=0.2,
            no_bid=0.19,
            sampled_at="2026-07-21T00:30:00+00:00",
        ))
        self.assertEqual(len(self.fetch_deals()), 1)

    def test_inactive_schedule_gate_boundaries_days_status_and_refresh(self):
        rule = app.create_rule(valid_rule(
            inactive_windows=[
                {"day_of_week": 0, "start_time": "10:00", "end_time": "12:00", "status": "active"},
                {"day_of_week": 1, "start_time": "22:00", "end_time": "02:00", "status": "active"},
                {"day_of_week": 3, "start_time": "00:00", "end_time": "00:00", "status": "active"},
                {"day_of_week": 4, "start_time": "10:00", "end_time": "12:00", "status": "inactive"},
            ],
        ))
        cases = [
            ("2026-07-20T06:59:00+00:00", True, "before"),
            ("2026-07-20T07:00:00+00:00", False, "start included"),
            ("2026-07-20T08:59:59+00:00", False, "inside"),
            ("2026-07-20T09:00:00+00:00", True, "end excluded"),
            ("2026-07-21T19:30:00+00:00", False, "cross midnight before"),
            ("2026-07-21T22:30:00+00:00", False, "cross midnight after"),
            ("2026-07-21T23:00:00+00:00", True, "cross midnight end"),
            ("2026-07-23T09:00:00+00:00", False, "all day"),
            ("2026-07-24T07:30:00+00:00", True, "disabled record ignored"),
        ]
        with app.get_conn() as conn:
            for current_time, expected, label in cases:
                with self.subTest(label=label):
                    allowed, loaded_rule, reason, _window = app.can_rule_open_new_deal(
                        conn, rule["id"], current_time
                    )
                    self.assertEqual(allowed, expected)
                    self.assertIsNotNone(loaded_rule)
                    self.assertEqual(reason, "" if expected else "rule_in_inactive_schedule")

            conn.execute(
                "UPDATE rule_inactive_windows SET start_time = '13:00:00', end_time = '14:00:00' "
                "WHERE rule_id = ? AND day_of_week = 0",
                (rule["id"],),
            )
            conn.commit()
            allowed, _rule, _reason, _window = app.can_rule_open_new_deal(
                conn, rule["id"], "2026-07-20T07:30:00+00:00"
            )
            self.assertTrue(allowed)

    def test_final_gate_reloads_schedule_before_insert(self):
        rule = app.create_rule(valid_rule())
        original = app.can_rule_open_new_deal
        calls = 0

        def add_window_then_gate(conn, rule_id, current_time):
            nonlocal calls
            calls += 1
            if calls == 1:
                now = app.now_iso()
                conn.execute(
                    "INSERT INTO rule_inactive_windows "
                    "(rule_id, day_of_week, start_time, end_time, status, created_at, updated_at) "
                    "VALUES (?, 4, '00:00:00', '00:00:00', 'active', ?, ?)",
                    (rule_id, now, now),
                )
            return original(conn, rule_id, current_time)

        with patch.object(app, "can_rule_open_new_deal", side_effect=add_window_then_gate):
            app.insert_orderbook_log(orderbook(
                "event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19,
                sampled_at="2026-07-17T09:00:01+00:00",
            ))
        self.assertEqual(calls, 1)
        self.assertEqual(len(self.fetch_deals()), 0)
        self.assertEqual(rule["status"], "active")

    def test_active_rule_starts_only_on_next_event_and_survives_restart(self):
        app.active_market = {"event_slug": "event-a", "condition_id": "condition-a"}
        rule = app.create_rule(valid_rule())
        self.assertEqual(rule["eligible_after_event_id"], "event-a")

        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 0)

        app.active_market = None
        app.insert_orderbook_log(orderbook("event-b", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        deals = self.fetch_deals()
        self.assertEqual(len(deals), 1)
        self.assertEqual(deals[0]["side"], "yes")

    def test_entry_rules_exact_prices_open_deal_limits_and_duplicates(self):
        app.create_rule(valid_rule(max_yes_entries_per_event=1, max_no_entries_per_event=1))

        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.76, yes_bid=0.75, no_ask=0.2, no_bid=0.19))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.78, yes_bid=0.75, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 0)

        first_sample = orderbook("event-a", yes_ask="0.770000", yes_bid=0.76, no_ask=0.2, no_bid=0.19)
        log_id = app.insert_orderbook_log(first_sample)
        with app.get_conn() as conn:
            app.process_demo_trading_for_orderbook(conn, first_sample, log_id)
            conn.commit()
        self.assertEqual(len(self.fetch_deals()), 1)

        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 1)

    def test_deal_entry_stores_latest_btc_volume_snapshot(self):
        app.create_rule(valid_rule())
        app.insert_btc_volume_log(btc_volume("2026-07-17T08:59:30+00:00", cumulative=10.0, delta=5.5))
        app.insert_btc_volume_log(btc_volume("2026-07-17T09:00:00+00:00", cumulative=16.0, delta=6.8))
        app.insert_btc_volume_log(btc_volume("2026-07-17T09:00:30+00:00", cumulative=20.0, delta=3.2))

        app.insert_orderbook_log(orderbook(
            "event-a",
            yes_ask=0.77,
            yes_bid=0.76,
            no_ask=0.2,
            no_bid=0.19,
            sampled_at="2026-07-17T09:00:01+00:00",
        ))

        deal = self.fetch_deals()[0]
        self.assertIsNotNone(deal["entry_btc_volume_log_id"])
        self.assertEqual(deal["entry_btc_volume_sampled_at"], "2026-07-17T09:00:00+00:00")
        self.assertAlmostEqual(deal["entry_btc_volume_btc_cumulative"], 16.0)
        self.assertAlmostEqual(deal["entry_btc_volume_btc_delta"], 6.8)
        self.assertEqual(deal["entry_btc_volume_status"], "success")

        overview = app.load_dashboard_overview(1)
        self.assertEqual(overview["btc_volume_gt_6_deals"], 1)
        self.assertEqual(overview["missing_btc_volume_snapshot_deals"], 0)

        snapshot = app.load_btc_volume_deal_snapshot()
        self.assertEqual(snapshot["deals_over_6_delta"], 1)
        self.assertEqual(snapshot["missing_snapshot_deals"], 0)
        self.assertEqual(snapshot["rows_over_6"][0]["id"], deal["id"])

        trends = app.load_btc_volume_trends()
        self.assertEqual(len(trends), 3)
        self.assertEqual(trends[-1]["label"], "17/07 12:00")
        self.assertAlmostEqual(trends[1]["volume_btc_delta"], 6.8)

    def test_backfill_script_updates_missing_btc_volume_snapshot(self):
        app.insert_btc_volume_log(btc_volume("2026-07-17T09:00:00+00:00", cumulative=16.0, delta=6.8))
        with app.get_conn() as conn:
            now = app.now_iso()
            cursor = conn.execute("""
                INSERT INTO rules (
                    name, created_at, updated_at, entry_price, stop_loss_price,
                    take_profit_price, max_yes_entries_per_event,
                    max_no_entries_per_event, status, eligible_after_event_id
                ) VALUES ('legacy-missing-volume', ?, ?, 0.77, 0.60, 0.90, 1, 1, 'inactive', NULL)
            """, (now, now))
            rule_id = cursor.lastrowid
            conn.execute("""
                INSERT INTO deals (
                    rule_id, rule_name, event_id, side, result, entry_at, entry_price,
                    entry_orderbook_log_id, created_at, updated_at
                ) VALUES (?, 'legacy-missing-volume', 'event-a', 'yes', 'open',
                    '2026-07-17T09:00:01+00:00', 0.77, 1, ?, ?)
            """, (rule_id, now, now))
            conn.commit()

        with patch.object(sys, "argv", [
            "backfill_deal_btc_volume_snapshots.py",
            "--apply",
            "--db-path",
            str(app.DB_PATH),
        ]):
            self.assertEqual(backfill_deal_btc_volume_snapshots.main(), 0)

        deal = self.fetch_deals()[0]
        self.assertEqual(deal["entry_btc_volume_sampled_at"], "2026-07-17T09:00:00+00:00")
        self.assertAlmostEqual(deal["entry_btc_volume_btc_delta"], 6.8)

    def test_no_entry_for_inactive_open_deal_and_both_sides_match(self):
        app.create_rule(valid_rule(name="inactive", status="inactive"))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 0)

        app.create_rule(valid_rule())
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.77, no_bid=0.76))
        self.assertEqual(len(self.fetch_deals()), 0)

    def test_two_rules_can_open_on_same_sample_and_side_quotas_reset_per_event(self):
        app.create_rule(valid_rule(name="rule 1", max_yes_entries_per_event=1, max_no_entries_per_event=1))
        app.create_rule(valid_rule(name="rule 2", max_yes_entries_per_event=1, max_no_entries_per_event=1))

        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 2)

        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.2, yes_bid=0.91, no_ask=0.2, no_bid=0.19))
        app.insert_orderbook_log(orderbook("event-a", no_ask=0.77, no_bid=0.76, yes_ask=0.2, yes_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 4)

        app.insert_orderbook_log(orderbook("event-a", no_ask=0.2, no_bid=0.91, yes_ask=0.2, yes_bid=0.19))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 4)

        app.insert_orderbook_log(orderbook("event-b", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        self.assertEqual(len(self.fetch_deals()), 6)

    def test_take_profit_and_stop_loss_use_same_side_bid_and_target_exit_price(self):
        app.create_rule(valid_rule())
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.99))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.2, yes_bid=0.93, no_ask=0.2, no_bid=0.1))
        deal = self.fetch_deals()[0]
        self.assertEqual(deal["result"], "win")
        self.assertEqual(deal["exit_reason"], "take_profit")
        self.assertEqual(deal["exit_price"], 0.9)
        self.assertIsNotNone(deal["exit_orderbook_log_id"])
        self.assertAlmostEqual(deal["price_change_points"], 13.0)

        app.create_rule(valid_rule(name="no side", entry_price="0.74", stop_loss_price="0.65", take_profit_price="0.95"))
        app.insert_orderbook_log(orderbook("event-b", no_ask=0.74, no_bid=0.73, yes_ask=0.2, yes_bid=0.99))
        app.insert_orderbook_log(orderbook("event-b", no_ask=0.2, no_bid=0.57, yes_ask=0.2, yes_bid=0.99))
        no_deal = self.fetch_deals()[-1]
        self.assertEqual(no_deal["side"], "no")
        self.assertEqual(no_deal["result"], "loss")
        self.assertEqual(no_deal["exit_reason"], "stop_loss")
        self.assertEqual(no_deal["exit_price"], 0.65)
        self.assertAlmostEqual(no_deal["price_change_points"], 9.0)

    def test_stop_loss_wins_if_both_thresholds_are_seen(self):
        with app.get_conn() as conn:
            now = app.now_iso()
            cursor = conn.execute("""
                INSERT INTO rules (
                    name, created_at, updated_at, entry_price, stop_loss_price,
                    take_profit_price, max_yes_entries_per_event,
                    max_no_entries_per_event, status, eligible_after_event_id
                ) VALUES ('legacy-invalid', ?, ?, 0.77, 0.80, 0.70, 1, 1, 'active', NULL)
            """, (now, now))
            rule_id = cursor.lastrowid
            conn.execute("""
                INSERT INTO deals (
                    rule_id, rule_name, event_id, side, result, entry_at, entry_price,
                    entry_orderbook_log_id, created_at, updated_at
                ) VALUES (?, ?, 'event-a', 'yes', 'open', ?, 0.77, 1, ?, ?)
            """, (rule_id, "legacy-invalid", now, now, now))
            conn.commit()

        app.insert_orderbook_log(orderbook("event-a", yes_bid=0.75, yes_ask=0.2, no_bid=0.1, no_ask=0.2))
        deal = self.fetch_deals()[0]
        self.assertEqual(deal["result"], "loss")
        self.assertEqual(deal["exit_reason"], "stop_loss")
        self.assertEqual(deal["exit_price"], 0.8)

    def test_event_resolution_closes_open_deals_once(self):
        app.create_rule(valid_rule())
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        market_row = {
            "polymarket_event_id": "poly-event",
            "polymarket_market_id": "poly-market",
            "condition_id": "condition-a",
            "event_slug": "event-a",
            "market_slug": "event-a",
            "title": "title",
            "question": "question",
            "event_url": "url",
            "start_time": "2026-07-17T09:00:00Z",
            "start_time_local": "17/07/2026 12:00:00",
            "end_time": "2026-07-17T09:05:00Z",
            "end_time_local": "17/07/2026 12:05:00",
            "yes_token_id": "yes-token",
            "no_token_id": "no-token",
            "outcomes": '["Up", "Down"]',
            "outcome_prices": '["1", "0"]',
            "active": 0,
            "closed": 1,
            "enable_order_book": 1,
            "accepting_orders": 0,
            "created_at_poly": "2026-07-17T08:59:00Z",
            "created_at_poly_local": "17/07/2026 11:59:00",
            "status": "closed",
            "notes": "",
            "raw_json": "{}",
        }
        app.upsert_event(market_row)
        app.upsert_event(market_row)
        deal = self.fetch_deals()[0]
        self.assertEqual(deal["result"], "win")
        self.assertEqual(deal["exit_reason"], "event_resolution")
        self.assertEqual(deal["exit_price"], 1)
        self.assertEqual(deal["market_result"], "yes")
        self.assertIsNone(deal["exit_orderbook_log_id"])

    def test_deactivate_rule_keeps_open_deal_closing_but_blocks_new_entries(self):
        rule = app.create_rule(valid_rule(max_yes_entries_per_event=2))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        app.deactivate_rule(rule["id"])
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.2, yes_bid=0.93, no_ask=0.2, no_bid=0.19))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        deals = self.fetch_deals()
        self.assertEqual(len(deals), 1)
        self.assertEqual(deals[0]["result"], "win")

    def test_deal_metric_examples(self):
        points, percent = app.calculate_deal_metrics(0.77, 0.95)
        self.assertEqual(points, 18.0)
        self.assertAlmostEqual(percent, ((0.95 - 0.77) / 0.77) * 100)

        points, percent = app.calculate_deal_metrics(0.74, 0.65)
        self.assertEqual(points, 9.0)
        self.assertAlmostEqual(percent, ((0.65 - 0.74) / 0.74) * 100)

        points, percent = app.calculate_deal_metrics(0.77, 1)
        self.assertEqual(points, 23.0)
        self.assertAlmostEqual(percent, ((1 - 0.77) / 0.77) * 100)

        points, percent = app.calculate_deal_metrics(0.77, 0)
        self.assertEqual(points, 77.0)
        self.assertAlmostEqual(percent, ((0 - 0.77) / 0.77) * 100)

        pnl, roi, shares = app.calculate_deal_pnl_usd(0.77, 0.90, 1)
        self.assertAlmostEqual(shares, 1 / 0.77)
        self.assertAlmostEqual(pnl, (1 / 0.77) * 0.90 - 1)
        self.assertAlmostEqual(roi, ((0.90 - 0.77) / 0.77) * 100)

        pnl, roi, shares = app.calculate_deal_pnl_usd(0.74, 0.65, 1)
        self.assertAlmostEqual(shares, 1 / 0.74)
        self.assertAlmostEqual(pnl, (1 / 0.74) * 0.65 - 1)
        self.assertAlmostEqual(roi, ((0.65 - 0.74) / 0.74) * 100)

    def test_demo_fee_calculation_policy(self):
        financials = app.calculate_demo_deal_financials(0.77, 0.90)
        self.assertEqual(financials["entry_liquidity_role"], "TAKER")
        self.assertEqual(financials["exit_liquidity_role"], "TAKER")
        self.assertEqual(float(financials["entry_fee_rate"]), 0.07)
        self.assertAlmostEqual(float(financials["entry_fee_usd"]), 0.01610)
        self.assertAlmostEqual(float(financials["exit_fee_usd"]), 0.00818)
        self.assertAlmostEqual(float(financials["gross_pnl_usd"]), (1 / 0.77) * 0.90 - 1)
        self.assertAlmostEqual(float(financials["net_pnl_usd"]), 0.14455117)

        maker = app.calculate_demo_deal_financials(0.77, 0.90, entry_liquidity_role="MAKER", exit_liquidity_role="MAKER")
        self.assertEqual(float(maker["total_fees_usd"]), 0.0)
        self.assertAlmostEqual(float(maker["net_pnl_usd"]), float(maker["gross_pnl_usd"]))

        no_fee = app.calculate_demo_deal_financials(0.77, 0.90, fee_rate=0)
        self.assertEqual(float(no_fee["total_fees_usd"]), 0.0)

    def test_gamma_fee_disabled_does_not_zero_demo_crypto_fees(self):
        with app.get_conn() as conn:
            now = app.now_iso()
            conn.execute("""
                INSERT INTO events (
                    event_slug, fees_enabled, fee_rate, fee_calculation_source,
                    fee_calculation_version, discovered_at, last_seen_at
                ) VALUES ('event-fee-disabled', 0, 0, 'MARKET_SNAPSHOT', ?, ?, ?)
            """, (app.DEMO_FEE_CALCULATION_VERSION, now, now))
            conn.commit()

        app.create_rule(valid_rule())
        app.insert_orderbook_log(orderbook("event-fee-disabled", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        app.insert_orderbook_log(orderbook("event-fee-disabled", yes_ask=0.2, yes_bid=0.93, no_ask=0.2, no_bid=0.19))

        deal = self.fetch_deals()[0]
        self.assertEqual(deal["entry_fee_rate"], 0.07)
        self.assertGreater(deal["total_fees_usd"], 0)
        overview = app.load_dashboard_overview(1)
        self.assertGreater(overview["total_fees_usd"], 0)

    def test_fee_backfill_updates_closed_zero_fee_deals(self):
        with app.get_conn() as conn:
            now = app.now_iso()
            conn.execute("""
                INSERT INTO rules (
                    name, created_at, updated_at, entry_price, stop_loss_price,
                    take_profit_price, max_yes_entries_per_event,
                    max_no_entries_per_event, status, eligible_after_event_id
                ) VALUES ('legacy-zero-fee', ?, ?, 0.77, 0.60, 0.90, 1, 1, 'inactive', NULL)
            """, (now, now))
            rule_id = conn.execute("SELECT id FROM rules WHERE name = 'legacy-zero-fee'").fetchone()["id"]
            conn.execute("""
                INSERT INTO deals (
                    rule_id, rule_name, event_id, side, result, entry_at, entry_price,
                    entry_orderbook_log_id, exit_at, exit_price, exit_reason,
                    entry_fee_rate, entry_fee_usd, exit_fee_rate, exit_fee_usd,
                    total_fees_usd, gross_pnl_usd, net_pnl_usd,
                    fee_calculation_source, created_at, updated_at
                ) VALUES (?, 'legacy-zero-fee', 'event-a', 'yes', 'win', ?, 0.77,
                    1, ?, 0.90, 'take_profit', 0, 0, 0, 0, 0, NULL, NULL,
                    'MARKET_SNAPSHOT', ?, ?)
            """, (rule_id, now, now, now, now))
            updated = app.backfill_demo_fee_snapshots(conn)
            conn.commit()

        self.assertEqual(updated, 1)
        deal = self.fetch_deals()[0]
        self.assertEqual(deal["fee_calculation_source"], app.DEMO_FEE_SOURCE_BACKFILL)
        self.assertGreater(deal["total_fees_usd"], 0)
        self.assertAlmostEqual(deal["net_pnl_usd"], expected_net_pnl(0.77, 0.90))

    def test_excel_and_dashboard_include_rules_and_deals(self):
        active = app.create_rule(valid_rule(name="active"))
        app.create_rule(valid_rule(name="inactive", status="inactive"))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.77, yes_bid=0.76, no_ask=0.2, no_bid=0.19))
        app.insert_orderbook_log(orderbook("event-a", yes_ask=0.2, yes_bid=0.93, no_ask=0.2, no_bid=0.19))
        app.deactivate_rule(active["id"])

        overview = app.load_dashboard_overview(1)
        self.assertEqual(overview["closed_deals"], 1)
        self.assertEqual(overview["open_deals"], 0)
        self.assertEqual(overview["wins"], 1)
        self.assertAlmostEqual(overview["gross_pnl_usd"], (1 / 0.77) * 0.90 - 1)
        self.assertAlmostEqual(overview["net_pnl_usd"], expected_net_pnl(0.77, 0.90))
        self.assertGreater(overview["total_fees_usd"], 0)
        self.assertEqual(overview["range"], "all")
        self.assertEqual(app.normalize_dashboard_range("bad-value"), "all")
        self.assertEqual(app.normalize_dashboard_range("custom"), "custom")

        custom_overview = app.load_dashboard_overview(1, "custom", "17/07/2026 11:00", "17/07/2026 13:00")
        self.assertEqual(custom_overview["closed_deals"], 1)
        self.assertEqual(custom_overview["custom_from"], "17/07/2026 11:00")
        self.assertEqual(custom_overview["custom_to"], "17/07/2026 13:00")

        excluded_overview = app.load_dashboard_overview(1, "custom", "18/07/2026 00:00", "18/07/2026 01:00")
        self.assertEqual(excluded_overview["closed_deals"], 0)

        rules_performance = app.load_rules_performance(1)
        self.assertEqual(len(rules_performance), 2)
        self.assertEqual(rules_performance[0]["rule_name"], "active")
        self.assertEqual(rules_performance[0]["closed_deals"], 1)
        self.assertEqual(rules_performance[0]["wins"], 1)
        self.assertAlmostEqual(rules_performance[0]["net_pnl_usd"], expected_net_pnl(0.77, 0.90))
        self.assertEqual(rules_performance[1]["rule_name"], "inactive")
        self.assertEqual(rules_performance[1]["closed_deals"], 0)

        risk = app.load_risk_snapshot(1)
        self.assertEqual(risk["closed_deals"], 1)
        self.assertAlmostEqual(risk["ending_equity_usd"], expected_net_pnl(0.77, 0.90))
        self.assertAlmostEqual(risk["max_drawdown_usd"], 0)
        self.assertEqual(risk["best_deal"]["deal_id"], 1)
        self.assertEqual(risk["worst_deal"]["deal_id"], 1)
        self.assertEqual(risk["exit_reasons"][0]["exit_reason"], "take_profit")

        conditions = app.load_market_conditions(1)
        self.assertEqual(conditions["by_side"][0]["label"], "YES")
        self.assertEqual(conditions["by_side"][0]["closed_deals"], 1)
        self.assertAlmostEqual(conditions["by_side"][0]["net_pnl_usd"], expected_net_pnl(0.77, 0.90))
        self.assertEqual(conditions["by_entry_price"][0]["label"], "0.70-0.79")
        self.assertEqual(conditions["by_entry_price"][0]["closed_deals"], 1)

        health = app.load_system_health_snapshot()
        self.assertEqual(health["status"], "ok")
        self.assertEqual(health["rules_count"], 2)
        self.assertEqual(health["deals_count"], 1)
        self.assertEqual(health["open_deals"], 0)
        self.assertEqual(health["orderbook_count"], 2)

        trends = app.load_time_trends(1)
        self.assertEqual(len(trends), 1)
        self.assertEqual(trends[0]["closed_deals"], 1)
        self.assertEqual(trends[0]["wins"], 1)
        self.assertAlmostEqual(trends[0]["net_pnl_usd"], expected_net_pnl(0.77, 0.90))

        data_quality = app.load_data_quality_snapshot()
        self.assertIn(data_quality["status"], {"ok", "needs_review"})
        self.assertEqual(len(data_quality["checks"]), 10)

        export_path, row_counts = app.write_xlsx_export()
        self.assertEqual(row_counts["rules"], 2)
        self.assertEqual(row_counts["deals"], 1)
        workbook = load_workbook(export_path, read_only=True)
        try:
            self.assertEqual(workbook.sheetnames, [
                "events",
                "orderbook_log",
                "btc_volume_log",
                "rules",
                "deals",
                "rule_inactive_windows",
                "fee_summary",
            ])
        finally:
            workbook.close()

        client = TestClient(app.app)
        dashboard = client.get("/")
        self.assertEqual(dashboard.status_code, 200)
        html = dashboard.text
        self.assertIn("<h2>סקירה ניהולית</h2>", html)
        self.assertIn("<h2>מגמות לאורך זמן</h2>", html)
        self.assertIn("<h2>איכות נתונים</h2>", html)
        self.assertIn('class="nav-link active"', html)
        self.assertIn("/rules-page", html)
        self.assertIn("/deals-page", html)
        self.assertIn("/market-data", html)
        self.assertIn("/system-page", html)
        self.assertIn("טווח תאריכים", html)
        self.assertIn("DD/MM/YYYY HH:mm", html)
        self.assertIn("custom_from", html)
        self.assertIn("custom_to", html)
        self.assertIn("רווח/הפסד נטו", html)
        self.assertIn("ROI ממוצע", html)
        self.assertIn("$0.17", html)
        self.assertIn("<h2>Interactive charts</h2>", html)
        self.assertNotIn('http-equiv="refresh"', html)
        self.assertIn("refreshDashboardContent", html)
        self.assertIn("renderDashboardCharts", html)
        self.assertIn("https://cdn.jsdelivr.net/npm/chart.js", html)
        self.assertIn('"timeTrendLabels"', html)
        self.assertNotIn("&quot;timeTrendLabels&quot;", html)

        dashboard_content = client.get("/dashboard-content")
        self.assertEqual(dashboard_content.status_code, 200)
        self.assertIn("<h2>סקירה ניהולית</h2>", dashboard_content.text)
        self.assertIn("<h2>מגמות לאורך זמן</h2>", dashboard_content.text)
        self.assertIn("<h2>Interactive charts</h2>", dashboard_content.text)
        self.assertIn("<h2>איכות נתונים</h2>", dashboard_content.text)

        rules_page = client.get("/rules-page")
        self.assertEqual(rules_page.status_code, 200)
        self.assertIn("<h2>ביצועי חוקים</h2>", rules_page.text)
        self.assertIn("<h2>Rules</h2>", rules_page.text)
        self.assertIn("Create Rule", rules_page.text)
        self.assertIn('data-content-path="/rules-page-content"', rules_page.text)

        deals_page = client.get("/deals-page")
        self.assertEqual(deals_page.status_code, 200)
        self.assertIn("<h2>תמונת סיכון</h2>", deals_page.text)
        self.assertIn("<h2>תנאי שוק</h2>", deals_page.text)
        self.assertIn("0.70-0.79", deals_page.text)
        self.assertIn("סיבות יציאה", deals_page.text)
        self.assertIn("<h2>BTC volume on deals</h2>", deals_page.text)
        self.assertIn("<h2>Deals</h2>", deals_page.text)
        self.assertIn('data-content-path="/deals-page-content"', deals_page.text)

        market_page = client.get("/market-data")
        self.assertEqual(market_page.status_code, 200)
        self.assertIn("<h2>Events / Markets</h2>", market_page.text)
        self.assertIn("<h2>Coinbase BTC Volume</h2>", market_page.text)
        self.assertIn("<h2>Orderbook Log</h2>", market_page.text)
        self.assertIn("Orderbook Log", market_page.text)
        self.assertIn('data-content-path="/market-data-content"', market_page.text)

        system_page = client.get("/system-page")
        self.assertEqual(system_page.status_code, 200)
        self.assertIn("<h2>איכות נתונים</h2>", system_page.text)
        self.assertIn("<h2>בריאות מערכת</h2>", system_page.text)
        self.assertIn("Generate Excel", system_page.text)
        self.assertIn('data-content-path="/system-page-content"', system_page.text)

        filtered_content = client.get("/dashboard-content?investment_usd=2&range_filter=all")
        self.assertEqual(filtered_content.status_code, 200)
        self.assertIn("טווח: כל התקופה", filtered_content.text)

        custom_content = client.get(
            "/dashboard-content?investment_usd=2&range_filter=custom&custom_from=17/07/2026%2011:00&custom_to=17/07/2026%2013:00"
        )
        self.assertEqual(custom_content.status_code, 200)
        self.assertIn("17/07/2026 11:00", custom_content.text)
        self.assertIn("17/07/2026 13:00", custom_content.text)

        app.export_state.update({"status": "ready", "filename": export_path.name, "path": str(export_path)})
        download = client.get("/download.xlsx")
        self.assertEqual(download.status_code, 200)
        downloaded = load_workbook(BytesIO(download.content), read_only=True)
        try:
            self.assertEqual(downloaded.sheetnames, [
                "events",
                "orderbook_log",
                "btc_volume_log",
                "rules",
                "deals",
                "rule_inactive_windows",
                "fee_summary",
            ])
        finally:
            downloaded.close()


if __name__ == "__main__":
    unittest.main()
